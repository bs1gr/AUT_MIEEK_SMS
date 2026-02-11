"""
Custom report generation service for Phase 6 Reporting Enhancements.

Handles report data retrieval and export to PDF/Excel/CSV.
Uses a background-task-friendly entrypoint that opens its own DB session.
"""

from __future__ import annotations

import csv
import logging
import os
import time
from datetime import date, datetime, timezone
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session, joinedload
from xml.sax.saxutils import escape

try:
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_CENTER
    from reportlab.lib.pagesizes import landscape, letter
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import inch
    from reportlab.platypus import Flowable, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False
    Flowable = Any  # type: ignore

from backend.config import settings
from backend.models import Attendance, Course, DailyPerformance, GeneratedReport, Grade, Report, Student, User
from backend.services.email_notification_service import EmailNotificationService
from backend.services.report_exporters import register_report_fonts

logger = logging.getLogger(__name__)


class CustomReportGenerationService:
    """Generate report files for custom report definitions."""

    FIELD_LABELS = {
        "first_name": {"en": "First Name", "el": "Όνομα"},
        "last_name": {"en": "Last Name", "el": "Επώνυμο"},
        "student_name": {"en": "Student Name", "el": "Όνομα Μαθητή"},
        "student_id": {"en": "Student ID", "el": "Αριθμός Μητρώου"},
        "email": {"en": "Email", "el": "Email"},
        "is_active": {"en": "Active", "el": "Ενεργός"},
        "course_code": {"en": "Course Code", "el": "Κωδικός Μαθήματος"},
        "course_name": {"en": "Course Name", "el": "Όνομα Μαθήματος"},
        "course_title": {"en": "Course Title", "el": "Τίτλος Μαθήματος"},
        "semester": {"en": "Semester", "el": "Εξάμηνο"},
        "credits": {"en": "Credits", "el": "Μονάδες"},
        "assignment_name": {"en": "Assignment Name", "el": "Όνομα Εργασίας"},
        "category": {"en": "Category", "el": "Κατηγορία"},
        "grade": {"en": "Grade", "el": "Βαθμός"},
        "max_grade": {"en": "Max Grade", "el": "Μέγιστος Βαθμός"},
        "weight": {"en": "Weight", "el": "Βαρύτητα"},
        "date_assigned": {"en": "Date Assigned", "el": "Ημερομηνία Ανάθεσης"},
        "date_submitted": {"en": "Date Submitted", "el": "Ημερομηνία Υποβολής"},
        "date": {"en": "Date", "el": "Ημερομηνία"},
        "status": {"en": "Status", "el": "Κατάσταση"},
        "score": {"en": "Score", "el": "Βαθμολογία"},
        "max_score": {"en": "Max Score", "el": "Μέγιστη Βαθμολογία"},
        "percentage": {"en": "Percentage", "el": "Ποσοστό"},
        "study_year": {"en": "Study Year", "el": "Έτος Σπουδών"},
        "academic_year": {"en": "Academic Year", "el": "Ακαδημαϊκό Έτος"},
        "class_division": {"en": "Class Division", "el": "Τμήμα"},
    }

    # Translation dictionary for data values (categories, statuses, etc.)
    VALUE_TRANSLATIONS = {
        # Assignment categories (case-insensitive matching)
        "exam": {"en": "Exam", "el": "Εξέταση"},
        "midterm": {"en": "Midterm", "el": "Ενδιάμεση"},
        "midterm exam": {"en": "Midterm Exam", "el": "Ενδιάμεση Εξέταση"},
        "final": {"en": "Final", "el": "Τελική"},
        "final exam": {"en": "Final Exam", "el": "Τελική Εξέταση"},
        "quiz": {"en": "Quiz", "el": "Κουίζ"},
        "test": {"en": "Test", "el": "Τεστ"},
        "assignment": {"en": "Assignment", "el": "Εργασία"},
        "homework": {"en": "Homework", "el": "Κατ' οίκον"},
        "homework/assignments": {"en": "Homework/Assignments", "el": "Εργασίες/Ασκήσεις"},
        "project": {"en": "Project", "el": "Έργο"},
        "lab": {"en": "Lab", "el": "Εργαστήριο"},
        "laboratory": {"en": "Laboratory", "el": "Εργαστήριο"},
        "participation": {"en": "Participation", "el": "Συμμετοχή"},
        "class participation": {"en": "Class Participation", "el": "Συμμετοχή στο Μάθημα"},
        "continuous assessment": {"en": "Continuous Assessment", "el": "Συνεχής Αξιολόγηση"},
        "presentation": {"en": "Presentation", "el": "Παρουσίαση"},
        "essay": {"en": "Essay", "el": "Δοκίμιο"},
        "report": {"en": "Report", "el": "Αναφορά"},
        "paper": {"en": "Paper", "el": "Εργασία"},
        "practical": {"en": "Practical", "el": "Πρακτική"},
        "oral": {"en": "Oral", "el": "Προφορική"},
        "written": {"en": "Written", "el": "Γραπτή"},
        # Attendance statuses
        "present": {"en": "Present", "el": "Παρών"},
        "absent": {"en": "Absent", "el": "Απών"},
        "late": {"en": "Late", "el": "Καθυστέρηση"},
        "excused": {"en": "Excused", "el": "Δικαιολογημένος"},
        "unexcused": {"en": "Unexcused", "el": "Αδικαιολόγητος"},
        "tardy": {"en": "Tardy", "el": "Καθυστέρηση"},
        # Common assignment name patterns
        "sample exam": {"en": "Sample Exam", "el": "Δοκιμαστική Εξέταση"},
        "practice exam": {"en": "Practice Exam", "el": "Πρακτική Εξέταση"},
        "mock exam": {"en": "Mock Exam", "el": "Δοκιμαστική Εξέταση"},
        "review": {"en": "Review", "el": "Επανάληψη"},
        "exercise": {"en": "Exercise", "el": "Άσκηση"},
        "worksheet": {"en": "Worksheet", "el": "Φύλλο Εργασίας"},
        # Semester/academic terms
        "fall": {"en": "Fall", "el": "Χειμερινό"},
        "spring": {"en": "Spring", "el": "Εαρινό"},
        "summer": {"en": "Summer", "el": "Θερινό"},
        "winter": {"en": "Winter", "el": "Χειμερινό"},
        "semester 1": {"en": "Semester 1", "el": "Εξάμηνο 1"},
        "semester 2": {"en": "Semester 2", "el": "Εξάμηνο 2"},
        # Academic status
        "active": {"en": "Active", "el": "Ενεργός"},
        "inactive": {"en": "Inactive", "el": "Ανενεργός"},
        "enrolled": {"en": "Enrolled", "el": "Εγγεγραμμένος"},
        "graduated": {"en": "Graduated", "el": "Απόφοιτος"},
        "withdrawn": {"en": "Withdrawn", "el": "Αποχώρησε"},
        "suspended": {"en": "Suspended", "el": "Αναστολή"},
        # Common grade/score terms
        "pass": {"en": "Pass", "el": "Επιτυχία"},
        "fail": {"en": "Fail", "el": "Αποτυχία"},
        "incomplete": {"en": "Incomplete", "el": "Ατελές"},
        "pending": {"en": "Pending", "el": "Εκκρεμεί"},
        "submitted": {"en": "Submitted", "el": "Υποβλήθηκε"},
        "graded": {"en": "Graded", "el": "Βαθμολογήθηκε"},
        "missing": {"en": "Missing", "el": "Λείπει"},
        # Course types
        "lecture": {"en": "Lecture", "el": "Διάλεξη"},
        "tutorial": {"en": "Tutorial", "el": "Φροντιστήριο"},
        "seminar": {"en": "Seminar", "el": "Σεμινάριο"},
        "workshop": {"en": "Workshop", "el": "Εργαστήριο"},
    }

    def __init__(self, db: Session):
        self.db = db
        self.reports_dir = os.path.join(os.path.dirname(__file__), "..", "reports")
        os.makedirs(self.reports_dir, exist_ok=True)
        self._allow_sensitive_fields = False
        self._language = "en"

    @staticmethod
    def _normalize_language(value: Optional[str]) -> str:
        if not value:
            return "en"
        normalized = value.strip().lower()
        if normalized.startswith("el"):
            return "el"
        return "en"

    def _set_language(self, value: Optional[str]) -> None:
        self._language = self._normalize_language(value)

    def _translate_value(self, value: Any, field_name: str = "") -> Any:
        """
        Translate data values like category names, assignment names, status values.
        Word-by-word translation with support for multi-word phrases.
        """
        if not isinstance(value, str) or not value:
            return value

        # Only translate specific fields
        translatable_fields = {"category", "status", "assignment_name"}
        if field_name.lower() not in translatable_fields:
            return value

        # Keep English if language is English
        if self._language == "en":
            return value

        value_lower = value.strip().lower()

        # Check for exact full match first
        if value_lower in self.VALUE_TRANSLATIONS:
            translation = self.VALUE_TRANSLATIONS[value_lower].get(self._language)
            if translation:
                return translation

        # For assignment_name and category, do word-by-word translation
        # Sort translation keys by length (descending) to match multi-word phrases first
        sorted_keys = sorted(self.VALUE_TRANSLATIONS.keys(), key=len, reverse=True)

        result = value
        result_lower = value_lower

        # Track what's been replaced to avoid double-translation
        already_replaced_positions: set[tuple[int, str]] = set()

        for key in sorted_keys:
            if key in result_lower:
                # Find all occurrences of this key
                start = 0
                while True:
                    pos = result_lower.find(key, start)
                    if pos == -1:
                        break

                    # Check if this position overlaps with already replaced text
                    if any(pos <= p < pos + len(key) or p <= pos < p + len(k) for p, k in already_replaced_positions):
                        start = pos + 1
                        continue

                    # Check word boundaries (don't replace partial words)
                    before_ok = pos == 0 or not result_lower[pos - 1].isalnum()
                    after_ok = pos + len(key) >= len(result_lower) or not result_lower[pos + len(key)].isalnum()

                    if before_ok and after_ok:
                        # Get the translation
                        translated = self.VALUE_TRANSLATIONS[key].get(self._language, key)

                        # Preserve capitalization style from original
                        original_fragment = result[pos : pos + len(key)]
                        if original_fragment.isupper():
                            # All caps → keep translation as-is
                            pass
                        elif original_fragment[0].isupper():
                            # Title case → capitalize first letter
                            translated = (
                                translated[0].upper() + translated[1:] if len(translated) > 1 else translated.upper()
                            )

                        # Replace in both result and result_lower
                        result = result[:pos] + translated + result[pos + len(key) :]
                        result_lower = result_lower[:pos] + translated.lower() + result_lower[pos + len(key) :]

                        # Mark this position as replaced
                        already_replaced_positions.add((pos, key))

                        # Move start forward
                        start = pos + len(translated)
                    else:
                        start = pos + 1

        return result

    @staticmethod
    def _normalize_field_key(field: str) -> str:
        return (field or "").strip().lower()

    @classmethod
    def _is_sensitive_field(cls, field: str) -> bool:
        sensitive = {
            "health_issue",
            "note",
            "mobile_phone",
            "phone",
            "email",
        }
        normalized = cls._normalize_field_key(field)
        if normalized in sensitive:
            return True
        if "." in normalized:
            suffix = normalized.split(".")[-1]
            return suffix in sensitive
        return False

    def _set_sensitive_field_access(self, user_id: int) -> None:
        user = self.db.query(User).filter(User.id == user_id).first()
        self._allow_sensitive_fields = bool(user and getattr(user, "role", None) == "admin")

    @staticmethod
    def run_generation_task(
        report_id: int,
        generated_report_id: int,
        user_id: int,
        export_format: str,
        include_charts: bool,
        email_recipients: Optional[List[str]] = None,
        email_enabled: Optional[bool] = None,
        language: Optional[str] = None,
    ) -> None:
        """Background task entrypoint with isolated DB session."""
        from backend.db import SessionLocal

        db = SessionLocal()
        try:
            service = CustomReportGenerationService(db)
            service.generate_report(
                report_id,
                generated_report_id,
                user_id,
                export_format,
                include_charts,
                email_recipients=email_recipients,
                email_enabled=email_enabled,
                language=language,
            )
        finally:
            db.close()

    def generate_report(
        self,
        report_id: int,
        generated_report_id: int,
        user_id: int,
        export_format: str,
        include_charts: bool,
        email_recipients: Optional[List[str]] = None,
        email_enabled: Optional[bool] = None,
        language: Optional[str] = None,
    ) -> None:
        """Generate a report file and update the GeneratedReport record."""
        start_time = time.perf_counter()
        generated = self._get_generated_report(report_id, generated_report_id, user_id)
        if not generated:
            logger.error("Generated report record not found: %s", generated_report_id)
            return

        report = self._get_report(report_id, user_id)
        if not report:
            self._mark_failed(generated, "Report not found or access denied")
            return

        self._set_sensitive_field_access(user_id)
        self._set_language(language)

        self._update_status(generated, "generating")

        try:
            rows, headers = self._build_report_rows(report)
            file_path = self._export_report(
                rows,
                headers,
                export_format,
                str(generated.file_name),
                title=str(report.name) if report.name else None,
            )  # type: ignore[arg-type]
            duration = time.perf_counter() - start_time

            generated.file_path = file_path  # type: ignore[assignment]
            generated.file_size_bytes = os.path.getsize(file_path) if file_path else None  # type: ignore[assignment]
            generated.record_count = len(rows)  # type: ignore[assignment]
            generated.generation_duration_seconds = round(float(duration), 3)  # type: ignore[arg-type,assignment]
            generated.status = "completed"  # type: ignore[assignment]
            generated.error_message = None  # type: ignore[assignment]

            report.last_run_at = datetime.now(timezone.utc)  # type: ignore[assignment]
            self.db.commit()

            self._send_report_email(
                report=report,
                generated=generated,
                export_format=export_format,
                email_recipients=email_recipients,
                email_enabled=email_enabled,
            )
        except Exception as exc:
            logger.error("Report generation failed: %s", exc, exc_info=True)
            self._mark_failed(generated, str(exc))

    def _send_report_email(
        self,
        report: Report,
        generated: GeneratedReport,
        export_format: str,
        email_recipients: Optional[List[str]] = None,
        email_enabled: Optional[bool] = None,
    ) -> None:
        recipients = email_recipients if email_recipients is not None else report.email_recipients
        if isinstance(recipients, str):
            recipients = [recipients]
        if not recipients:
            return

        send_enabled = email_enabled if email_enabled is not None else bool(report.email_enabled)
        if not send_enabled:
            return

        if not EmailNotificationService.is_enabled():
            logger.warning("Email delivery skipped: SMTP not configured")
            return

        file_path = str(generated.file_path or "")
        attachment_paths: list[str] = []
        attachment_note: Optional[str] = None
        max_mb = int(getattr(settings, "SMTP_ATTACHMENT_MAX_MB", 10) or 10)
        max_bytes = max_mb * 1024 * 1024

        if file_path and os.path.exists(file_path):
            try:
                file_size = os.path.getsize(file_path)
                if file_size <= max_bytes:
                    attachment_paths = [file_path]
                else:
                    attachment_note = (
                        f"Attachment not included because the file exceeds {max_mb} MB. "
                        "You can download it from the reports page."
                    )
            except Exception as exc:
                logger.warning("Failed to prepare attachment: %s", exc)
        else:
            attachment_note = "Attachment not found on disk. You can download it from the reports page."

        generated_at = generated.generated_at.isoformat() if generated.generated_at else ""
        success_count, failed = EmailNotificationService.send_report_ready(
            recipients=recipients,
            report_name=str(report.name),
            export_format=export_format,
            record_count=int(generated.record_count) if generated.record_count is not None else None,
            generated_at=generated_at,
            attachment_paths=attachment_paths or None,
            attachment_note=attachment_note,
        )

        if failed:
            generated.email_sent = False  # type: ignore[assignment]
            generated.email_error = f"Failed to send to: {', '.join(failed)}"  # type: ignore[assignment]
        else:
            generated.email_sent = True  # type: ignore[assignment]
            generated.email_error = None  # type: ignore[assignment]
        if success_count > 0:
            generated.email_sent_at = datetime.now(timezone.utc)  # type: ignore[assignment]
        self.db.commit()

    def _get_report(self, report_id: int, user_id: int) -> Optional[Report]:
        return (
            self.db.query(Report)
            .filter(Report.id == report_id, Report.user_id == user_id, Report.deleted_at.is_(None))
            .first()
        )

    def _get_generated_report(self, report_id: int, generated_id: int, user_id: int) -> Optional[GeneratedReport]:
        return (
            self.db.query(GeneratedReport)
            .filter(
                GeneratedReport.id == generated_id,
                GeneratedReport.report_id == report_id,
                GeneratedReport.user_id == user_id,
            )
            .first()
        )

    def _update_status(self, generated: GeneratedReport, status: str) -> None:
        generated.status = status  # type: ignore[assignment]
        self.db.commit()

    def _mark_failed(self, generated: GeneratedReport, message: str) -> None:
        generated.status = "failed"  # type: ignore[assignment]
        generated.error_message = message  # type: ignore[assignment]
        self.db.commit()

    def _build_report_rows(self, report: Report) -> Tuple[List[List[Any]], List[str]]:
        model, query = self._build_query(report)
        # Handle filters as a list of filter objects or dict
        filters_list: list = report.filters or []  # type: ignore[assignment]
        sort_list: list = report.sort_by or []  # type: ignore[assignment]

        query = self._apply_filters(query, model, filters_list)
        query = self._apply_sort(query, model, sort_list)

        columns = self._normalize_columns(str(report.report_type), report.fields)  # type: ignore[arg-type]
        if not self._allow_sensitive_fields:
            columns = [(key, label) for key, label in columns if not self._is_sensitive_field(key)]
        headers = [label for _, label in columns]

        rows: List[List[Any]] = []
        for record in query.all():
            row = [self._resolve_field(record, key) for key, _ in columns]
            rows.append(row)

        return rows, headers

    def _build_query(self, report: Report):
        report_type = str(report.report_type or "").lower()

        if report_type == "student":
            query = self.db.query(Student).filter(Student.deleted_at.is_(None))
            return Student, query

        if report_type == "course":
            query = self.db.query(Course).filter(Course.deleted_at.is_(None))
            return Course, query

        if report_type == "grade":
            query = (
                self.db.query(Grade)
                .options(joinedload(Grade.student), joinedload(Grade.course))
                .filter(Grade.deleted_at.is_(None))
            )
            return Grade, query

        if report_type == "attendance":
            query = (
                self.db.query(Attendance)
                .options(joinedload(Attendance.student), joinedload(Attendance.course))
                .filter(Attendance.deleted_at.is_(None))
            )
            return Attendance, query

        if report_type == "daily_performance":
            query = (
                self.db.query(DailyPerformance)
                .options(joinedload(DailyPerformance.student), joinedload(DailyPerformance.course))
                .filter(DailyPerformance.deleted_at.is_(None))
            )
            return DailyPerformance, query

        raise ValueError(f"Unsupported report type: {report.report_type}")

    def _apply_filters(self, query, model, filters: Any) -> Any:  # type: ignore[no-untyped-def]
        """Apply filters to the query. Handles multiple filter formats."""
        if not filters:
            return query

        # Convert list of filter objects to proper queries
        if isinstance(filters, list):
            for filter_obj in filters:
                if not isinstance(filter_obj, dict):
                    continue

                field = filter_obj.get("field")
                operator = filter_obj.get("operator", "equals")
                value = filter_obj.get("value")

                if not field or not hasattr(model, field):
                    continue

                column = getattr(model, field)

                # Apply operator-specific logic
                if operator == "equals":
                    query = query.filter(column == value)  # type: ignore[operator]
                elif operator == "not_equals":
                    query = query.filter(column != value)  # type: ignore[operator]
                elif operator == "contains":
                    query = query.filter(column.ilike(f"%{value}%"))  # type: ignore[operator]
                elif operator == "not_contains":
                    query = query.filter(~column.ilike(f"%{value}%"))  # type: ignore[operator]
                elif operator == "starts_with":
                    query = query.filter(column.ilike(f"{value}%"))  # type: ignore[operator]
                elif operator == "ends_with":
                    query = query.filter(column.ilike(f"%{value}"))  # type: ignore[operator]
                elif operator == "greater_than" or operator == "gt":
                    query = query.filter(column > value)  # type: ignore[operator]
                elif operator == "less_than" or operator == "lt":
                    query = query.filter(column < value)  # type: ignore[operator]
                elif operator == "greater_than_or_equal":
                    query = query.filter(column >= value)  # type: ignore[operator]
                elif operator == "less_than_or_equal":
                    query = query.filter(column <= value)  # type: ignore[operator]
                elif operator == "in":
                    if isinstance(value, (list, tuple)):
                        query = query.filter(column.in_(value))  # type: ignore[operator]
                elif operator == "between":
                    if isinstance(value, dict) and "from" in value and "to" in value:
                        query = query.filter(column.between(value["from"], value["to"]))  # type: ignore[operator]

        # Nested dict format: {"grade": {"operator": "less_than", "value": 60}}
        elif isinstance(filters, dict):
            for field_name, filter_spec in filters.items():
                # If value is a dict with operator/value, it's the new format
                if isinstance(filter_spec, dict) and "operator" in filter_spec and "value" in filter_spec:
                    operator = filter_spec.get("operator", "equals")
                    value = filter_spec.get("value")

                    if not hasattr(model, field_name):
                        continue

                    column = getattr(model, field_name)

                    # Apply operator-specific logic
                    if operator == "equals":
                        query = query.filter(column == value)  # type: ignore[operator]
                    elif operator == "not_equals":
                        query = query.filter(column != value)  # type: ignore[operator]
                    elif operator == "contains":
                        query = query.filter(column.ilike(f"%{value}%"))  # type: ignore[operator]
                    elif operator == "not_contains":
                        query = query.filter(~column.ilike(f"%{value}%"))  # type: ignore[operator]
                    elif operator == "starts_with":
                        query = query.filter(column.ilike(f"{value}%"))  # type: ignore[operator]
                    elif operator == "ends_with":
                        query = query.filter(column.ilike(f"%{value}"))  # type: ignore[operator]
                    elif operator == "greater_than" or operator == "gt":
                        query = query.filter(column > value)  # type: ignore[operator]
                    elif operator == "less_than" or operator == "lt":
                        query = query.filter(column < value)  # type: ignore[operator]
                    elif operator == "greater_than_or_equal":
                        query = query.filter(column >= value)  # type: ignore[operator]
                    elif operator == "less_than_or_equal":
                        query = query.filter(column <= value)  # type: ignore[operator]
                    elif operator == "in":
                        if isinstance(value, (list, tuple)):
                            query = query.filter(column.in_(value))  # type: ignore[operator]
                    elif operator == "between":
                        if isinstance(value, dict) and "from" in value and "to" in value:
                            query = query.filter(column.between(value["from"], value["to"]))  # type: ignore[operator]
                # Legacy simple format: {"field_name": value}
                else:
                    if hasattr(model, field_name):
                        column = getattr(model, field_name)
                        query = query.filter(column == filter_spec)  # type: ignore[operator]

        return query

    def _apply_sort(self, query, model, sort_by: Any) -> Any:  # type: ignore[no-untyped-def]
        """Apply sorting to the query. Handles both list and dict formats."""
        # Handle list of sort objects
        if isinstance(sort_by, list):
            for sort_obj in sort_by:
                if not isinstance(sort_obj, dict):
                    continue

                field = sort_obj.get("field")
                order = str(sort_obj.get("order", "asc")).lower()

                if field and hasattr(model, field):
                    column = getattr(model, field)
                    query = query.order_by(column.desc() if order == "desc" else column.asc())  # type: ignore[operator]

        # Legacy dict format support
        elif isinstance(sort_by, dict):
            field = sort_by.get("field") if isinstance(sort_by, dict) else None
            direction = str(sort_by.get("direction", "asc")).lower() if isinstance(sort_by, dict) else "asc"
            if field and hasattr(model, field):
                column = getattr(model, field)
                query = query.order_by(column.desc() if direction == "desc" else column.asc())  # type: ignore[operator]

        return query

    def _normalize_columns(self, report_type: str, fields: Any) -> List[Tuple[str, str]]:  # type: ignore[no-untyped-def]
        columns = self._extract_columns(fields)
        if columns:
            return columns

        defaults: Dict[str, List[str]] = {
            "student": ["first_name", "last_name", "email", "student_id", "is_active"],
            "course": ["course_code", "course_name", "semester", "credits"],
            "grade": ["student_name", "course_code", "grade", "max_grade", "date_assigned"],
            "attendance": ["student_name", "course_code", "date", "status"],
            "daily_performance": [
                "student_name",
                "course_code",
                "date",
                "category",
                "score",
                "max_score",
                "percentage",
            ],
        }
        fallback = defaults.get(report_type.lower(), [])
        return [(key, self._label_from_key(key, self._language)) for key in fallback]

    def _extract_columns(self, fields: Any) -> List[Tuple[str, str]]:
        columns: Sequence[Any] = []
        if isinstance(fields, dict):
            columns = fields.get("columns") or fields.get("fields") or []
        elif isinstance(fields, list):
            columns = fields

        result: List[Tuple[str, str]] = []
        for col in columns:
            if isinstance(col, dict):
                key = col.get("key") or col.get("field") or col.get("name")
                if not key:
                    continue
                label = col.get("label") or ""
                resolved_label = self._localize_label(str(key), str(label))
                result.append((str(key), resolved_label))
            else:
                key = str(col)
                result.append((key, self._label_from_key(key, self._language)))
        return result

    def _label_from_key(self, key: str, language: Optional[str] = None) -> str:
        normalized = self._normalize_field_key(key)
        labels = self.FIELD_LABELS.get(normalized)
        default_label = key.replace("_", " ").replace(".", " ").title()
        if not labels:
            return default_label

        target_language = self._normalize_language(language)
        if target_language == "el":
            return labels.get("el") or labels.get("en") or default_label
        return labels.get("en") or default_label

    def _localize_label(self, key: str, label: str) -> str:
        if not label:
            return self._label_from_key(key, self._language)

        if self._normalize_language(self._language) != "el":
            return label

        english_default = self._label_from_key(key, "en")
        normalized_label = label.strip().lower()
        if normalized_label == english_default.strip().lower():
            return self._label_from_key(key, "el")
        return label

    def _resolve_field(self, record: Any, field: str) -> Any:
        """Resolve a field value from a record, handling relationships and nested fields."""

        if not self._allow_sensitive_fields and self._is_sensitive_field(field):
            return ""

        if field == "percentage":
            value = None
            if hasattr(record, "score") and hasattr(record, "max_score"):
                score = getattr(record, "score", None)
                max_score = getattr(record, "max_score", None)
                if score is not None and max_score:
                    try:
                        value = (float(score) / float(max_score)) * 100
                    except Exception:
                        value = None
            elif hasattr(record, "grade") and hasattr(record, "max_grade"):
                grade = getattr(record, "grade", None)
                max_grade = getattr(record, "max_grade", None)
                if grade is not None and max_grade:
                    try:
                        value = (float(grade) / float(max_grade)) * 100
                    except Exception:
                        value = None
            elif hasattr(record, "percentage"):
                try:
                    value = float(getattr(record, "percentage"))
                except Exception:
                    value = None

            if value is None:
                return ""
            return f"{value:.2f}%"

        # Special handling for computed fields
        if field == "student_name":
            student = getattr(record, "student", None)
            if student:
                return f"{student.first_name} {student.last_name}".strip()
            if hasattr(record, "first_name") and hasattr(record, "last_name"):
                return f"{record.first_name} {record.last_name}".strip()
            return ""

        if field == "course_code":
            course = getattr(record, "course", None)
            if course and hasattr(course, "course_code"):
                return course.course_code
            if hasattr(record, "course_code"):
                return record.course_code
            return ""

        if field == "course_name":
            course = getattr(record, "course", None)
            if course and hasattr(course, "course_name"):
                return course.course_name
            if hasattr(record, "course_name"):
                return record.course_name
            return ""

        # Handle nested dot notation
        if "." in field:
            value = record
            for part in field.split("."):
                value = getattr(value, part, None)
                if value is None:
                    return ""
            return value

        # Try to get directly from record
        if hasattr(record, field):
            value = getattr(record, field, None)
            if value is None:
                return ""
            if isinstance(value, (datetime, date)):
                return value.isoformat()
            # Translate value if applicable (category, status, assignment_name)
            return self._translate_value(value, field)

        # For Grade records, try to resolve from related models
        # e.g., first_name -> student.first_name, last_name -> student.last_name
        if hasattr(record, "student"):
            student = getattr(record, "student", None)
            if student and hasattr(student, field):
                value = getattr(student, field, None)
                if value is None:
                    return ""
                if isinstance(value, (datetime, date)):
                    return value.isoformat()
                return self._translate_value(value, field)

        # Try Course relationship
        if hasattr(record, "course"):
            course = getattr(record, "course", None)
            if course and hasattr(course, field):
                value = getattr(course, field, None)
                if value is None:
                    return ""
                if isinstance(value, (datetime, date)):
                    return value.isoformat()
                return self._translate_value(value, field)

        return ""

    def _export_report(
        self,
        rows: List[List[Any]],
        headers: List[str],
        export_format: str,
        file_name: str,
        title: Optional[str] = None,
    ) -> str:  # type: ignore[override]
        format_lower = (export_format or "pdf").lower()
        file_path = os.path.join(self.reports_dir, file_name)

        if format_lower == "csv":
            self._export_csv(rows, headers, file_path)
            return file_path

        if format_lower == "excel":
            self._export_excel(rows, headers, file_path)
            return file_path

        if format_lower == "pdf":
            self._export_pdf(rows, headers, file_path, title=title)
            return file_path

        raise ValueError(f"Unsupported export format: {export_format}")

    def _export_csv(self, rows: Iterable[Sequence[Any]], headers: Sequence[str], file_path: str) -> None:
        with open(file_path, "w", newline="", encoding="utf-8") as handle:
            writer = csv.writer(handle)
            writer.writerow(headers)
            for row in rows:
                writer.writerow([self._stringify_value(value) for value in row])

    def _export_excel(self, rows: Iterable[Sequence[Any]], headers: Sequence[str], file_path: str) -> None:
        wb = Workbook()
        ws = wb.active
        ws.title = "Report"

        header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        for col_index, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_index, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        for row_index, row in enumerate(rows, 2):
            for col_index, value in enumerate(row, 1):
                ws.cell(row=row_index, column=col_index, value=self._stringify_value(value))

        for col_index, header in enumerate(headers, 1):
            width = max(len(str(header)), 10)
            ws.column_dimensions[get_column_letter(col_index)].width = max(width, 12)

        wb.save(file_path)

    def _export_pdf(
        self, rows: List[List[Any]], headers: List[str], file_path: str, title: Optional[str] = None
    ) -> None:
        if not REPORTLAB_AVAILABLE:
            raise ImportError("ReportLab is required for PDF generation. Install with: pip install reportlab")

        page_size = letter
        if len(headers) > 8:
            page_size = landscape(letter)

        doc = SimpleDocTemplate(
            file_path,
            pagesize=page_size,
            leftMargin=0.5 * inch,
            rightMargin=0.5 * inch,
            topMargin=0.5 * inch,
            bottomMargin=0.5 * inch,
        )
        styles = getSampleStyleSheet()
        base_font, base_font_bold = register_report_fonts()
        body_font_size = 10
        if len(headers) > 6:
            body_font_size = 9
        if len(headers) > 8:
            body_font_size = 8
        if len(headers) > 10:
            body_font_size = 7

        title_style = ParagraphStyle(
            "ReportTitle",
            parent=styles["Heading1"],
            fontSize=16,
            textColor=colors.HexColor("#1f77b4"),
            spaceAfter=12,
            fontName=base_font_bold,
            alignment=TA_CENTER,
        )
        header_style = ParagraphStyle(
            "ReportHeader",
            parent=styles["BodyText"],
            fontName=base_font_bold,
            fontSize=body_font_size,
            leading=body_font_size + 2,
            textColor=colors.whitesmoke,
            alignment=TA_CENTER,
        )
        cell_style = ParagraphStyle(
            "ReportCell",
            parent=styles["BodyText"],
            fontName=base_font,
            fontSize=body_font_size,
            leading=body_font_size + 2,
        )

        report_title = title
        if not report_title:
            report_title = "Προσαρμοσμένη Αναφορά" if self._language == "el" else "Custom Report"
        elements: List[Flowable] = [Paragraph(report_title, title_style), Spacer(1, 0.2 * inch)]

        safe_headers: List[Flowable] = [Paragraph(escape(str(header)), header_style) for header in headers]
        safe_rows: List[List[Flowable]] = []
        for row in rows:
            safe_row: List[Flowable] = [
                Paragraph(escape(str(self._stringify_value(value))), cell_style) for value in row
            ]
            safe_rows.append(safe_row)

        table_data: List[List[Flowable]] = [safe_headers] + safe_rows
        col_count = max(len(headers), 1)
        col_width = doc.width / col_count
        col_widths = [col_width for _ in range(col_count)]

        table = Table(table_data, repeatRows=1, colWidths=col_widths)
        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1f77b4")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                    ("ALIGN", (0, 0), (-1, 0), "CENTER"),
                    ("ALIGN", (0, 1), (-1, -1), "LEFT"),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("FONTNAME", (0, 0), (-1, 0), base_font_bold),
                    ("FONTNAME", (0, 1), (-1, -1), base_font),
                    ("FONTSIZE", (0, 0), (-1, 0), body_font_size),
                    ("FONTSIZE", (0, 1), (-1, -1), body_font_size),
                    ("BOTTOMPADDING", (0, 0), (-1, 0), 8),
                    ("BOTTOMPADDING", (0, 1), (-1, -1), 6),
                    ("TOPPADDING", (0, 1), (-1, -1), 6),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F3F4F6")]),
                ]
            )
        )
        elements.append(table)
        doc.build(elements)

    def _stringify_value(self, value: Any) -> Any:
        if isinstance(value, (datetime, date)):
            return value.isoformat()
        if value is None:
            return ""
        if isinstance(value, bool):
            if self._language == "el":
                return "Ναι" if value else "Όχι"
            return "Yes" if value else "No"
        return value
