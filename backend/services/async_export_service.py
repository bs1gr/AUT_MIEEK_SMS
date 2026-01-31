"""
Async export service for background Excel/CSV/PDF generation.

Provides task queue management for large exports to improve response times.
Uses in-memory queue pattern with background processing.
Supports multiple export formats: Excel (XLSX), CSV, and PDF.
"""

import logging
import os
from datetime import datetime, timezone
from typing import Optional, Dict, Any
from enum import Enum
import csv
import io

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

try:
    from reportlab.lib.pagesizes import letter
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False

from backend.models import Student, Course, Grade, ExportJob

logger = logging.getLogger(__name__)


class ExportStatus(str, Enum):
    """Export job status states."""

    PENDING = "pending"
    PROCESSING = "processing"
    COMPLETED = "completed"
    FAILED = "failed"
    CANCELLED = "cancelled"


class AsyncExportService:
    """Service for managing async export tasks.

    Handles background generation of Excel/PDF files with status tracking.
    Uses background task queue pattern for non-blocking exports.
    """

    def __init__(self):
        self.exports_dir = os.path.join(os.path.dirname(__file__), "..", "exports")
        os.makedirs(self.exports_dir, exist_ok=True)

    def get_export_path(self, export_id: int, format: str) -> str:
        """Generate file path for export."""
        ext = "xlsx" if format == "excel" else format
        return os.path.join(self.exports_dir, f"export_{export_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.{ext}")

    def generate_students_excel(
        self, db: Session, export_job_id: int, filters: Optional[Dict[str, Any]] = None, limit: int = 10000
    ) -> tuple[bool, Optional[str], str]:
        """Generate Excel export of students.

        Returns: (success, file_path, error_message)
        """
        try:
            query = db.query(Student).filter(Student.deleted_at.is_(None))

            # Apply filters if provided
            if filters:
                if filters.get("is_active") is not None:
                    query = query.filter(Student.is_active == filters.get("is_active"))
                if filters.get("status"):
                    query = query.filter(Student.status == filters.get("status"))

            # Apply limit and fetch
            students = query.order_by(Student.last_name, Student.first_name).limit(limit).all()

            # Generate workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Students"

            # Headers
            headers = ["ID", "First Name", "Last Name", "Email", "Student ID", "Enrollment Date", "Status"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")

            # Data rows with progress tracking
            total_students = len(students)
            for row, student in enumerate(students, 2):
                ws.cell(row=row, column=1, value=student.id)
                ws.cell(row=row, column=2, value=student.first_name)
                ws.cell(row=row, column=3, value=student.last_name)
                ws.cell(row=row, column=4, value=student.email)
                ws.cell(row=row, column=5, value=student.student_id)
                ws.cell(row=row, column=6, value=str(student.enrollment_date) if student.enrollment_date else "")
                ws.cell(row=row, column=7, value="Active" if student.is_active else "Inactive")

                # Update progress every 10% or every 100 records (whichever is smaller)
                current_index = row - 1  # row starts at 2, so index is row-1
                update_interval = min(max(total_students // 10, 1), 100)
                if current_index % update_interval == 0 or current_index == total_students:
                    progress = int((current_index / total_students) * 100)
                    export_job = db.query(ExportJob).filter(ExportJob.id == export_job_id).first()
                    if export_job:
                        export_job.progress_percent = progress
                        db.commit()

            # Adjust column widths
            for col in range(1, len(headers) + 1):
                ws.column_dimensions[get_column_letter(col)].width = 18

            # Save file
            file_path = self.get_export_path(export_job_id, "excel")
            wb.save(file_path)

            # Update export job
            export_job = db.query(ExportJob).filter(ExportJob.id == export_job_id).first()
            if export_job:
                export_job.status = ExportStatus.COMPLETED
                export_job.file_path = file_path
                export_job.total_records = len(students)
                export_job.completed_at = datetime.now(timezone.utc)
                db.commit()

            logger.info(f"Export {export_job_id} completed: {len(students)} students")
            return True, file_path, ""

        except Exception as e:
            logger.error(f"Export {export_job_id} failed: {str(e)}", exc_info=True)

            # Update export job status
            export_job = db.query(ExportJob).filter(ExportJob.id == export_job_id).first()
            if export_job:
                export_job.status = ExportStatus.FAILED
                export_job.completed_at = datetime.now(timezone.utc)
                db.commit()

            return False, None, str(e)

    def generate_courses_excel(
        self, db: Session, export_job_id: int, filters: Optional[Dict[str, Any]] = None, limit: int = 10000
    ) -> tuple[bool, Optional[str], str]:
        """Generate Excel export of courses."""
        try:
            query = db.query(Course).filter(Course.deleted_at.is_(None))

            if filters:
                if filters.get("is_active") is not None:
                    query = query.filter(Course.is_active == filters.get("is_active"))

            courses = query.order_by(Course.code).limit(limit).all()

            # Generate workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Courses"

            headers = ["Code", "Name", "Description", "Instructor", "Credits", "Status"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="059669", end_color="059669", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")

            for row, course in enumerate(courses, 2):
                ws.cell(row=row, column=1, value=course.code)
                ws.cell(row=row, column=2, value=course.name)
                ws.cell(row=row, column=3, value=course.description or "")
                ws.cell(row=row, column=4, value=course.instructor or "")
                ws.cell(row=row, column=5, value=course.credits or 0)
                ws.cell(row=row, column=6, value="Active" if course.is_active else "Inactive")

                # Update progress every 10% or every 100 records
                current_index = row - 1
                total_courses = len(courses)
                update_interval = min(max(total_courses // 10, 1), 100)
                if current_index % update_interval == 0 or current_index == total_courses:
                    progress = int((current_index / total_courses) * 100)
                    export_job = db.query(ExportJob).filter(ExportJob.id == export_job_id).first()
                    if export_job:
                        export_job.progress_percent = progress
                        db.commit()

            for col in range(1, len(headers) + 1):
                ws.column_dimensions[get_column_letter(col)].width = 18

            file_path = self.get_export_path(export_job_id, "excel")
            wb.save(file_path)

            export_job = db.query(ExportJob).filter(ExportJob.id == export_job_id).first()
            if export_job:
                export_job.status = ExportStatus.COMPLETED
                export_job.file_path = file_path
                export_job.total_records = len(courses)
                export_job.completed_at = datetime.now(timezone.utc)
                db.commit()

            logger.info(f"Export {export_job_id} completed: {len(courses)} courses")
            return True, file_path, ""

        except Exception as e:
            logger.error(f"Export {export_job_id} failed: {str(e)}", exc_info=True)

            export_job = db.query(ExportJob).filter(ExportJob.id == export_job_id).first()
            if export_job:
                export_job.status = ExportStatus.FAILED
                export_job.completed_at = datetime.now(timezone.utc)
                db.commit()

            return False, None, str(e)

    def generate_grades_excel(
        self, db: Session, export_job_id: int, filters: Optional[Dict[str, Any]] = None, limit: int = 10000
    ) -> tuple[bool, Optional[str], str]:
        """Generate Excel export of grades."""
        try:
            query = db.query(Grade).filter(Grade.deleted_at.is_(None))

            if filters:
                if filters.get("student_id"):
                    query = query.filter(Grade.student_id == filters.get("student_id"))
                if filters.get("course_id"):
                    query = query.filter(Grade.course_id == filters.get("course_id"))

            grades = query.order_by(Grade.created_at.desc()).limit(limit).all()

            wb = Workbook()
            ws = wb.active
            ws.title = "Grades"

            headers = ["Student ID", "Course Code", "Grade", "Points", "Recorded Date"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="7C3AED", end_color="7C3AED", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")

            for row, grade in enumerate(grades, 2):
                ws.cell(row=row, column=1, value=grade.student_id if grade.student else "")
                ws.cell(row=row, column=2, value=grade.course.code if grade.course else "")
                ws.cell(row=row, column=3, value=grade.grade)
                ws.cell(row=row, column=4, value=grade.points)
                ws.cell(row=row, column=5, value=str(grade.created_at) if grade.created_at else "")

                # Update progress every 10% or every 100 records
                current_index = row - 1
                total_grades = len(grades)
                update_interval = min(max(total_grades // 10, 1), 100)
                if current_index % update_interval == 0 or current_index == total_grades:
                    progress = int((current_index / total_grades) * 100)
                    export_job = db.query(ExportJob).filter(ExportJob.id == export_job_id).first()
                    if export_job:
                        export_job.progress_percent = progress
                        db.commit()

            for col in range(1, len(headers) + 1):
                ws.column_dimensions[get_column_letter(col)].width = 18

            file_path = self.get_export_path(export_job_id, "excel")
            wb.save(file_path)

            export_job = db.query(ExportJob).filter(ExportJob.id == export_job_id).first()
            if export_job:
                export_job.status = ExportStatus.COMPLETED
                export_job.file_path = file_path
                export_job.total_records = len(grades)
                export_job.completed_at = datetime.now(timezone.utc)
                db.commit()

            logger.info(f"Export {export_job_id} completed: {len(grades)} grades")
            return True, file_path, ""

        except Exception as e:
            logger.error(f"Export {export_job_id} failed: {str(e)}", exc_info=True)

            export_job = db.query(ExportJob).filter(ExportJob.id == export_job_id).first()
            if export_job:
                export_job.status = ExportStatus.FAILED
                export_job.completed_at = datetime.now(timezone.utc)
                db.commit()

            return False, None, str(e)

    # ===== CSV EXPORT METHODS =====

    def generate_students_csv(
        self, db: Session, export_job_id: int, filters: Optional[Dict[str, Any]] = None, limit: int = 10000
    ) -> tuple[bool, Optional[str], str]:
        """Generate CSV export of students."""
        try:
            query = db.query(Student)
            if filters:
                if filters.get("status"):
                    query = query.filter(Student.status == filters["status"])
                if filters.get("search"):
                    search_term = f"%{filters['search']}%"
                    query = query.filter(
                        (Student.first_name.ilike(search_term)) | (Student.last_name.ilike(search_term)) | (Student.email.ilike(search_term))
                    )

            students = query.limit(limit).all()

            # Write CSV file
            file_path = self.get_export_path(export_job_id, "csv")
            with open(file_path, "w", newline="", encoding="utf-8") as f:
                writer = csv.writer(f)
                writer.writerow(["Student ID", "First Name", "Last Name", "Email", "Status", "Created At"])
                for student in students:
                    writer.writerow([
                        student.id,
                        student.first_name,
                        student.last_name,
                        student.email,
                        student.status,
                        student.created_at.isoformat() if student.created_at else ""
                    ])

            # Update export job
            export_job = db.query(ExportJob).filter(ExportJob.id == export_job_id).first()
            if export_job:
                export_job.status = ExportStatus.COMPLETED
                export_job.file_path = file_path
                export_job.total_records = len(students)
                export_job.progress_percent = 100
                export_job.completed_at = datetime.now(timezone.utc)
                db.commit()

            return True, file_path, "CSV export completed"
        except Exception as e:
            logger.error(f"CSV export failed: {e}")
            export_job = db.query(ExportJob).filter(ExportJob.id == export_job_id).first()
            if export_job:
                export_job.status = ExportStatus.FAILED
                export_job.error_message = str(e)
                db.commit()
            return False, None, str(e)

    def generate_courses_csv(
        self, db: Session, export_job_id: int, filters: Optional[Dict[str, Any]] = None, limit: int = 10000
    ) -> tuple[bool, Optional[str], str]:
        """Generate CSV export of courses."""
        try:
            query = db.query(Course)
            if filters:
                if filters.get("status"):
                    query = query.filter(Course.status == filters["status"])
                if filters.get("search"):
                    search_term = f"%{filters['search']}%"
                    query = query.filter(
                        (Course.course_code.ilike(search_term)) | (Course.course_name.ilike(search_term))
                    )

            courses = query.limit(limit).all()

            # Write CSV file
            file_path = self.get_export_path(export_job_id, "csv")
            with open(file_path, "w", newline="", encoding="utf-8") as f:
                writer = csv.writer(f)
                writer.writerow(["Course Code", "Course Name", "Instructor", "Status", "Created At"])
                for course in courses:
                    writer.writerow([
                        course.course_code,
                        course.course_name,
                        course.instructor_name or "",
                        course.status,
                        course.created_at.isoformat() if course.created_at else ""
                    ])

            # Update export job
            export_job = db.query(ExportJob).filter(ExportJob.id == export_job_id).first()
            if export_job:
                export_job.status = ExportStatus.COMPLETED
                export_job.file_path = file_path
                export_job.total_records = len(courses)
                export_job.progress_percent = 100
                export_job.completed_at = datetime.now(timezone.utc)
                db.commit()

            return True, file_path, "CSV export completed"
        except Exception as e:
            logger.error(f"CSV export failed: {e}")
            export_job = db.query(ExportJob).filter(ExportJob.id == export_job_id).first()
            if export_job:
                export_job.status = ExportStatus.FAILED
                export_job.error_message = str(e)
                db.commit()
            return False, None, str(e)

    def generate_grades_csv(
        self, db: Session, export_job_id: int, filters: Optional[Dict[str, Any]] = None, limit: int = 10000
    ) -> tuple[bool, Optional[str], str]:
        """Generate CSV export of grades."""
        try:
            query = db.query(Grade).join(Student).join(Course)
            if filters:
                if filters.get("student_id"):
                    query = query.filter(Grade.student_id == filters["student_id"])
                if filters.get("course_id"):
                    query = query.filter(Grade.course_id == filters["course_id"])

            grades = query.limit(limit).all()

            # Write CSV file
            file_path = self.get_export_path(export_job_id, "csv")
            with open(file_path, "w", newline="", encoding="utf-8") as f:
                writer = csv.writer(f)
                writer.writerow(["Student ID", "Student Name", "Course Code", "Grade", "Points", "Recorded At"])
                for grade in grades:
                    writer.writerow([
                        grade.student.id if grade.student else "",
                        f"{grade.student.first_name} {grade.student.last_name}" if grade.student else "",
                        grade.course.course_code if grade.course else "",
                        grade.grade,
                        grade.points or "",
                        grade.recorded_at.isoformat() if grade.recorded_at else ""
                    ])

            # Update export job
            export_job = db.query(ExportJob).filter(ExportJob.id == export_job_id).first()
            if export_job:
                export_job.status = ExportStatus.COMPLETED
                export_job.file_path = file_path
                export_job.total_records = len(grades)
                export_job.progress_percent = 100
                export_job.completed_at = datetime.now(timezone.utc)
                db.commit()

            return True, file_path, "CSV export completed"
        except Exception as e:
            logger.error(f"CSV export failed: {e}")
            export_job = db.query(ExportJob).filter(ExportJob.id == export_job_id).first()
            if export_job:
                export_job.status = ExportStatus.FAILED
                export_job.error_message = str(e)
                db.commit()
            return False, None, str(e)

    # ===== PDF EXPORT METHODS =====

    def generate_students_pdf(
        self, db: Session, export_job_id: int, filters: Optional[Dict[str, Any]] = None, limit: int = 10000
    ) -> tuple[bool, Optional[str], str]:
        """Generate PDF export of students."""
        if not REPORTLAB_AVAILABLE:
            return False, None, "ReportLab not installed. Install with: pip install reportlab"

        try:
            query = db.query(Student)
            if filters:
                if filters.get("status"):
                    query = query.filter(Student.status == filters["status"])
                if filters.get("search"):
                    search_term = f"%{filters['search']}%"
                    query = query.filter(
                        (Student.first_name.ilike(search_term)) | (Student.last_name.ilike(search_term)) | (Student.email.ilike(search_term))
                    )

            students = query.limit(limit).all()

            # Create PDF file
            file_path = self.get_export_path(export_job_id, "pdf")
            doc = SimpleDocTemplate(file_path, pagesize=letter)
            elements = []
            styles = getSampleStyleSheet()

            # Add title
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=16,
                textColor=colors.HexColor('#1f77b4'),
                spaceAfter=20
            )
            elements.append(Paragraph("Student Export Report", title_style))
            elements.append(Spacer(1, 0.3*inch))

            # Prepare table data
            table_data = [["ID", "First Name", "Last Name", "Email", "Status"]]
            for student in students:
                table_data.append([
                    str(student.id),
                    student.first_name,
                    student.last_name,
                    student.email,
                    student.status
                ])

            # Create and style table
            table = Table(table_data, colWidths=[0.6*inch, 1.2*inch, 1.2*inch, 2*inch, 0.8*inch])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1f77b4')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 9),
            ]))
            elements.append(table)

            # Build PDF
            doc.build(elements)

            # Update export job
            export_job = db.query(ExportJob).filter(ExportJob.id == export_job_id).first()
            if export_job:
                export_job.status = ExportStatus.COMPLETED
                export_job.file_path = file_path
                export_job.total_records = len(students)
                export_job.progress_percent = 100
                export_job.completed_at = datetime.now(timezone.utc)
                db.commit()

            return True, file_path, "PDF export completed"
        except Exception as e:
            logger.error(f"PDF export failed: {e}")
            export_job = db.query(ExportJob).filter(ExportJob.id == export_job_id).first()
            if export_job:
                export_job.status = ExportStatus.FAILED
                export_job.error_message = str(e)
                db.commit()
            return False, None, str(e)

    def generate_courses_pdf(
        self, db: Session, export_job_id: int, filters: Optional[Dict[str, Any]] = None, limit: int = 10000
    ) -> tuple[bool, Optional[str], str]:
        """Generate PDF export of courses."""
        if not REPORTLAB_AVAILABLE:
            return False, None, "ReportLab not installed. Install with: pip install reportlab"

        try:
            query = db.query(Course)
            if filters:
                if filters.get("status"):
                    query = query.filter(Course.status == filters["status"])
                if filters.get("search"):
                    search_term = f"%{filters['search']}%"
                    query = query.filter(
                        (Course.course_code.ilike(search_term)) | (Course.course_name.ilike(search_term))
                    )

            courses = query.limit(limit).all()

            # Create PDF file
            file_path = self.get_export_path(export_job_id, "pdf")
            doc = SimpleDocTemplate(file_path, pagesize=letter)
            elements = []
            styles = getSampleStyleSheet()

            # Add title
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=16,
                textColor=colors.HexColor('#1f77b4'),
                spaceAfter=20
            )
            elements.append(Paragraph("Course Export Report", title_style))
            elements.append(Spacer(1, 0.3*inch))

            # Prepare table data
            table_data = [["Code", "Name", "Instructor", "Status"]]
            for course in courses:
                table_data.append([
                    course.course_code,
                    course.course_name[:30],  # Truncate long names
                    course.instructor_name or "N/A",
                    course.status
                ])

            # Create and style table
            table = Table(table_data, colWidths=[1*inch, 2*inch, 1.5*inch, 1*inch])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1f77b4')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 9),
            ]))
            elements.append(table)

            # Build PDF
            doc.build(elements)

            # Update export job
            export_job = db.query(ExportJob).filter(ExportJob.id == export_job_id).first()
            if export_job:
                export_job.status = ExportStatus.COMPLETED
                export_job.file_path = file_path
                export_job.total_records = len(courses)
                export_job.progress_percent = 100
                export_job.completed_at = datetime.now(timezone.utc)
                db.commit()

            return True, file_path, "PDF export completed"
        except Exception as e:
            logger.error(f"PDF export failed: {e}")
            export_job = db.query(ExportJob).filter(ExportJob.id == export_job_id).first()
            if export_job:
                export_job.status = ExportStatus.FAILED
                export_job.error_message = str(e)
                db.commit()
            return False, None, str(e)

    def generate_grades_pdf(
        self, db: Session, export_job_id: int, filters: Optional[Dict[str, Any]] = None, limit: int = 10000
    ) -> tuple[bool, Optional[str], str]:
        """Generate PDF export of grades."""
        if not REPORTLAB_AVAILABLE:
            return False, None, "ReportLab not installed. Install with: pip install reportlab"

        try:
            query = db.query(Grade).join(Student).join(Course)
            if filters:
                if filters.get("student_id"):
                    query = query.filter(Grade.student_id == filters["student_id"])
                if filters.get("course_id"):
                    query = query.filter(Grade.course_id == filters["course_id"])

            grades = query.limit(limit).all()

            # Create PDF file
            file_path = self.get_export_path(export_job_id, "pdf")
            doc = SimpleDocTemplate(file_path, pagesize=letter)
            elements = []
            styles = getSampleStyleSheet()

            # Add title
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=16,
                textColor=colors.HexColor('#1f77b4'),
                spaceAfter=20
            )
            elements.append(Paragraph("Grade Export Report", title_style))
            elements.append(Spacer(1, 0.3*inch))

            # Prepare table data
            table_data = [["Student", "Course", "Grade", "Points"]]
            for grade in grades:
                student_name = f"{grade.student.first_name} {grade.student.last_name}" if grade.student else "N/A"
                course_code = grade.course.course_code if grade.course else "N/A"
                table_data.append([
                    student_name[:25],
                    course_code,
                    grade.grade,
                    str(grade.points) if grade.points else "N/A"
                ])

            # Create and style table
            table = Table(table_data, colWidths=[2*inch, 1.5*inch, 0.8*inch, 0.8*inch])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1f77b4')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 9),
            ]))
            elements.append(table)

            # Build PDF
            doc.build(elements)

            # Update export job
            export_job = db.query(ExportJob).filter(ExportJob.id == export_job_id).first()
            if export_job:
                export_job.status = ExportStatus.COMPLETED
                export_job.file_path = file_path
                export_job.total_records = len(grades)
                export_job.progress_percent = 100
                export_job.completed_at = datetime.now(timezone.utc)
                db.commit()

            return True, file_path, "PDF export completed"
        except Exception as e:
            logger.error(f"PDF export failed: {e}")
            export_job = db.query(ExportJob).filter(ExportJob.id == export_job_id).first()
            if export_job:
                export_job.status = ExportStatus.FAILED
                export_job.error_message = str(e)
                db.commit()
            return False, None, str(e)

    def process_export_task(
        self, export_job_id: int, export_type: str, export_format: str = "excel",
        filters: Optional[Dict[str, Any]] = None, limit: int = 10000
    ) -> tuple[bool, Optional[str], str]:
        """Process export task in background.

        This is called as a background task and handles the heavy lifting.
        Supports multiple formats: 'excel', 'csv', 'pdf'
        """
        # Create fresh session for background task
        from backend.db import SessionLocal

        db = SessionLocal()

        try:
            # Update status to processing
            export_job = db.query(ExportJob).filter(ExportJob.id == export_job_id).first()
            if export_job:
                export_job.status = ExportStatus.PROCESSING
                db.commit()

            # Route to appropriate generator based on format
            format_lower = (export_format or "excel").lower()

            if format_lower == "csv":
                if export_type == "students":
                    return self.generate_students_csv(db, export_job_id, filters, limit)
                elif export_type == "courses":
                    return self.generate_courses_csv(db, export_job_id, filters, limit)
                elif export_type == "grades":
                    return self.generate_grades_csv(db, export_job_id, filters, limit)
            elif format_lower == "pdf":
                if export_type == "students":
                    return self.generate_students_pdf(db, export_job_id, filters, limit)
                elif export_type == "courses":
                    return self.generate_courses_pdf(db, export_job_id, filters, limit)
                elif export_type == "grades":
                    return self.generate_grades_pdf(db, export_job_id, filters, limit)
            else:  # Default to Excel
                if export_type == "students":
                    return self.generate_students_excel(db, export_job_id, filters, limit)
                elif export_type == "courses":
                    return self.generate_courses_excel(db, export_job_id, filters, limit)
                elif export_type == "grades":
                    return self.generate_grades_excel(db, export_job_id, filters, limit)

            raise ValueError(f"Unknown export type: {export_type}")

        finally:
            db.close()

    def get_export_download_url(self, export_job_id: int) -> Optional[str]:
        """Get download URL for completed export."""
        # This would be the endpoint where users can download the file
        # Format: /api/v1/export-jobs/{export_job_id}/download
        return f"/api/v1/export-jobs/{export_job_id}/download"
