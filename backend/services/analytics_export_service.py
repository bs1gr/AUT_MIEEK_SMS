# -*- coding: utf-8 -*-
"""
Analytics Export Service
Provides functionality to export analytics data to PDF and Excel formats.
"""

import io
import logging
from datetime import datetime as dt, timezone
from typing import Any, Dict, List, Optional

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Flowable
from reportlab.lib import colors
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from sqlalchemy.orm import Session

logger = logging.getLogger(__name__)




def _register_analytics_fonts() -> tuple[str, str]:
    """Register Unicode-capable fonts for PDF rendering - exact pattern from report_exporters.py"""
    font_regular = "DejaVuSans"
    font_bold = "DejaVuSans-Bold"

    from pathlib import Path

    fonts_dir = Path(__file__).resolve().parents[1] / "fonts"
    regular_path = fonts_dir / "DejaVuSans.ttf"
    bold_path = fonts_dir / "DejaVuSans-Bold.ttf"

    if regular_path.exists() and bold_path.exists():
        registered = set(pdfmetrics.getRegisteredFontNames())
        if font_regular not in registered:
            pdfmetrics.registerFont(TTFont(font_regular, str(regular_path)))
        if font_bold not in registered:
            pdfmetrics.registerFont(TTFont(font_bold, str(bold_path)))
        return font_regular, font_bold

    logger.warning(f"DejaVuSans fonts not found at {fonts_dir}. Greek PDFs will render with Helvetica fallback.")
    return "Helvetica", "Helvetica-Bold"

    return "Helvetica", "Helvetica-Bold"


class AnalyticsExportService:
    """Service for exporting analytics data to various formats."""

    # Localization strings
    TRANSLATIONS = {
        "en": {
            "title": "Analytics Dashboard Report",
            "exported": "Exported",
            "generated": "Generated",
            "summary_statistics": "Summary Statistics",
            "total_students": "Total Students",
            "total_courses": "Total Courses",
            "average_grade": "Average Grade %",
            "average_attendance": "Average Attendance %",
            "class_averages": "Class Averages",
            "class": "Class",
            "student_count": "Student Count",
            "course_averages": "Course Averages",
            "course_name": "Course Name",
            "enrollments": "Enrollments",
        },
        "el": {
            "title": "Αναφορά Αναλυτικών Δεδομένων",
            "exported": "Εξαγωγή",
            "generated": "Δημιουργήθηκε",
            "summary_statistics": "Στατιστικά Σύνοψης",
            "total_students": "Σύνολο Φοιτητών",
            "total_courses": "Σύνολο Μαθημάτων",
            "average_grade": "Μέσος Όρος Βαθμών %",
            "average_attendance": "Μέσος Όρος Παρουσίας %",
            "class_averages": "Μέσοι Όροι Τάξης",
            "class": "Τάξη",
            "student_count": "Αριθμός Φοιτητών",
            "course_averages": "Μέσοι Όροι Μαθήματος",
            "course_name": "Όνομα Μαθήματος",
            "enrollments": "Εγγραφές",
        },
    }

    def __init__(self, db: Session, language: str = "en", timezone: str = "Europe/Athens"):
        self.db = db
        self.language = language if language in self.TRANSLATIONS else "en"
        self.timezone = timezone
        self.t = self.TRANSLATIONS[self.language]
        logger.info(f"AnalyticsExportService initialized with language={language}, timezone={timezone}")

    def _get_bold_font_name(self, font_name: str) -> str:
        """Get the appropriate bold font name for the given font."""
        if font_name == "DejaVuSans":
            return "DejaVuSans-Bold"
        elif font_name == "Helvetica":
            return "Helvetica-Bold"
        elif "-Bold" in font_name:
            return font_name
        else:
            return f"{font_name}-Bold"

    def format_datetime(self, dt_obj: Optional[Any] = None) -> str:
        """Format datetime with localized date format matching frontend settings."""
        if dt_obj is None:
            dt_obj = dt.now(timezone.utc).replace(tzinfo=None)

        if not isinstance(dt_obj, dt):
            logger.error(f"format_datetime called with non-datetime: {type(dt_obj)}")
            return "Invalid date"

        try:
            from zoneinfo import ZoneInfo

            # Convert UTC to the specified timezone
            utc_tz = ZoneInfo("UTC")
            local_tz = ZoneInfo(self.timezone)

            # If dt_obj is naive, assume it's UTC
            if dt_obj.tzinfo is None:
                dt_obj = dt_obj.replace(tzinfo=utc_tz)

            local_dt = dt_obj.astimezone(local_tz)
            tz_abbr = local_dt.strftime("%Z")  # e.g., 'EEST' or 'EET'

            if self.language == "el":
                # Greek date format: DD/MM/YYYY HH:MM:SS TZ (matching gr-ddmmyyyy frontend setting)
                return local_dt.strftime(f"%d/%m/%Y %H:%M:%S {tz_abbr}")
            else:
                # English date format: MM/DD/YYYY HH:MM:SS TZ (matching en-us frontend setting)
                return local_dt.strftime(f"%m/%d/%Y %H:%M:%S {tz_abbr}")
        except Exception as e:
            logger.warning(f"Failed to format datetime with timezone {self.timezone}: {e}, falling back to UTC")
            try:
                if self.language == "el":
                    return dt_obj.strftime("%d/%m/%Y %H:%M:%S UTC")
                else:
                    return dt_obj.strftime("%m/%d/%Y %H:%M:%S UTC")
            except (AttributeError, TypeError):
                return "Invalid date"

    def export_dashboard_to_excel(
        self,
        filename: str = "analytics_dashboard.xlsx",
        data: Optional[Dict[str, Any]] = None,
    ) -> bytes:
        """Export dashboard summary data to Excel format."""
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Analytics Dashboard"

            # Style definitions
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=12)
            subheader_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            subheader_font = Font(bold=True, size=11)
            border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )

            # Title
            ws["A1"] = self.t["title"]
            ws["A1"].font = Font(bold=True, size=14)
            ws.merge_cells("A1:D1")

            # Date exported
            ws["A2"] = f"{self.t['exported']}: {self.format_datetime()}"
            ws["A2"].font = Font(italic=True, size=10)
            ws.merge_cells("A2:D2")

            row = 4

            # Summary Section
            if data and "summary" in data:
                summary = data["summary"]
                ws[f"A{row}"] = self.t["summary_statistics"]
                ws[f"A{row}"].fill = subheader_fill
                ws[f"A{row}"].font = subheader_font
                ws.merge_cells(f"A{row}:B{row}")
                row += 1

                summary_items = [
                    (self.t["total_students"], summary.get("total_students", 0)),
                    (self.t["total_courses"], summary.get("total_courses", 0)),
                    (self.t["average_grade"], f"{summary.get('average_grade', 0):.2f}%"),
                    (self.t["average_attendance"], f"{summary.get('average_attendance', 0):.2f}%"),
                ]

                for label, value in summary_items:
                    ws[f"A{row}"] = label
                    ws[f"B{row}"] = value
                    ws[f"A{row}"].fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
                    ws[f"A{row}"].border = border
                    ws[f"B{row}"].border = border
                    row += 1

            # Class Averages Section
            if data and "class_averages" in data:
                row += 1
                ws[f"A{row}"] = self.t["class_averages"]
                ws[f"A{row}"].fill = subheader_fill
                ws[f"A{row}"].font = subheader_font
                ws.merge_cells(f"A{row}:C{row}")
                row += 1

                # Headers
                headers = [self.t["class"], self.t["student_count"], self.t["average_grade"]]
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=row, column=col)
                    cell.value = header
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.border = border
                    cell.alignment = Alignment(horizontal="center")

                row += 1

                # Data
                for item in data["class_averages"]:
                    ws[f"A{row}"] = item.get("label", "")
                    ws[f"B{row}"] = item.get("count", 0)
                    ws[f"C{row}"] = f"{item.get('average', 0):.2f}%"
                    for col in range(1, 4):
                        ws.cell(row=row, column=col).border = border
                        ws.cell(row=row, column=col).alignment = Alignment(horizontal="center")
                    row += 1

            # Course Averages Section
            if data and "course_averages" in data:
                row += 1
                ws[f"A{row}"] = self.t["course_averages"]
                ws[f"A{row}"].fill = subheader_fill
                ws[f"A{row}"].font = subheader_font
                ws.merge_cells(f"A{row}:C{row}")
                row += 1

                # Headers
                headers = [self.t["course_name"], self.t["enrollments"], self.t["average_grade"]]
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=row, column=col)
                    cell.value = header
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.border = border
                    cell.alignment = Alignment(horizontal="center")

                row += 1

                # Data
                for item in data["course_averages"][:20]:  # Limit to first 20 for readability
                    ws[f"A{row}"] = item.get("label", "")
                    ws[f"B{row}"] = item.get("count", 0)
                    ws[f"C{row}"] = f"{item.get('average', 0):.2f}%"
                    for col in range(1, 4):
                        ws.cell(row=row, column=col).border = border
                        ws.cell(row=row, column=col).alignment = Alignment(horizontal="center")
                    row += 1

            # Adjust column widths
            ws.column_dimensions["A"].width = 25
            ws.column_dimensions["B"].width = 18
            ws.column_dimensions["C"].width = 18
            ws.column_dimensions["D"].width = 18

            # Save to bytes
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            return output.getvalue()

        except Exception as exc:
            logger.error("Failed to export dashboard to Excel: %s", exc, exc_info=True)
            raise

    def export_dashboard_to_pdf(
        self,
        filename: str = "analytics_dashboard.pdf",
        data: Optional[Dict[str, Any]] = None,
    ) -> bytes:
        """Export dashboard summary data to PDF format using proven report_exporters pattern."""
        try:
            # Register fonts each time (same as report_exporters.py)
            base_font, base_font_bold = _register_analytics_fonts()

            output = io.BytesIO()
            doc = SimpleDocTemplate(
                output,
                pagesize=letter,
                topMargin=0.5 * inch,
                bottomMargin=0.5 * inch,
            )

            elements: List[Flowable] = []
            styles = getSampleStyleSheet()

            # Styles - matching report_exporters.py pattern
            title_style = ParagraphStyle(
                "CustomTitle",
                parent=styles["Heading1"],
                fontSize=24,
                textColor=colors.HexColor("#4472C4"),
                spaceAfter=12,
                fontName=base_font,
            )
            heading_style = ParagraphStyle(
                "CustomHeading",
                parent=styles["Heading2"],
                fontSize=16,
                textColor=colors.HexColor("#2F5496"),
                spaceAfter=10,
                spaceBefore=10,
                fontName=base_font,
            )
            normal_style = ParagraphStyle(
                "AnalyticsNormal",
                parent=styles["Normal"],
                fontName=base_font,
                fontSize=10,
                leading=12,
            )
            table_header_style = ParagraphStyle(
                "AnalyticsTableHeader",
                parent=styles["Normal"],
                fontName=base_font_bold,
                fontSize=10,
                leading=12,
                alignment=1,
                textColor=colors.whitesmoke,
            )
            table_cell_style = ParagraphStyle(
                "AnalyticsTableCell",
                parent=styles["Normal"],
                fontName=base_font,
                fontSize=10,
                leading=12,
                alignment=1,
            )

            # Title
            elements.append(Paragraph(self.t["title"], title_style))
            elements.append(Spacer(1, 0.2 * inch))

            # Summary Section
            if data and "summary" in data:
                summary = data["summary"]
                elements.append(Paragraph(self.t["summary_statistics"], heading_style))

                summary_data = [
                    [
                        Paragraph(self.t["total_students"], table_header_style),
                        Paragraph(str(summary.get("total_students", 0)), table_cell_style),
                    ],
                    [
                        Paragraph(self.t["total_courses"], table_header_style),
                        Paragraph(str(summary.get("total_courses", 0)), table_cell_style),
                    ],
                    [
                        Paragraph(self.t["average_grade"], table_header_style),
                        Paragraph(f"{summary.get('average_grade', 0):.2f}%", table_cell_style),
                    ],
                    [
                        Paragraph(self.t["average_attendance"], table_header_style),
                        Paragraph(f"{summary.get('average_attendance', 0):.2f}%", table_cell_style),
                    ],
                ]

                summary_table = Table(summary_data, colWidths=[3 * inch, 3 * inch])
                summary_table.setStyle(
                    TableStyle(
                        [
                            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4472C4")),
                            ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                            ("FONTNAME", (0, 0), (-1, -1), base_font),
                            ("FONTSIZE", (0, 0), (-1, -1), 10),
                            ("FONTNAME", (0, 0), (-1, 0), base_font_bold),
                            ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
                            ("TOPPADDING", (0, 0), (-1, -1), 8),
                            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F3F4F6")]),
                        ]
                    )
                )
                elements.append(summary_table)
                elements.append(Spacer(1, 0.3 * inch))

            # Class Averages Section
            if data and "class_averages" in data:
                elements.append(Paragraph(self.t["class_averages"], heading_style))

                class_data = [
                    [
                        Paragraph(self.t["class"], table_header_style),
                        Paragraph(self.t["student_count"], table_header_style),
                        Paragraph(self.t["average_grade"], table_header_style),
                    ]
                ]
                for item in data["class_averages"][:10]:
                    class_data.append(
                        [
                            Paragraph(item.get("label", ""), table_cell_style),
                            Paragraph(str(item.get("count", 0)), table_cell_style),
                            Paragraph(f"{item.get('average', 0):.2f}%", table_cell_style),
                        ]
                    )

                class_table = Table(class_data, colWidths=[2.5 * inch, 1.5 * inch, 1.5 * inch])
                class_table.setStyle(
                    TableStyle(
                        [
                            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4472C4")),
                            ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                            ("FONTNAME", (0, 0), (-1, -1), base_font),
                            ("FONTSIZE", (0, 0), (-1, -1), 10),
                            ("FONTNAME", (0, 0), (-1, 0), base_font_bold),
                            ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
                            ("TOPPADDING", (0, 0), (-1, -1), 8),
                            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F3F4F6")]),
                        ]
                    )
                )
                elements.append(class_table)
                elements.append(Spacer(1, 0.3 * inch))

            # Course Averages Section
            if data and "course_averages" in data:
                elements.append(Paragraph(self.t["course_averages"], heading_style))

                course_data = [
                    [
                        Paragraph(self.t["course_name"], table_header_style),
                        Paragraph(self.t["enrollments"], table_header_style),
                        Paragraph(self.t["average_grade"], table_header_style),
                    ]
                ]
                for item in data["course_averages"][:15]:
                    course_data.append(
                        [
                            Paragraph(item.get("label", "")[:30], table_cell_style),
                            Paragraph(str(item.get("count", 0)), table_cell_style),
                            Paragraph(f"{item.get('average', 0):.2f}%", table_cell_style),
                        ]
                    )

                course_table = Table(course_data, colWidths=[2.5 * inch, 1.5 * inch, 1.5 * inch])
                course_table.setStyle(
                    TableStyle(
                        [
                            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4472C4")),
                            ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                            ("FONTNAME", (0, 0), (-1, -1), base_font),
                            ("FONTSIZE", (0, 0), (-1, -1), 10),
                            ("FONTNAME", (0, 0), (-1, 0), base_font_bold),
                            ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
                            ("TOPPADDING", (0, 0), (-1, -1), 8),
                            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F3F4F6")]),
                        ]
                    )
                )
                elements.append(course_table)
                elements.append(Spacer(1, 0.2 * inch))

            # Add generated timestamp at end
            elements.append(Paragraph(f"{self.t['generated']}: {self.format_datetime()}", normal_style))

            doc.build(elements)
            output.seek(0)
            return output.getvalue()

        except Exception as exc:
            logger.error("Failed to export dashboard to PDF: %s", exc, exc_info=True)
            raise

    def export_student_performance_to_excel(
        self,
        student_id: int,
        performance_data: List[Dict[str, Any]],
        student_name: str = "Student",
        filename: str = "student_performance.xlsx",
    ) -> bytes:
        """Export student performance data to Excel."""
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Performance"

            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )

            # Header
            ws["A1"] = f"Performance Report: {student_name}"
            ws["A1"].font = Font(bold=True, size=12)
            ws.merge_cells("A1:D1")

            ws["A2"] = f"Generated: {dt.utcnow().strftime('%Y-%m-%d %H:%M:%S UTC')}"
            ws["A2"].font = Font(italic=True, size=9)
            ws.merge_cells("A2:D2")

            row = 4

            # Headers
            headers = ["Date", "Course", "Grade", "Percentage %"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
                cell.border = border
                cell.alignment = Alignment(horizontal="center")

            row += 1

            # Data
            for item in performance_data:
                ws[f"A{row}"] = item.get("date", "")
                ws[f"B{row}"] = item.get("course", "")
                ws[f"C{row}"] = item.get("grade", 0)
                ws[f"D{row}"] = f"{item.get('percentage', 0):.2f}%"
                for col in range(1, 5):
                    ws.cell(row=row, column=col).border = border
                row += 1

            # Adjust widths
            ws.column_dimensions["A"].width = 15
            ws.column_dimensions["B"].width = 20
            ws.column_dimensions["C"].width = 12
            ws.column_dimensions["D"].width = 15

            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            return output.getvalue()

        except Exception as exc:
            logger.error("Failed to export student performance: %s", exc, exc_info=True)
            raise
