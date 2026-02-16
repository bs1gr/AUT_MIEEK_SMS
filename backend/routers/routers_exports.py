import os
import csv
import zipfile
import re
from typing import Any

from fastapi import APIRouter, Depends, HTTPException, Request, Query
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

# Simple i18n dict for EN/EL (expand as needed)
TRANSLATIONS = {
    "en": {
        "report_title": "Comprehensive Student Report",
        "attendance_summary": "Attendance Summary",
        "total_classes": "Total Classes",
        "present": "Present",
        "absent": "Absent",
        "late": "Late",
        "excused": "Excused",
        "attendance_rate": "Attendance Rate",
        "grades_by_course": "Grades by Course",
        "assignment": "Assignment",
        "category": "Category",
        "score": "Score",
        "max_score": "Max Score",
        "percentage": "Percentage",
        "letter": "Letter",
        "average": "Average:",
        "daily_performance_summary": "Daily Performance Summary",
        "total_daily_entries": "Total Daily Performance Entries",
        "avg_performance": "Average Performance",
        "generated_on": "Generated on",
        "system": "Student Management System",
        "sheet_students": "Students",
        "sheet_grades": "Grades",
        "sheet_attendance": "Attendance",
        "sheet_overview": "Overview",
        "sheet_course_summary": "Course Summary",
        "sheet_period_summary": "Period Summary",
        "sheet_course_periods": "Course Periods",
        "sheet_student_summary": "Student Summary",
        "sheet_daily_overview": "Daily Overview",
        "sheet_courses": "Courses",
        "sheet_enrollments": "Enrollments",
        "sheet_all_grades": "All Grades",
        "sheet_daily_performance": "Daily Performance",
        "sheet_highlights": "Highlights",
        "title_attendance_export": "Attendance Analytics Export",
        "title_course_catalog": "Course Catalog",
        "title_course_analytics": "Course Analytics Report",
        "title_course_statistics": "Course Statistics",
        "title_grade_distribution": "Grade Distribution",
        "student_directory_title": "Student Directory",
        "grade_report_title": "Grade Report: {student_name} ({student_code})",
        "label_metric": "Metric",
        "label_value": "Value",
        "label_total_records": "Total Records",
        "label_unique_students": "Unique Students",
        "label_unique_courses": "Unique Courses",
        "label_date_range": "Date Range",
        "label_present_share": "Present Share",
        "label_status": "Status",
        "label_count": "Count",
        "label_notice": "Notice",
        "label_no_attendance": "No attendance records were found in the system.",
        "label_total_students_enrolled": "Total Students Enrolled",
        "label_total_assignments": "Total Assignments",
        "label_average_grade": "Average Grade",
        "label_highest_grade": "Highest Grade",
        "label_lowest_grade": "Lowest Grade",
        "label_letter_grade": "Letter Grade",
        "label_percentage": "Percentage",
        "label_most_frequent_status": "Most Frequent Status",
        "status_active": "Active",
        "status_inactive": "Inactive",
        "not_available": "N/A",
        "yes_label": "Yes",
        "no_label": "No",
        "attendance_rate_percent": "Attendance Rate (%)",
        "label_enrolled_date": "Enrolled Date",
        "label_date_submitted": "Date Submitted",
        "label_hours_per_week": "Hours/Week",
        "label_periods_per_week": "Periods/Week",
        "label_description": "Description",
        "label_highlight_text": "Highlight Text",
        "label_date_created": "Date Created",
        "label_is_positive": "Is Positive",
        "label_rating": "Rating",
        "label_course_code": "Course Code",
        "label_course_name": "Course Name",
        "label_course_id": "Course ID",
        "label_student_name": "Student Name",
        "label_student_id": "Student ID",
        "label_first_name": "First Name",
        "label_last_name": "Last Name",
        "label_email": "Email",
        "label_enrollment_date": "Enrollment Date",
        "label_period": "Period",
        "label_date": "Date",
        "label_notes": "Notes",
        "label_semester": "Semester",
        "label_credits": "Credits",
        "label_hours": "Hours/Week",
        "label_generated_on_table": "Generated On",
        "label_daily_entries": "Daily Performance Entries",
        "label_avg_performance_short": "Average Performance",
        "distribution_a": "A (90-100%)",
        "distribution_b": "B (80-89%)",
        "distribution_c": "C (70-79%)",
        "distribution_d": "D (60-69%)",
        "distribution_f": "F (<60%)",
        "header_id": "ID",
        "label_weight": "Weight",
        "label_grade": "Grade",
    },
    "el": {
        "report_title": "Αναλυτική Αναφορά Σπουδαστή",
        "attendance_summary": "Σύνοψη Παρουσιών",
        "total_classes": "Σύνολο Μαθημάτων",
        "present": "Παρόν",
        "absent": "Απών",
        "late": "Καθυστέρηση",
        "excused": "Δικαιολογημένη",
        "attendance_rate": "Ποσοστό Παρουσιών",
        "grades_by_course": "Βαθμοί ανά Μάθημα",
        "assignment": "Άσκηση",
        "category": "Κατηγορία",
        "score": "Βαθμός",
        "max_score": "Μέγιστος Βαθμός",
        "percentage": "Ποσοστό",
        "letter": "Γράμμα",
        "average": "Μέσος όρος:",
        "daily_performance_summary": "Σύνοψη Ημερήσιας Επίδοσης",
        "total_daily_entries": "Σύνολο Καταχωρήσεων Ημερήσιας Επίδοσης",
        "avg_performance": "Μέση Επίδοση",
        "generated_on": "Δημιουργήθηκε στις",
        "system": "Σύστημα Διαχείρισης Σπουδαστών",
        "sheet_students": "Φοιτητές",
        "sheet_grades": "Βαθμοί",
        "sheet_attendance": "Παρουσίες",
        "sheet_overview": "Επισκόπηση",
        "sheet_course_summary": "Σύνοψη Μαθημάτων",
        "sheet_period_summary": "Σύνοψη Περιόδων",
        "sheet_course_periods": "Περίοδοι Μαθημάτων",
        "sheet_student_summary": "Σύνοψη Σπουδαστών",
        "sheet_daily_overview": "Ημερήσια Επισκόπηση",
        "sheet_courses": "Μαθήματα",
        "sheet_enrollments": "Εγγραφές",
        "sheet_all_grades": "Όλες οι Βαθμολογίες",
        "sheet_daily_performance": "Ημερήσια Επίδοση",
        "sheet_highlights": "Επισημάνσεις",
        "title_attendance_export": "Εξαγωγή Αναλύσεων Παρουσιών",
        "title_course_catalog": "Κατάλογος Μαθημάτων",
        "title_course_analytics": "Αναφορά Αναλυτικών Μαθήματος",
        "title_course_statistics": "Στατιστικά Μαθήματος",
        "title_grade_distribution": "Κατανομή Βαθμών",
        "student_directory_title": "Κατάλογος Σπουδαστών",
        "grade_report_title": "Αναφορά Βαθμολογίας: {student_name} ({student_code})",
        "label_metric": "Μετρική",
        "label_value": "Τιμή",
        "label_total_records": "Σύνολο Εγγραφών",
        "label_unique_students": "Μοναδικοί Σπουδαστές",
        "label_unique_courses": "Μοναδικά Μαθήματα",
        "label_date_range": "Εύρος Ημερομηνιών",
        "label_present_share": "Ποσοστό Παρόντων",
        "label_status": "Κατάσταση",
        "label_count": "Πλήθος",
        "label_notice": "Σημείωση",
        "label_no_attendance": "Δεν βρέθηκαν εγγραφές παρουσιών στο σύστημα.",
        "label_total_students_enrolled": "Σύνολο Εγγεγραμμένων Σπουδαστών",
        "label_total_assignments": "Σύνολο Εργασιών",
        "label_average_grade": "Μέσος Όρος Βαθμολογίας",
        "label_highest_grade": "Υψηλότερη Βαθμολογία",
        "label_lowest_grade": "Χαμηλότερη Βαθμολογία",
        "label_letter_grade": "Βαθμός Γράμματος",
        "label_percentage": "Ποσοστό",
        "label_most_frequent_status": "Συχνότερη Κατάσταση",
        "status_active": "Ενεργός",
        "status_inactive": "Ανενεργός",
        "not_available": "Μ/Δ",
        "yes_label": "Ναι",
        "no_label": "Όχι",
        "attendance_rate_percent": "Ποσοστό Παρουσιών (%)",
        "label_enrolled_date": "Ημερομηνία Εγγραφής",
        "label_date_submitted": "Ημερομηνία Υποβολής",
        "label_hours_per_week": "Ώρες/Εβδομάδα",
        "label_periods_per_week": "Περίοδοι/Εβδομάδα",
        "label_description": "Περιγραφή",
        "label_highlight_text": "Κείμενο Επισήμανσης",
        "label_date_created": "Ημερομηνία Δημιουργίας",
        "label_is_positive": "Θετική Αναφορά",
        "label_rating": "Αξιολόγηση",
        "label_course_code": "Κωδικός Μαθήματος",
        "label_course_name": "Τίτλος Μαθήματος",
        "label_course_id": "Αναγνωριστικό Μαθήματος",
        "label_student_name": "Όνομα Σπουδαστή",
        "label_student_id": "Αριθμός Μητρώου",
        "label_first_name": "Όνομα",
        "label_last_name": "Επώνυμο",
        "label_email": "Email",
        "label_enrollment_date": "Ημερομηνία Εγγραφής",
        "label_period": "Περίοδος",
        "label_date": "Ημερομηνία",
        "label_notes": "Σημειώσεις",
        "label_semester": "Εξάμηνο",
        "label_credits": "Πιστωτικές Μονάδες",
        "label_hours": "Ώρες/Εβδομάδα",
        "label_generated_on_table": "Δημιουργήθηκε",
        "label_daily_entries": "Καταχωρήσεις Ημερήσιας Επίδοσης",
        "label_avg_performance_short": "Μέση Επίδοση",
        "distribution_a": "A (90-100%)",
        "distribution_b": "B (80-89%)",
        "distribution_c": "C (70-79%)",
        "distribution_d": "D (60-69%)",
        "distribution_f": "F (<60%)",
        "header_id": "Κωδικός",
        "label_weight": "Βαρύτητα",
        "label_grade": "Βαθμός",
    },
}


ATTENDANCE_STATUSES = ("Present", "Absent", "Late", "Excused")
STATUS_LOOKUP = {status.upper(): status for status in ATTENDANCE_STATUSES}


def get_lang(request: Request):
    # Try query param, then Accept-Language header
    lang = request.query_params.get("lang")
    if lang and lang.lower() in TRANSLATIONS:
        return lang.lower()
    accept = request.headers.get("accept-language", "").lower()
    if "el" in accept:
        return "el"
    return "en"


def t(key: str, lang: str) -> str:
    return TRANSLATIONS.get(lang, TRANSLATIONS["en"]).get(key, key)


HEADER_DEFINITIONS = {
    "students": [
        "header_id",
        "label_first_name",
        "label_last_name",
        "label_email",
        "label_student_id",
        "label_enrollment_date",
        "label_status",
    ],
    "student_grades": [
        "assignment",
        "category",
        "score",
        "max_score",
        "percentage",
        "label_weight",
        "label_grade",
        "label_date",
    ],
    "attendance": [
        "header_id",
        "label_student_id",
        "label_course_id",
        "label_date",
        "label_status",
        "label_period",
        "label_notes",
    ],
    "overview": ["label_metric", "label_value"],
    "status_counts": ["label_status", "label_count"],
    "course_summary": [
        "label_course_code",
        "label_course_name",
        "label_total_records",
        "present",
        "absent",
        "late",
        "excused",
        "attendance_rate_percent",
    ],
    "period_summary": [
        "label_period",
        "label_total_records",
        "present",
        "absent",
        "late",
        "excused",
        "attendance_rate_percent",
    ],
    "course_periods": [
        "label_course_code",
        "label_course_name",
        "label_period",
        "label_total_records",
        "present",
        "absent",
        "late",
        "excused",
        "attendance_rate_percent",
    ],
    "student_summary": [
        "label_student_id",
        "label_student_name",
        "label_total_records",
        "present",
        "absent",
        "late",
        "excused",
        "label_most_frequent_status",
    ],
    "daily_overview": [
        "label_date",
        "label_total_records",
        "present",
        "absent",
        "late",
        "excused",
    ],
    "courses": [
        "header_id",
        "label_course_code",
        "label_course_name",
        "label_semester",
        "label_credits",
        "label_hours_per_week",
        "label_periods_per_week",
        "label_description",
    ],
    "enrollments": [
        "header_id",
        "label_student_id",
        "label_student_name",
        "label_course_id",
        "label_course_code",
        "label_course_name",
        "label_enrolled_date",
    ],
    "all_grades": [
        "header_id",
        "label_student_id",
        "label_student_name",
        "label_course_id",
        "label_course_name",
        "assignment",
        "category",
        "score",
        "max_score",
        "percentage",
        "label_weight",
        "label_date_submitted",
    ],
    "daily_performance": [
        "header_id",
        "label_student_id",
        "label_student_name",
        "label_course_id",
        "label_course_name",
        "label_date",
        "category",
        "score",
        "max_score",
        "percentage",
        "label_notes",
    ],
    "highlights": [
        "header_id",
        "label_student_id",
        "label_student_name",
        "label_semester",
        "category",
        "label_rating",
        "label_highlight_text",
        "label_date_created",
        "label_is_positive",
    ],
}


def get_header_row(key: str, lang: str) -> list[str]:
    keys = HEADER_DEFINITIONS.get(key, [])
    return [t(k, lang) for k in keys]


def translate_status_value(status: str, lang: str) -> str:
    normalized = (status or "").strip().lower()
    if normalized in {"present", "absent", "late", "excused"}:
        return t(normalized, lang)
    return status or ""


def not_available(lang: str) -> str:
    return t("not_available", lang)


def yes_no(value: bool, lang: str) -> str:
    return t("yes_label", lang) if value else t("no_label", lang)


def grade_report_heading(student_name: str, student_code: str, lang: str) -> str:
    template = t("grade_report_title", lang)
    return template.format(student_name=student_name, student_code=student_code)


def format_date_value(value: Any, lang: str, include_time: bool = False) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        dt = value
    elif isinstance(value, date):
        dt = datetime.combine(value, datetime.min.time())
    elif isinstance(value, str):
        try:
            dt = datetime.fromisoformat(value)
        except ValueError:
            return value
    else:
        return str(value)

    if lang == "el":
        fmt = "%d/%m/%Y %H:%M" if include_time else "%d/%m/%Y"
    else:
        fmt = "%Y-%m-%d %H:%M" if include_time else "%Y-%m-%d"
    return dt.strftime(fmt)


def translate_grade_category(value: str | None, lang: str) -> str:
    if not value:
        return ""
    normalized = value.strip().lower()
    if lang != "el":
        return value
    mapping = {
        "midterm": "Ενδιάμεση Εξέταση",
        "midterm exam": "Ενδιάμεση Εξέταση",
        "final": "Τελική Εξέταση",
        "final exam": "Τελική Εξέταση",
        "exam": "Εξέταση",
        "assignment": "Εργασία",
        "quiz": "Κουίζ",
        "project": "Εργασία",
        "homework": "Εργασία",
        "homework/assignments": "Εργασία",
        "homework / assignments": "Εργασία",
        "class participation": "Συμμετοχή στην Τάξη",
        "continuous assessment": "Συνεχής Αξιολόγηση",
    }
    return mapping.get(normalized, value)


def translate_assignment_name(value: str | None, lang: str) -> str:
    if not value:
        return ""
    if lang != "el":
        return value
    replacements = {
        "sample": "Δείγμα",
        "exam": "Εξέταση",
        "midterm": "Ενδιάμεση",
        "final": "Τελική",
        "assignment": "Εργασία",
        "project": "Εργασία",
        "quiz": "Κουίζ",
    }
    text = value
    for src, dest in replacements.items():
        text = re.sub(rf"\b{re.escape(src)}\b", dest, text, flags=re.IGNORECASE)
    return text


"""
Export Routes
Provides endpoints to export data to Excel/PDF.
"""

import logging
from datetime import datetime, date
from io import BytesIO, StringIO

import openpyxl
from fastapi.responses import StreamingResponse
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import Flowable, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from sqlalchemy.orm import Session

HEADER_FILL = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")


def _init_status_counts():
    return {status: 0 for status in ATTENDANCE_STATUSES}


def _normalize_status(status: str | None) -> str:
    if not status:
        return "Present"
    if not isinstance(status, str):
        return "Present"
    normalized = STATUS_LOOKUP.get(status.strip().upper())
    return normalized or "Present"


def _apply_table_header(ws: Worksheet, headers, row: int = 1):
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _auto_fit_columns(ws: Any):
    for column_cells in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            try:
                cell_length = len(str(cell.value)) if cell.value is not None else 0
            except Exception:
                cell_length = 0
            if cell_length > max_length:
                max_length = cell_length
        ws.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 45)


def _dominant_status(counts: dict[str, int]) -> str:
    best_status = ATTENDANCE_STATUSES[0]
    best_value = -1
    for status in ATTENDANCE_STATUSES:
        value = counts.get(status, 0)
        if value > best_value:
            best_status = status
            best_value = value
    return best_status


def _csv_response(headers: list[str], rows: list[list[Any]], filename: str) -> StreamingResponse:
    output = StringIO()
    writer = csv.writer(output)
    writer.writerow(headers)
    writer.writerows(rows)
    csv_content = "\ufeff" + output.getvalue()
    buffer = BytesIO(csv_content.encode("utf-8"))
    return StreamingResponse(
        buffer,
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


def _pdf_table_response(title: str, headers: list[str], rows: list[list[Any]], filename: str) -> StreamingResponse:
    font_path = os.path.join(os.path.dirname(__file__), "../fonts/DejaVuSans.ttf")
    pdfmetrics.registerFont(TTFont("DejaVuSans", font_path))
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, leftMargin=0.6 * inch, rightMargin=0.6 * inch)
    elements: list[Flowable] = []
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "CustomTitle",
        parent=styles["Heading1"],
        fontName="DejaVuSans",
        fontSize=20,
        textColor=colors.HexColor("#4F46E5"),
        spaceAfter=20,
        alignment=1,
    )
    header_style = ParagraphStyle(
        "TableHeader",
        parent=styles["Normal"],
        fontName="DejaVuSans",
        fontSize=9.5,
        leading=11,
        textColor=colors.white,
        alignment=1,
    )
    cell_style = ParagraphStyle(
        "TableCell",
        parent=styles["Normal"],
        fontName="DejaVuSans",
        fontSize=9,
        leading=11,
        alignment=1,
    )
    elements.append(Paragraph(title, title_style))
    elements.append(Spacer(1, 0.2 * inch))
    wrapped_headers = [Paragraph(str(h), header_style) for h in headers]
    wrapped_rows = [[Paragraph(str(cell or ""), cell_style) for cell in row] for row in rows]
    data = [wrapped_headers] + wrapped_rows
    col_widths = _fit_pdf_column_widths([headers] + rows, doc.width)
    table = Table(data, repeatRows=1, colWidths=col_widths)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4F46E5")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("FONTNAME", (0, 0), (-1, -1), "DejaVuSans"),
                ("FONTSIZE", (0, 0), (-1, 0), 10),
                ("BOTTOMPADDING", (0, 0), (-1, 0), 8),
                ("BACKGROUND", (0, 1), (-1, -1), colors.beige),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ]
        )
    )
    elements.append(table)
    doc.build(elements)
    buffer.seek(0)
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


def _fit_pdf_column_widths(rows: list[list[Any]], available_width: float) -> list[float]:
    if not rows:
        return []
    column_count = max(len(row) for row in rows)
    max_lengths = [0] * column_count
    for row in rows:
        for idx in range(column_count):
            value = row[idx] if idx < len(row) else ""
            text = str(value or "")
            max_lengths[idx] = max(max_lengths[idx], len(text))

    char_width = 5.2
    min_width = 0.8 * inch
    max_width = 2.6 * inch
    widths = [min(max_width, max(min_width, length * char_width)) for length in max_lengths]

    total = sum(widths)
    if total > available_width and total > 0:
        scale = available_width / total
        widths = [max(min_width, width * scale) for width in widths]
        total = sum(widths)
        if total > available_width:
            overflow = total - available_width
            widths[-1] = max(min_width, widths[-1] - overflow)

    return widths


def _build_wrapped_table(
    headers: list[str],
    rows: list[list[Any]],
    available_width: float,
    align: str = "CENTER",
    header_font_size: float = 9.5,
    cell_font_size: float = 9,
    header_bg: Any = colors.HexColor("#4F46E5"),
    body_bg: Any = colors.beige,
    grid_color: Any = colors.grey,
    extra_styles: list[tuple[Any, ...]] | None = None,
) -> Table:
    styles = getSampleStyleSheet()
    header_style = ParagraphStyle(
        "TableHeader",
        parent=styles["Normal"],
        fontName="DejaVuSans",
        fontSize=header_font_size,
        leading=header_font_size + 2,
        textColor=colors.white,
        alignment=1,
    )
    cell_style = ParagraphStyle(
        "TableCell",
        parent=styles["Normal"],
        fontName="DejaVuSans",
        fontSize=cell_font_size,
        leading=cell_font_size + 2,
        alignment=1 if align == "CENTER" else 0,
    )
    wrapped_headers = [Paragraph(str(h), header_style) for h in headers]
    wrapped_rows = [[Paragraph(str(cell or ""), cell_style) for cell in row] for row in rows]
    data = [wrapped_headers] + wrapped_rows
    col_widths = _fit_pdf_column_widths([headers] + rows, available_width)
    table = Table(data, repeatRows=1, colWidths=col_widths)
    style_commands: list[tuple[Any, ...]] = [
        ("BACKGROUND", (0, 0), (-1, 0), header_bg),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
        ("ALIGN", (0, 0), (-1, -1), align),
        ("FONTNAME", (0, 0), (-1, -1), "DejaVuSans"),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 8),
        ("BACKGROUND", (0, 1), (-1, -1), body_bg),
        ("GRID", (0, 0), (-1, -1), 0.5, grid_color),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ]
    if extra_styles:
        style_commands.extend(extra_styles)
    table.setStyle(TableStyle(style_commands))
    return table


def _csv_content(headers: list[str], rows: list[list[Any]]) -> str:
    output = StringIO()
    writer = csv.writer(output)
    if headers:
        writer.writerow(headers)
    writer.writerows(rows)
    return "\ufeff" + output.getvalue()


def _build_attendance_analytics_csv_rows(
    rows: list[Any],
    lang: str,
    na_value: str,
) -> list[list[Any]]:
    overall_counts = _init_status_counts()
    course_summary: dict[int, dict[str, Any]] = {}
    period_summary: dict[int, dict[str, Any]] = {}
    course_period_summary: dict[tuple[int, int], dict[str, Any]] = {}
    student_summary: dict[int, dict[str, Any]] = {}
    daily_summary: dict[Any, dict[str, Any]] = {}
    unique_students = set()
    unique_courses = set()
    date_values = set()

    for (
        _attendance_id,
        att_date,
        status,
        period_number,
        course_id,
        student_id,
        student_code,
        first_name,
        last_name,
        course_code,
        course_name,
    ) in rows:
        safe_status = _normalize_status(status)
        period = period_number or 1
        student_label = (f"{first_name or ''} {last_name or ''}").strip() or na_value
        course_label = course_code or na_value
        course_title = course_name or na_value

        overall_counts[safe_status] += 1
        unique_students.add(student_id)
        unique_courses.add(course_id)
        if att_date:
            date_values.add(att_date)

        course_entry = course_summary.setdefault(
            course_id,
            {"course_code": course_label, "course_name": course_title, "counts": _init_status_counts(), "total": 0},
        )
        course_entry["counts"][safe_status] += 1
        course_entry["total"] += 1

        period_entry = period_summary.setdefault(
            period,
            {"counts": _init_status_counts(), "total": 0},
        )
        period_entry["counts"][safe_status] += 1
        period_entry["total"] += 1

        course_period_entry = course_period_summary.setdefault(
            (course_id, period),
            {
                "course_code": course_label,
                "course_name": course_title,
                "period": period,
                "counts": _init_status_counts(),
                "total": 0,
            },
        )
        course_period_entry["counts"][safe_status] += 1
        course_period_entry["total"] += 1

        student_entry = student_summary.setdefault(
            student_id,
            {
                "student_code": student_code or str(student_id),
                "student_name": student_label,
                "counts": _init_status_counts(),
                "total": 0,
            },
        )
        student_entry["counts"][safe_status] += 1
        student_entry["total"] += 1

        if att_date:
            daily_entry = daily_summary.setdefault(att_date, {"counts": _init_status_counts(), "total": 0})
            daily_entry["counts"][safe_status] += 1
            daily_entry["total"] += 1

    total_records = sum(overall_counts.values())
    present_share = (overall_counts["Present"] / total_records * 100) if total_records else 0
    date_range = na_value
    if date_values:
        sorted_dates = sorted(date_values)
        date_range = f"{format_date_value(sorted_dates[0], lang)} → {format_date_value(sorted_dates[-1], lang)}"

    sections: list[list[list[Any]]] = []
    sections.append([[t("title_attendance_export", lang)]])
    sections.append([get_header_row("overview", lang)])
    sections.append(
        [
            [t("label_total_records", lang), total_records],
            [t("label_unique_students", lang), len(unique_students)],
            [t("label_unique_courses", lang), len(unique_courses)],
            [t("label_date_range", lang), date_range],
            [t("label_present_share", lang), f"{present_share:.1f}%"],
        ]
    )
    sections.append([get_header_row("status_counts", lang)])
    sections.append([[t(status.lower(), lang), overall_counts[status]] for status in ATTENDANCE_STATUSES])

    sections.append([get_header_row("course_summary", lang)])
    sections.append(
        [
            [
                data["course_code"],
                data["course_name"],
                data["total"],
                data["counts"]["Present"],
                data["counts"]["Absent"],
                data["counts"]["Late"],
                data["counts"]["Excused"],
                f"{(data['counts']['Present'] / data['total'] * 100) if data['total'] else 0:.1f}%",
            ]
            for data in sorted(course_summary.values(), key=lambda item: item["course_code"])
        ]
    )

    sections.append([get_header_row("period_summary", lang)])
    sections.append(
        [
            [
                period,
                data["total"],
                data["counts"]["Present"],
                data["counts"]["Absent"],
                data["counts"]["Late"],
                data["counts"]["Excused"],
                f"{(data['counts']['Present'] / data['total'] * 100) if data['total'] else 0:.1f}%",
            ]
            for period, data in sorted(period_summary.items())
        ]
    )

    sections.append([get_header_row("course_periods", lang)])
    sections.append(
        [
            [
                data["course_code"],
                data["course_name"],
                data["period"],
                data["total"],
                data["counts"]["Present"],
                data["counts"]["Absent"],
                data["counts"]["Late"],
                data["counts"]["Excused"],
                f"{(data['counts']['Present'] / data['total'] * 100) if data['total'] else 0:.1f}%",
            ]
            for _, data in sorted(
                course_period_summary.items(), key=lambda item: (item[1]["course_code"], item[1]["period"])
            )
        ]
    )

    sections.append([get_header_row("student_summary", lang)])
    sections.append(
        [
            [
                data["student_code"],
                data["student_name"],
                data["total"],
                data["counts"]["Present"],
                data["counts"]["Absent"],
                data["counts"]["Late"],
                data["counts"]["Excused"],
                translate_status_value(_dominant_status(data["counts"]), lang),
            ]
            for data in sorted(student_summary.values(), key=lambda item: item["student_name"])
        ]
    )

    sections.append([get_header_row("daily_overview", lang)])
    sections.append(
        [
            [
                format_date_value(day, lang),
                data["total"],
                data["counts"]["Present"],
                data["counts"]["Absent"],
                data["counts"]["Late"],
                data["counts"]["Excused"],
            ]
            for day, data in sorted(daily_summary.items())
        ]
    )

    flat_rows: list[list[Any]] = []
    for section in sections:
        for row in section:
            flat_rows.append(row)
        flat_rows.append([])

    return flat_rows


logger = logging.getLogger(__name__)

router = APIRouter(
    prefix="/export",
    tags=["Export"],
    responses={404: {"description": "Not found"}},
)


from backend.db import get_session as get_db
from backend.errors import ErrorCode, http_error
from backend.import_resolver import import_names
from backend.schemas.audit import AuditAction, AuditResource
from backend.rbac import require_permission
from backend.services.audit_service import AuditLogger


@router.get("/all/zip")
@require_permission("exports:generate")
async def export_all_zip(request: Request, db: Session = Depends(get_db)):
    """Export all data as a single ZIP (CSV files)."""
    audit = AuditLogger(db)
    try:
        Student, Course, Grade, Attendance, CourseEnrollment, DailyPerformance, Highlight = import_names(
            "models",
            "Student",
            "Course",
            "Grade",
            "Attendance",
            "CourseEnrollment",
            "DailyPerformance",
            "Highlight",
        )
        lang = get_lang(request)
        na_value = not_available(lang)

        buf = BytesIO()
        with zipfile.ZipFile(buf, mode="w", compression=zipfile.ZIP_DEFLATED) as zf:
            # Students
            students = (
                db.query(Student)
                .filter(Student.deleted_at.is_(None))
                .order_by(Student.last_name, Student.first_name)
                .all()
            )
            student_headers = get_header_row("students", lang)
            student_rows = [
                [
                    s.id,
                    s.first_name,
                    s.last_name,
                    s.email,
                    s.student_id,
                    format_date_value(s.enrollment_date, lang),
                    t("status_active", lang) if s.is_active else t("status_inactive", lang),
                ]
                for s in students
            ]
            zf.writestr("students.csv", _csv_content(student_headers, student_rows))

            # Courses
            courses = db.query(Course).filter(Course.deleted_at.is_(None)).all()
            course_headers = get_header_row("courses", lang)
            course_rows = [
                [
                    c.id,
                    c.course_code,
                    c.course_name,
                    c.semester,
                    c.credits,
                    c.hours_per_week,
                    c.periods_per_week,
                    c.description or "",
                ]
                for c in courses
            ]
            zf.writestr("courses.csv", _csv_content(course_headers, course_rows))

            # Attendance
            attendance = db.query(Attendance).filter(Attendance.deleted_at.is_(None)).all()
            attendance_headers = get_header_row("attendance", lang)
            attendance_rows = [
                [
                    r.id,
                    r.student_id,
                    r.course_id,
                    format_date_value(r.date, lang),
                    translate_status_value(r.status, lang),
                    r.period_number,
                    r.notes or "",
                ]
                for r in attendance
            ]
            zf.writestr("attendance.csv", _csv_content(attendance_headers, attendance_rows))

            # Attendance analytics (summary)
            analytics_rows = (
                db.query(
                    Attendance.id,
                    Attendance.date,
                    Attendance.status,
                    Attendance.period_number,
                    Attendance.course_id,
                    Attendance.student_id,
                    Student.student_id.label("student_code"),
                    Student.first_name,
                    Student.last_name,
                    Course.course_code,
                    Course.course_name,
                )
                .join(Student, Student.id == Attendance.student_id)
                .join(Course, Course.id == Attendance.course_id)
                .filter(Attendance.deleted_at.is_(None))
                .all()
            )
            analytics_csv_rows = _build_attendance_analytics_csv_rows(analytics_rows, lang, na_value)
            zf.writestr("attendance_analytics.csv", _csv_content([], analytics_csv_rows))

            # Enrollments
            enrollments = db.query(CourseEnrollment).filter(CourseEnrollment.deleted_at.is_(None)).all()
            enrollment_headers = get_header_row("enrollments", lang)
            enrollment_rows = []
            for e in enrollments:
                student = db.query(Student).filter(Student.id == e.student_id, Student.deleted_at.is_(None)).first()
                course = db.query(Course).filter(Course.id == e.course_id, Course.deleted_at.is_(None)).first()
                enrollment_rows.append(
                    [
                        e.id,
                        e.student_id,
                        f"{student.first_name} {student.last_name}" if student else na_value,
                        e.course_id,
                        course.course_code if course else na_value,
                        course.course_name if course else na_value,
                        format_date_value(e.enrolled_at, lang),
                    ]
                )
            zf.writestr("enrollments.csv", _csv_content(enrollment_headers, enrollment_rows))

            # All grades
            grades = db.query(Grade).filter(Grade.deleted_at.is_(None)).all()
            grade_headers = get_header_row("all_grades", lang)
            grade_rows = []
            for g in grades:
                student = db.query(Student).filter(Student.id == g.student_id, Student.deleted_at.is_(None)).first()
                course = db.query(Course).filter(Course.id == g.course_id, Course.deleted_at.is_(None)).first()
                pct = (g.grade / g.max_grade) * 100 if g.max_grade else 0
                grade_rows.append(
                    [
                        g.id,
                        g.student_id,
                        f"{student.first_name} {student.last_name}" if student else na_value,
                        g.course_id,
                        course.course_name if course else na_value,
                        translate_assignment_name(g.assignment_name, lang),
                        translate_grade_category(g.category or na_value, lang),
                        g.grade,
                        g.max_grade,
                        f"{pct:.2f}%",
                        g.weight,
                        format_date_value(g.date_submitted, lang) if g.date_submitted else na_value,
                    ]
                )
            zf.writestr("all_grades.csv", _csv_content(grade_headers, grade_rows))

            # Daily performance
            performances = db.query(DailyPerformance).filter(DailyPerformance.deleted_at.is_(None)).all()
            performance_headers = get_header_row("daily_performance", lang)
            performance_rows = []
            for p in performances:
                student = db.query(Student).filter(Student.id == p.student_id, Student.deleted_at.is_(None)).first()
                course = db.query(Course).filter(Course.id == p.course_id, Course.deleted_at.is_(None)).first()
                performance_rows.append(
                    [
                        p.id,
                        p.student_id,
                        f"{student.first_name} {student.last_name}" if student else na_value,
                        p.course_id,
                        course.course_name if course else na_value,
                        format_date_value(p.date, lang),
                        translate_grade_category(p.category or na_value, lang),
                        p.score,
                        p.max_score,
                        f"{p.percentage:.2f}%",
                        p.notes or "",
                    ]
                )
            zf.writestr("daily_performance.csv", _csv_content(performance_headers, performance_rows))

            # Highlights
            highlights = db.query(Highlight).filter(Highlight.deleted_at.is_(None)).all()
            highlight_headers = get_header_row("highlights", lang)
            highlight_rows = []
            for h in highlights:
                student = db.query(Student).filter(Student.id == h.student_id, Student.deleted_at.is_(None)).first()
                highlight_rows.append(
                    [
                        h.id,
                        h.student_id,
                        f"{student.first_name} {student.last_name}" if student else na_value,
                        h.semester,
                        translate_grade_category(h.category or na_value, lang),
                        h.rating or na_value,
                        h.highlight_text,
                        format_date_value(h.date_created, lang),
                        yes_no(bool(h.is_positive), lang),
                    ]
                )
            zf.writestr("highlights.csv", _csv_content(highlight_headers, highlight_rows))

        buf.seek(0)
        filename = f"all_data_{datetime.now().strftime('%Y%m%d_%H%M%S')}.zip"
        audit.log_from_request(
            request=request,
            action=AuditAction.BULK_EXPORT,
            resource=AuditResource.SYSTEM,
            details={"format": "zip", "filename": filename},
            success=True,
        )
        return StreamingResponse(
            buf,
            media_type="application/zip",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )
    except Exception as exc:
        logger.error("Export all zip failed: %s", exc, exc_info=True)
        audit.log_from_request(
            request=request,
            action=AuditAction.BULK_EXPORT,
            resource=AuditResource.SYSTEM,
            details={"error": str(exc), "format": "zip"},
            success=False,
        )
        raise http_error(500, ErrorCode.EXPORT_FAILED, "Export failed", request, context={"error": str(exc)})


@router.get("/students/excel")
@require_permission("exports:generate")
async def export_students_excel(
    request: Request,
    limit: int = Query(100, ge=1, le=10000, description="Max records to export (default 100, max 10000)"),
    db: Session = Depends(get_db),
):
    audit = AuditLogger(db)
    try:
        (Student,) = import_names("models", "Student")
        lang = get_lang(request)

        # Query with pagination to reduce memory overhead
        students = (
            db.query(Student)
            .filter(Student.deleted_at.is_(None))
            .order_by(Student.last_name, Student.first_name)
            .limit(limit)
            .all()
        )

        wb = openpyxl.Workbook()
        ws: Any = wb.active
        ws.title = t("sheet_students", lang)
        headers = get_header_row("students", lang)
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        for row, s in enumerate(students, 2):
            ws.cell(row=row, column=1, value=s.id)
            ws.cell(row=row, column=2, value=s.first_name)
            ws.cell(row=row, column=3, value=s.last_name)
            ws.cell(row=row, column=4, value=s.email)
            ws.cell(row=row, column=5, value=s.student_id)
            ws.cell(row=row, column=6, value=format_date_value(s.enrollment_date, lang))
            ws.cell(row=row, column=7, value=t("status_active", lang) if s.is_active else t("status_inactive", lang))
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 18
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        filename = f"students_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        # Log successful export
        audit.log_from_request(
            request=request,
            action=AuditAction.BULK_EXPORT,
            resource=AuditResource.STUDENT,
            details={"count": len(students), "requested_limit": limit, "format": "excel", "filename": filename},
            success=True,
        )
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )
    except Exception as exc:
        logger.error("Export students excel failed: %s", exc, exc_info=True)
        # Log failed export
        audit.log_from_request(
            request=request,
            action=AuditAction.BULK_EXPORT,
            resource=AuditResource.STUDENT,
            details={"error": str(exc), "format": "excel"},
            success=False,
        )
        raise http_error(500, ErrorCode.EXPORT_FAILED, "Export failed", request)


@router.get("/students/csv")
@require_permission("exports:generate")
async def export_students_csv(
    request: Request,
    limit: int = Query(100, ge=1, le=10000, description="Max records to export (default 100, max 10000)"),
    db: Session = Depends(get_db),
):
    audit = AuditLogger(db)
    try:
        (Student,) = import_names("models", "Student")
        lang = get_lang(request)
        students = (
            db.query(Student)
            .filter(Student.deleted_at.is_(None))
            .order_by(Student.last_name, Student.first_name)
            .limit(limit)
            .all()
        )
        headers = get_header_row("students", lang)
        rows = [
            [
                s.id,
                s.first_name,
                s.last_name,
                s.email,
                s.student_id,
                format_date_value(s.enrollment_date, lang),
                t("status_active", lang) if s.is_active else t("status_inactive", lang),
            ]
            for s in students
        ]
        filename = f"students_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        audit.log_from_request(
            request=request,
            action=AuditAction.BULK_EXPORT,
            resource=AuditResource.STUDENT,
            details={"count": len(students), "requested_limit": limit, "format": "csv", "filename": filename},
            success=True,
        )
        return _csv_response(headers, rows, filename)
    except Exception as exc:
        logger.error("Export students csv failed: %s", exc, exc_info=True)
        audit.log_from_request(
            request=request,
            action=AuditAction.BULK_EXPORT,
            resource=AuditResource.STUDENT,
            details={"error": str(exc), "format": "csv"},
            success=False,
        )
        raise http_error(500, ErrorCode.EXPORT_FAILED, "Export failed", request)


@router.get("/grades/excel/{student_id}")
@require_permission("exports:generate")
async def export_student_grades_excel(student_id: int, request: Request, db: Session = Depends(get_db)):
    try:
        Student, Grade = import_names("models", "Student", "Grade")
        lang = get_lang(request)

        student = db.query(Student).filter(Student.id == student_id, Student.deleted_at.is_(None)).first()
        if not student:
            raise http_error(
                404,
                ErrorCode.STUDENT_NOT_FOUND,
                "Student not found",
                request,
                context={"student_id": student_id},
            )
        grades = db.query(Grade).filter(Grade.student_id == student_id, Grade.deleted_at.is_(None)).all()

        wb = openpyxl.Workbook()
        ws: Any = wb.active
        ws.title = t("sheet_grades", lang)
        ws.merge_cells("A1:H1")
        title = ws["A1"]
        student_name = f"{student.first_name} {student.last_name}".strip()
        title.value = grade_report_heading(student_name, student.student_id, lang)
        title.font = Font(size=16, bold=True)
        title.alignment = Alignment(horizontal="center")
        headers = get_header_row("student_grades", lang)
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
        total_percentage = 0
        for row, g in enumerate(grades, 4):
            pct = (g.grade / g.max_grade) * 100 if g.max_grade else 0
            total_percentage += pct
            ws.cell(row=row, column=1, value=translate_assignment_name(g.assignment_name, lang))
            ws.cell(row=row, column=2, value=translate_grade_category(g.category or not_available(lang), lang))
            ws.cell(row=row, column=3, value=g.grade)
            ws.cell(row=row, column=4, value=g.max_grade)
            ws.cell(row=row, column=5, value=f"{pct:.2f}%")
            ws.cell(row=row, column=6, value=g.weight)
            ws.cell(row=row, column=7, value=_letter_grade(pct))
            ws.cell(
                row=row,
                column=8,
                value=format_date_value(g.date_submitted, lang) if g.date_submitted else not_available(lang),
            )
        for col in range(1, 9):
            ws.column_dimensions[get_column_letter(col)].width = 18
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        filename = f"grades_{student.student_id}_{datetime.now().strftime('%Y%m%d')}.xlsx"
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )
    except HTTPException:
        raise
    except Exception as exc:
        logger.error("Export grades excel failed: %s", exc, exc_info=True)
        raise http_error(
            500,
            ErrorCode.EXPORT_FAILED,
            "Export failed",
            request,
            context={"error": str(exc)},
        )


@router.get("/attendance/excel/{student_id}")
@require_permission("exports:generate")
async def export_student_attendance_excel(student_id: int, request: Request, db: Session = Depends(get_db)):
    try:
        Student, Attendance = import_names("models", "Student", "Attendance")
        lang = get_lang(request)

        student = db.query(Student).filter(Student.id == student_id, Student.deleted_at.is_(None)).first()
        if not student:
            raise http_error(
                404,
                ErrorCode.STUDENT_NOT_FOUND,
                "Student not found",
                request,
                context={"student_id": student_id},
            )

        records = (
            db.query(Attendance).filter(Attendance.student_id == student_id, Attendance.deleted_at.is_(None)).all()
        )

        wb = openpyxl.Workbook()
        ws: Any = wb.active
        ws.title = t("sheet_attendance", lang)
        headers = get_header_row("attendance", lang)
        _apply_table_header(ws, headers)
        for row, r in enumerate(records, 2):
            ws.cell(row=row, column=1, value=r.id)
            ws.cell(row=row, column=2, value=r.student_id)
            ws.cell(row=row, column=3, value=r.course_id)
            ws.cell(row=row, column=4, value=format_date_value(r.date, lang))
            ws.cell(row=row, column=5, value=translate_status_value(r.status, lang))
            ws.cell(row=row, column=6, value=r.period_number)
            ws.cell(row=row, column=7, value=r.notes or "")
        _auto_fit_columns(ws)

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        filename = f"attendance_{student.student_id}_{datetime.now().strftime('%Y%m%d')}.xlsx"
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )
    except HTTPException:
        raise
    except Exception as exc:
        logger.error("Export student attendance excel failed: %s", exc, exc_info=True)
        raise http_error(
            500,
            ErrorCode.EXPORT_FAILED,
            "Export failed",
            request,
            context={"error": str(exc)},
        )


@router.get("/performance/excel/{student_id}")
@require_permission("exports:generate")
async def export_student_performance_excel(student_id: int, request: Request, db: Session = Depends(get_db)):
    try:
        Student, Course, DailyPerformance = import_names("models", "Student", "Course", "DailyPerformance")
        lang = get_lang(request)
        na_value = not_available(lang)

        student = db.query(Student).filter(Student.id == student_id, Student.deleted_at.is_(None)).first()
        if not student:
            raise http_error(
                404,
                ErrorCode.STUDENT_NOT_FOUND,
                "Student not found",
                request,
                context={"student_id": student_id},
            )

        records = (
            db.query(DailyPerformance)
            .filter(DailyPerformance.student_id == student_id, DailyPerformance.deleted_at.is_(None))
            .all()
        )

        wb = openpyxl.Workbook()
        ws: Any = wb.active
        ws.title = t("sheet_daily_performance", lang)
        headers = get_header_row("daily_performance", lang)
        _apply_table_header(ws, headers)
        for row, r in enumerate(records, 2):
            course = db.query(Course).filter(Course.id == r.course_id, Course.deleted_at.is_(None)).first()
            percentage = r.percentage
            if percentage is None:
                percentage = (r.score / r.max_score) * 100 if r.max_score else 0
            ws.cell(row=row, column=1, value=r.id)
            ws.cell(row=row, column=2, value=r.student_id)
            ws.cell(row=row, column=3, value=f"{student.first_name} {student.last_name}" if student else na_value)
            ws.cell(row=row, column=4, value=r.course_id)
            ws.cell(row=row, column=5, value=course.course_name if course else na_value)
            ws.cell(row=row, column=6, value=format_date_value(r.date, lang))
            ws.cell(row=row, column=7, value=translate_grade_category(r.category or na_value, lang))
            ws.cell(row=row, column=8, value=r.score)
            ws.cell(row=row, column=9, value=r.max_score)
            ws.cell(row=row, column=10, value=f"{percentage:.2f}%")
            ws.cell(row=row, column=11, value=r.notes or "")
        _auto_fit_columns(ws)

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        filename = f"performance_{student.student_id}_{datetime.now().strftime('%Y%m%d')}.xlsx"
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )
    except HTTPException:
        raise
    except Exception as exc:
        logger.error("Export student performance excel failed: %s", exc, exc_info=True)
        raise http_error(
            500,
            ErrorCode.EXPORT_FAILED,
            "Export failed",
            request,
            context={"error": str(exc)},
        )


@router.get("/highlights/excel/{student_id}")
@require_permission("exports:generate")
async def export_student_highlights_excel(student_id: int, request: Request, db: Session = Depends(get_db)):
    try:
        Student, Highlight = import_names("models", "Student", "Highlight")
        lang = get_lang(request)
        na_value = not_available(lang)

        student = db.query(Student).filter(Student.id == student_id, Student.deleted_at.is_(None)).first()
        if not student:
            raise http_error(
                404,
                ErrorCode.STUDENT_NOT_FOUND,
                "Student not found",
                request,
                context={"student_id": student_id},
            )

        records = db.query(Highlight).filter(Highlight.student_id == student_id, Highlight.deleted_at.is_(None)).all()

        wb = openpyxl.Workbook()
        ws: Any = wb.active
        ws.title = t("sheet_highlights", lang)
        headers = get_header_row("highlights", lang)
        _apply_table_header(ws, headers)
        for row, h in enumerate(records, 2):
            ws.cell(row=row, column=1, value=h.id)
            ws.cell(row=row, column=2, value=h.student_id)
            ws.cell(row=row, column=3, value=f"{student.first_name} {student.last_name}" if student else na_value)
            ws.cell(row=row, column=4, value=h.semester)
            ws.cell(row=row, column=5, value=translate_grade_category(h.category or na_value, lang))
            ws.cell(row=row, column=6, value=h.rating or na_value)
            ws.cell(row=row, column=7, value=h.highlight_text)
            ws.cell(row=row, column=8, value=format_date_value(h.date_created, lang))
            ws.cell(row=row, column=9, value=yes_no(bool(h.is_positive), lang))
        _auto_fit_columns(ws)

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        filename = f"highlights_{student.student_id}_{datetime.now().strftime('%Y%m%d')}.xlsx"
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )
    except HTTPException:
        raise
    except Exception as exc:
        logger.error("Export student highlights excel failed: %s", exc, exc_info=True)
        raise http_error(
            500,
            ErrorCode.EXPORT_FAILED,
            "Export failed",
            request,
            context={"error": str(exc)},
        )


@router.get("/enrollments/excel/{student_id}")
@require_permission("exports:generate")
async def export_student_enrollments_excel(student_id: int, request: Request, db: Session = Depends(get_db)):
    try:
        Student, Course, CourseEnrollment = import_names("models", "Student", "Course", "CourseEnrollment")
        lang = get_lang(request)
        na_value = not_available(lang)

        student = db.query(Student).filter(Student.id == student_id, Student.deleted_at.is_(None)).first()
        if not student:
            raise http_error(
                404,
                ErrorCode.STUDENT_NOT_FOUND,
                "Student not found",
                request,
                context={"student_id": student_id},
            )

        enrollments = (
            db.query(CourseEnrollment)
            .filter(CourseEnrollment.student_id == student_id, CourseEnrollment.deleted_at.is_(None))
            .all()
        )

        wb = openpyxl.Workbook()
        ws: Any = wb.active
        ws.title = t("sheet_enrollments", lang)
        headers = get_header_row("enrollments", lang)
        _apply_table_header(ws, headers)
        for row, e in enumerate(enrollments, 2):
            course = db.query(Course).filter(Course.id == e.course_id, Course.deleted_at.is_(None)).first()
            ws.cell(row=row, column=1, value=e.id)
            ws.cell(row=row, column=2, value=e.student_id)
            ws.cell(row=row, column=3, value=f"{student.first_name} {student.last_name}" if student else na_value)
            ws.cell(row=row, column=4, value=e.course_id)
            ws.cell(row=row, column=5, value=course.course_code if course else na_value)
            ws.cell(row=row, column=6, value=course.course_name if course else na_value)
            ws.cell(row=row, column=7, value=format_date_value(e.enrolled_at, lang))
        _auto_fit_columns(ws)

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        filename = f"enrollments_{student.student_id}_{datetime.now().strftime('%Y%m%d')}.xlsx"
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )
    except HTTPException:
        raise
    except Exception as exc:
        logger.error("Export student enrollments excel failed: %s", exc, exc_info=True)
        raise http_error(
            500,
            ErrorCode.EXPORT_FAILED,
            "Export failed",
            request,
            context={"error": str(exc)},
        )


def _letter_grade(percentage: float) -> str:
    """
    Convert a percentage grade to a letter grade using standard academic scale.

    Args:
        percentage: Numeric grade as a percentage (0-100)

    Returns:
        Letter grade: A+ (97-100), A (93-96), A- (90-92), B+ (87-89), B (83-86),
                     B- (80-82), C+ (77-79), C (70-76), D (60-69), F (0-59)
    """
    if percentage >= 97:
        return "A+"
    if percentage >= 93:
        return "A"
    if percentage >= 90:
        return "A-"
    if percentage >= 87:
        return "B+"
    if percentage >= 83:
        return "B"
    if percentage >= 80:
        return "B-"
    if percentage >= 77:
        return "C+"
    if percentage >= 70:
        return "C"
    if percentage >= 60:
        return "D"
    return "F"


@router.get("/students/pdf")
@require_permission("exports:generate")
async def export_students_pdf(request: Request, db: Session = Depends(get_db)):
    try:
        (Student,) = import_names("models", "Student")
        lang = get_lang(request)
        students = db.query(Student).filter(Student.deleted_at.is_(None)).all()
        headers = [
            t("header_id", lang),
            t("label_student_name", lang),
            t("label_student_id", lang),
            t("label_email", lang),
            t("label_enrollment_date", lang),
            t("label_status", lang),
        ]
        rows = [
            [
                str(s.id),
                f"{s.first_name} {s.last_name}",
                s.student_id,
                s.email,
                format_date_value(s.enrollment_date, lang),
                t("status_active", lang) if s.is_active else t("status_inactive", lang),
            ]
            for s in students
        ]
        filename = f"students_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        return _pdf_table_response(t("student_directory_title", lang), headers, rows, filename)
    except Exception as exc:
        logger.error("Export students pdf failed: %s", exc, exc_info=True)
        raise http_error(
            500,
            ErrorCode.EXPORT_FAILED,
            "Export failed",
            request,
            context={"error": str(exc)},
        )


@router.get("/attendance/excel")
@require_permission("exports:generate")
async def export_attendance_excel(request: Request, db: Session = Depends(get_db)):
    audit = AuditLogger(db)
    try:
        (Attendance,) = import_names("models", "Attendance")
        lang = get_lang(request)

        records = db.query(Attendance).filter(Attendance.deleted_at.is_(None)).all()
        wb = openpyxl.Workbook()
        ws: Any = wb.active
        ws.title = t("sheet_attendance", lang)
        headers = get_header_row("attendance", lang)
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
        for row, r in enumerate(records, 2):
            ws.cell(row=row, column=1, value=r.id)
            ws.cell(row=row, column=2, value=r.student_id)
            ws.cell(row=row, column=3, value=r.course_id)
            ws.cell(row=row, column=4, value=format_date_value(r.date, lang))
            ws.cell(row=row, column=5, value=translate_status_value(r.status, lang))
            ws.cell(row=row, column=6, value=r.period_number)
            ws.cell(row=row, column=7, value=r.notes or "")
        for col in range(1, 8):
            ws.column_dimensions[get_column_letter(col)].width = 15
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        filename = f"attendance_{datetime.now().strftime('%Y%m%d')}.xlsx"
        # Log successful export
        audit.log_from_request(
            request=request,
            action=AuditAction.BULK_EXPORT,
            resource=AuditResource.ATTENDANCE,
            details={"count": len(records), "format": "excel", "filename": filename},
            success=True,
        )
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )
    except Exception as exc:
        logger.error("Export attendance excel failed: %s", exc, exc_info=True)
        # Log failed export
        audit.log_from_request(
            request=request,
            action=AuditAction.BULK_EXPORT,
            resource=AuditResource.ATTENDANCE,
            details={"error": str(exc), "format": "excel"},
            success=False,
        )
        raise http_error(
            500,
            ErrorCode.EXPORT_FAILED,
            "Export failed",
            request,
            context={"error": str(exc)},
        )


@router.get("/attendance/csv")
@require_permission("exports:generate")
async def export_attendance_csv(request: Request, db: Session = Depends(get_db)):
    audit = AuditLogger(db)
    try:
        (Attendance,) = import_names("models", "Attendance")
        lang = get_lang(request)
        records = db.query(Attendance).filter(Attendance.deleted_at.is_(None)).all()
        headers = get_header_row("attendance", lang)
        rows = [
            [
                r.id,
                r.student_id,
                r.course_id,
                format_date_value(r.date, lang),
                translate_status_value(r.status, lang),
                r.period_number,
                r.notes or "",
            ]
            for r in records
        ]
        filename = f"attendance_{datetime.now().strftime('%Y%m%d')}.csv"
        audit.log_from_request(
            request=request,
            action=AuditAction.BULK_EXPORT,
            resource=AuditResource.ATTENDANCE,
            details={"count": len(records), "format": "csv", "filename": filename},
            success=True,
        )
        return _csv_response(headers, rows, filename)
    except Exception as exc:
        logger.error("Export attendance csv failed: %s", exc, exc_info=True)
        audit.log_from_request(
            request=request,
            action=AuditAction.BULK_EXPORT,
            resource=AuditResource.ATTENDANCE,
            details={"error": str(exc), "format": "csv"},
            success=False,
        )
        raise http_error(500, ErrorCode.EXPORT_FAILED, "Export failed", request, context={"error": str(exc)})


@router.get("/attendance/pdf")
@require_permission("exports:generate")
async def export_attendance_pdf(request: Request, db: Session = Depends(get_db)):
    audit = AuditLogger(db)
    try:
        (Attendance,) = import_names("models", "Attendance")
        lang = get_lang(request)
        records = db.query(Attendance).filter(Attendance.deleted_at.is_(None)).all()
        headers = get_header_row("attendance", lang)
        rows = [
            [
                str(r.id),
                str(r.student_id),
                str(r.course_id),
                format_date_value(r.date, lang),
                translate_status_value(r.status, lang),
                str(r.period_number),
                r.notes or "",
            ]
            for r in records
        ]
        filename = f"attendance_{datetime.now().strftime('%Y%m%d')}.pdf"
        audit.log_from_request(
            request=request,
            action=AuditAction.BULK_EXPORT,
            resource=AuditResource.ATTENDANCE,
            details={"count": len(records), "format": "pdf", "filename": filename},
            success=True,
        )
        return _pdf_table_response(t("sheet_attendance", lang), headers, rows, filename)
    except Exception as exc:
        logger.error("Export attendance pdf failed: %s", exc, exc_info=True)
        audit.log_from_request(
            request=request,
            action=AuditAction.BULK_EXPORT,
            resource=AuditResource.ATTENDANCE,
            details={"error": str(exc), "format": "pdf"},
            success=False,
        )
        raise http_error(500, ErrorCode.EXPORT_FAILED, "Export failed", request, context={"error": str(exc)})


@router.get("/attendance/analytics/excel")
@require_permission("exports:generate")
async def export_attendance_analytics_excel(request: Request, db: Session = Depends(get_db)):
    try:
        Attendance, Student, Course = import_names("models", "Attendance", "Student", "Course")
        lang = get_lang(request)
        na_value = not_available(lang)

        rows = (
            db.query(
                Attendance.id,
                Attendance.date,
                Attendance.status,
                Attendance.period_number,
                Attendance.course_id,
                Attendance.student_id,
                Student.student_id.label("student_code"),
                Student.first_name,
                Student.last_name,
                Course.course_code,
                Course.course_name,
            )
            .join(Student, Student.id == Attendance.student_id)
            .join(Course, Course.id == Attendance.course_id)
            .filter(Attendance.deleted_at.is_(None))
            .all()
        )

        overall_counts = _init_status_counts()
        course_summary: dict[int, dict[str, Any]] = {}
        period_summary: dict[int, dict[str, Any]] = {}
        course_period_summary: dict[tuple[int, int], dict[str, Any]] = {}
        student_summary: dict[int, dict[str, Any]] = {}
        daily_summary: dict[Any, dict[str, Any]] = {}
        unique_students = set()
        unique_courses = set()
        date_values = set()

        for (
            _attendance_id,
            att_date,
            status,
            period_number,
            course_id,
            student_id,
            student_code,
            first_name,
            last_name,
            course_code,
            course_name,
        ) in rows:
            safe_status = _normalize_status(status)
            period = period_number or 1
            student_label = (f"{first_name or ''} {last_name or ''}").strip() or na_value
            course_label = course_code or na_value
            course_title = course_name or na_value

            overall_counts[safe_status] += 1
            unique_students.add(student_id)
            unique_courses.add(course_id)
            if att_date:
                date_values.add(att_date)

            course_entry = course_summary.setdefault(
                course_id,
                {"course_code": course_label, "course_name": course_title, "counts": _init_status_counts(), "total": 0},
            )
            course_entry["counts"][safe_status] += 1
            course_entry["total"] += 1

            period_entry = period_summary.setdefault(
                period,
                {"counts": _init_status_counts(), "total": 0},
            )
            period_entry["counts"][safe_status] += 1
            period_entry["total"] += 1

            course_period_entry = course_period_summary.setdefault(
                (course_id, period),
                {
                    "course_code": course_label,
                    "course_name": course_title,
                    "period": period,
                    "counts": _init_status_counts(),
                    "total": 0,
                },
            )
            course_period_entry["counts"][safe_status] += 1
            course_period_entry["total"] += 1

            student_entry = student_summary.setdefault(
                student_id,
                {
                    "student_code": student_code or str(student_id),
                    "student_name": student_label,
                    "counts": _init_status_counts(),
                    "total": 0,
                },
            )
            student_entry["counts"][safe_status] += 1
            student_entry["total"] += 1

            if att_date:
                daily_entry = daily_summary.setdefault(att_date, {"counts": _init_status_counts(), "total": 0})
                daily_entry["counts"][safe_status] += 1
                daily_entry["total"] += 1

        total_records = sum(overall_counts.values())
        present_share = (overall_counts["Present"] / total_records * 100) if total_records else 0
        date_range = na_value
        if date_values:
            sorted_dates = sorted(date_values)
            date_range = f"{format_date_value(sorted_dates[0], lang)} → {format_date_value(sorted_dates[-1], lang)}"

        wb = openpyxl.Workbook()
        overview_ws: Any = wb.active
        overview_ws.title = t("sheet_overview", lang)
        overview_ws["A1"] = t("title_attendance_export", lang)
        overview_ws["A1"].font = Font(size=16, bold=True)
        overview_ws.append(get_header_row("overview", lang))
        overview_ws.append([t("label_total_records", lang), total_records])
        overview_ws.append([t("label_unique_students", lang), len(unique_students)])
        overview_ws.append([t("label_unique_courses", lang), len(unique_courses)])
        overview_ws.append([t("label_date_range", lang), date_range])
        overview_ws.append([t("label_present_share", lang), f"{present_share:.1f}%"])
        overview_ws.append(
            [t("label_generated_on_table", lang), format_date_value(datetime.now(), lang, include_time=True)]
        )
        overview_ws.append([])
        overview_ws.append(get_header_row("status_counts", lang))
        for status in ATTENDANCE_STATUSES:
            overview_ws.append([t(status.lower(), lang), overall_counts[status]])
        if total_records == 0:
            overview_ws.append([])
            overview_ws.append([t("label_notice", lang), t("label_no_attendance", lang)])
        _auto_fit_columns(overview_ws)

        course_ws = wb.create_sheet(t("sheet_course_summary", lang))
        course_headers = get_header_row("course_summary", lang)
        _apply_table_header(course_ws, course_headers)
        for row_idx, data in enumerate(sorted(course_summary.values(), key=lambda item: item["course_code"]), start=2):
            counts = data["counts"]
            total = data["total"]
            rate = (counts["Present"] / total * 100) if total else 0
            values = [
                data["course_code"],
                data["course_name"],
                total,
                counts["Present"],
                counts["Absent"],
                counts["Late"],
                counts["Excused"],
                f"{rate:.1f}%",
            ]
            for col, value in enumerate(values, 1):
                course_ws.cell(row=row_idx, column=col, value=value)
        _auto_fit_columns(course_ws)

        period_ws = wb.create_sheet(t("sheet_period_summary", lang))
        period_headers = get_header_row("period_summary", lang)
        _apply_table_header(period_ws, period_headers)
        for row_idx, period in enumerate(sorted(period_summary.keys()), start=2):
            counts = period_summary[period]["counts"]
            total = period_summary[period]["total"]
            rate = (counts["Present"] / total * 100) if total else 0
            values = [
                period,
                total,
                counts["Present"],
                counts["Absent"],
                counts["Late"],
                counts["Excused"],
                f"{rate:.1f}%",
            ]
            for col, value in enumerate(values, 1):
                period_ws.cell(row=row_idx, column=col, value=value)
        _auto_fit_columns(period_ws)

        course_period_ws = wb.create_sheet(t("sheet_course_periods", lang))
        course_period_headers = get_header_row("course_periods", lang)
        _apply_table_header(course_period_ws, course_period_headers)
        for row_idx, key in enumerate(
            sorted(
                course_period_summary.keys(), key=lambda item: (course_period_summary[item]["course_code"], item[1])
            ),
            start=2,
        ):
            data = course_period_summary[key]
            counts = data["counts"]
            total = data["total"]
            rate = (counts["Present"] / total * 100) if total else 0
            values = [
                data["course_code"],
                data["course_name"],
                data["period"],
                total,
                counts["Present"],
                counts["Absent"],
                counts["Late"],
                counts["Excused"],
                f"{rate:.1f}%",
            ]
            for col, value in enumerate(values, 1):
                course_period_ws.cell(row=row_idx, column=col, value=value)
        _auto_fit_columns(course_period_ws)

        student_ws = wb.create_sheet(t("sheet_student_summary", lang))
        student_headers = get_header_row("student_summary", lang)
        _apply_table_header(student_ws, student_headers)
        for row_idx, data in enumerate(
            sorted(student_summary.values(), key=lambda item: item["student_name"]), start=2
        ):
            counts = data["counts"]
            most_common = _dominant_status(counts)
            values = [
                data["student_code"],
                data["student_name"],
                data["total"],
                counts["Present"],
                counts["Absent"],
                counts["Late"],
                counts["Excused"],
                translate_status_value(most_common, lang),
            ]
            for col, value in enumerate(values, 1):
                student_ws.cell(row=row_idx, column=col, value=value)
        _auto_fit_columns(student_ws)

        daily_ws = wb.create_sheet(t("sheet_daily_overview", lang))
        daily_headers = get_header_row("daily_overview", lang)
        _apply_table_header(daily_ws, daily_headers)
        for row_idx, day in enumerate(sorted(daily_summary.keys()), start=2):
            counts = daily_summary[day]["counts"]
            values = [
                format_date_value(day, lang),
                daily_summary[day]["total"],
                counts["Present"],
                counts["Absent"],
                counts["Late"],
                counts["Excused"],
            ]
            for col, value in enumerate(values, 1):
                daily_ws.cell(row=row_idx, column=col, value=value)
        _auto_fit_columns(daily_ws)

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        filename = f"attendance_analytics_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )
    except Exception as exc:
        logger.error("Export attendance analytics excel failed: %s", exc, exc_info=True)
        raise http_error(
            500,
            ErrorCode.EXPORT_FAILED,
            "Export failed",
            request,
            context={"error": str(exc)},
        )


@router.get("/attendance/analytics/csv")
@require_permission("exports:generate")
async def export_attendance_analytics_csv(request: Request, db: Session = Depends(get_db)):
    try:
        Attendance, Student, Course = import_names("models", "Attendance", "Student", "Course")
        lang = get_lang(request)
        na_value = not_available(lang)

        rows = (
            db.query(
                Attendance.id,
                Attendance.date,
                Attendance.status,
                Attendance.period_number,
                Attendance.course_id,
                Attendance.student_id,
                Student.student_id.label("student_code"),
                Student.first_name,
                Student.last_name,
                Course.course_code,
                Course.course_name,
            )
            .join(Student, Student.id == Attendance.student_id)
            .join(Course, Course.id == Attendance.course_id)
            .filter(Attendance.deleted_at.is_(None))
            .all()
        )

        flat_rows = _build_attendance_analytics_csv_rows(rows, lang, na_value)

        filename = f"attendance_analytics_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        headers = []
        return _csv_response(headers, flat_rows, filename)
    except Exception as exc:
        logger.error("Export attendance analytics csv failed: %s", exc, exc_info=True)
        raise http_error(500, ErrorCode.EXPORT_FAILED, "Export failed", request, context={"error": str(exc)})


@router.get("/attendance/analytics/pdf")
@require_permission("exports:generate")
async def export_attendance_analytics_pdf(request: Request, db: Session = Depends(get_db)):
    try:
        Attendance, Student, Course = import_names("models", "Attendance", "Student", "Course")
        lang = get_lang(request)
        na_value = not_available(lang)

        rows = (
            db.query(
                Attendance.id,
                Attendance.date,
                Attendance.status,
                Attendance.period_number,
                Attendance.course_id,
                Attendance.student_id,
                Student.student_id.label("student_code"),
                Student.first_name,
                Student.last_name,
                Course.course_code,
                Course.course_name,
            )
            .join(Student, Student.id == Attendance.student_id)
            .join(Course, Course.id == Attendance.course_id)
            .filter(Attendance.deleted_at.is_(None))
            .all()
        )

        overall_counts = _init_status_counts()
        unique_students = set()
        unique_courses = set()
        date_values = set()

        for (
            _attendance_id,
            att_date,
            status,
            _period_number,
            course_id,
            student_id,
            _student_code,
            _first_name,
            _last_name,
            _course_code,
            _course_name,
        ) in rows:
            safe_status = _normalize_status(status)
            overall_counts[safe_status] += 1
            unique_students.add(student_id)
            unique_courses.add(course_id)
            if att_date:
                date_values.add(att_date)

        total_records = sum(overall_counts.values())
        present_share = (overall_counts["Present"] / total_records * 100) if total_records else 0
        date_range = na_value
        if date_values:
            sorted_dates = sorted(date_values)
            date_range = f"{format_date_value(sorted_dates[0], lang)} → {format_date_value(sorted_dates[-1], lang)}"

        headers = [t("label_metric", lang), t("label_value", lang)]
        rows_data = [
            [t("label_total_records", lang), total_records],
            [t("label_unique_students", lang), len(unique_students)],
            [t("label_unique_courses", lang), len(unique_courses)],
            [t("label_date_range", lang), date_range],
            [t("label_present_share", lang), f"{present_share:.1f}%"],
        ]
        filename = f"attendance_analytics_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        return _pdf_table_response(t("title_attendance_export", lang), headers, rows_data, filename)
    except Exception as exc:
        logger.error("Export attendance analytics pdf failed: %s", exc, exc_info=True)
        raise http_error(500, ErrorCode.EXPORT_FAILED, "Export failed", request, context={"error": str(exc)})


@router.get("/courses/excel")
@require_permission("exports:generate")
async def export_courses_excel(request: Request, db: Session = Depends(get_db)):
    """Export all courses to Excel"""
    audit = AuditLogger(db)
    try:
        (Course,) = import_names("models", "Course")
        lang = get_lang(request)

        courses = db.query(Course).filter(Course.deleted_at.is_(None)).all()

        wb = openpyxl.Workbook()
        ws: Any = wb.active
        ws.title = t("sheet_courses", lang)
        headers = get_header_row("courses", lang)
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        for row, c in enumerate(courses, 2):
            ws.cell(row=row, column=1, value=c.id)
            ws.cell(row=row, column=2, value=c.course_code)
            ws.cell(row=row, column=3, value=c.course_name)
            ws.cell(row=row, column=4, value=c.semester)
            ws.cell(row=row, column=5, value=c.credits)
            ws.cell(row=row, column=6, value=c.hours_per_week)
            ws.cell(row=row, column=7, value=c.periods_per_week)
            ws.cell(row=row, column=8, value=c.description or "")

        for col in range(1, 9):
            ws.column_dimensions[get_column_letter(col)].width = 20

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        filename = f"courses_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        # Log successful export
        audit.log_from_request(
            request=request,
            action=AuditAction.BULK_EXPORT,
            resource=AuditResource.COURSE,
            details={"count": len(courses), "format": "excel", "filename": filename},
            success=True,
        )
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )
    except Exception as exc:
        logger.error("Export courses excel failed: %s", exc, exc_info=True)
        # Log failed export
        audit.log_from_request(
            request=request,
            action=AuditAction.BULK_EXPORT,
            resource=AuditResource.COURSE,
            details={"error": str(exc), "format": "excel"},
            success=False,
        )
        raise http_error(
            500,
            ErrorCode.EXPORT_FAILED,
            "Export failed",
            request,
            context={"error": str(exc)},
        )


@router.get("/courses/csv")
@require_permission("exports:generate")
async def export_courses_csv(request: Request, db: Session = Depends(get_db)):
    audit = AuditLogger(db)
    try:
        (Course,) = import_names("models", "Course")
        lang = get_lang(request)
        courses = db.query(Course).filter(Course.deleted_at.is_(None)).all()
        headers = get_header_row("courses", lang)
        rows = [
            [
                c.id,
                c.course_code,
                c.course_name,
                c.semester,
                c.credits,
                c.hours_per_week,
                c.periods_per_week,
                c.description or "",
            ]
            for c in courses
        ]
        filename = f"courses_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        audit.log_from_request(
            request=request,
            action=AuditAction.BULK_EXPORT,
            resource=AuditResource.COURSE,
            details={"count": len(courses), "format": "csv", "filename": filename},
            success=True,
        )
        return _csv_response(headers, rows, filename)
    except Exception as exc:
        logger.error("Export courses csv failed: %s", exc, exc_info=True)
        audit.log_from_request(
            request=request,
            action=AuditAction.BULK_EXPORT,
            resource=AuditResource.COURSE,
            details={"error": str(exc), "format": "csv"},
            success=False,
        )
        raise http_error(500, ErrorCode.EXPORT_FAILED, "Export failed", request, context={"error": str(exc)})


@router.get("/enrollments/excel")
@require_permission("exports:generate")
async def export_enrollments_excel(request: Request, db: Session = Depends(get_db)):
    """Export all course enrollments to Excel"""
    audit = AuditLogger(db)
    try:
        CourseEnrollment, Student, Course = import_names("models", "CourseEnrollment", "Student", "Course")
        lang = get_lang(request)
        na_value = not_available(lang)

        enrollments = db.query(CourseEnrollment).filter(CourseEnrollment.deleted_at.is_(None)).all()

        wb = openpyxl.Workbook()
        ws: Any = wb.active
        ws.title = t("sheet_enrollments", lang)
        headers = get_header_row("enrollments", lang)
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        for row, e in enumerate(enrollments, 2):
            student = db.query(Student).filter(Student.id == e.student_id, Student.deleted_at.is_(None)).first()
            course = db.query(Course).filter(Course.id == e.course_id, Course.deleted_at.is_(None)).first()

            ws.cell(row=row, column=1, value=e.id)
            ws.cell(row=row, column=2, value=e.student_id)
            ws.cell(row=row, column=3, value=f"{student.first_name} {student.last_name}" if student else na_value)
            ws.cell(row=row, column=4, value=e.course_id)
            ws.cell(row=row, column=5, value=course.course_code if course else na_value)
            ws.cell(row=row, column=6, value=course.course_name if course else na_value)
            ws.cell(row=row, column=7, value=format_date_value(e.enrolled_at, lang))

        for col in range(1, 8):
            ws.column_dimensions[get_column_letter(col)].width = 18

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        filename = f"enrollments_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        # Log successful export
        audit.log_from_request(
            request=request,
            action=AuditAction.BULK_EXPORT,
            resource=AuditResource.ENROLLMENT,
            details={"count": len(enrollments), "format": "excel", "filename": filename},
            success=True,
        )
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )
    except Exception as exc:
        logger.error("Export enrollments excel failed: %s", exc, exc_info=True)
        # Log failed export
        audit.log_from_request(
            request=request,
            action=AuditAction.BULK_EXPORT,
            resource=AuditResource.ENROLLMENT,
            details={"error": str(exc), "format": "excel"},
            success=False,
        )
        raise http_error(
            500,
            ErrorCode.EXPORT_FAILED,
            "Export failed",
            request,
            context={"error": str(exc)},
        )


@router.get("/enrollments/csv")
@require_permission("exports:generate")
async def export_enrollments_csv(request: Request, db: Session = Depends(get_db)):
    audit = AuditLogger(db)
    try:
        CourseEnrollment, Student, Course = import_names("models", "CourseEnrollment", "Student", "Course")
        lang = get_lang(request)
        na_value = not_available(lang)
        enrollments = db.query(CourseEnrollment).filter(CourseEnrollment.deleted_at.is_(None)).all()
        headers = get_header_row("enrollments", lang)
        rows = []
        for e in enrollments:
            student = db.query(Student).filter(Student.id == e.student_id, Student.deleted_at.is_(None)).first()
            course = db.query(Course).filter(Course.id == e.course_id, Course.deleted_at.is_(None)).first()
            rows.append(
                [
                    e.id,
                    e.student_id,
                    f"{student.first_name} {student.last_name}" if student else na_value,
                    e.course_id,
                    course.course_code if course else na_value,
                    course.course_name if course else na_value,
                    format_date_value(e.enrolled_at, lang),
                ]
            )
        filename = f"enrollments_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        audit.log_from_request(
            request=request,
            action=AuditAction.BULK_EXPORT,
            resource=AuditResource.ENROLLMENT,
            details={"count": len(enrollments), "format": "csv", "filename": filename},
            success=True,
        )
        return _csv_response(headers, rows, filename)
    except Exception as exc:
        logger.error("Export enrollments csv failed: %s", exc, exc_info=True)
        audit.log_from_request(
            request=request,
            action=AuditAction.BULK_EXPORT,
            resource=AuditResource.ENROLLMENT,
            details={"error": str(exc), "format": "csv"},
            success=False,
        )
        raise http_error(500, ErrorCode.EXPORT_FAILED, "Export failed", request, context={"error": str(exc)})


@router.get("/enrollments/pdf")
@require_permission("exports:generate")
async def export_enrollments_pdf(request: Request, db: Session = Depends(get_db)):
    audit = AuditLogger(db)
    try:
        CourseEnrollment, Student, Course = import_names("models", "CourseEnrollment", "Student", "Course")
        lang = get_lang(request)
        na_value = not_available(lang)
        enrollments = db.query(CourseEnrollment).filter(CourseEnrollment.deleted_at.is_(None)).all()
        headers = get_header_row("enrollments", lang)
        rows = []
        for e in enrollments:
            student = db.query(Student).filter(Student.id == e.student_id, Student.deleted_at.is_(None)).first()
            course = db.query(Course).filter(Course.id == e.course_id, Course.deleted_at.is_(None)).first()
            rows.append(
                [
                    str(e.id),
                    str(e.student_id),
                    f"{student.first_name} {student.last_name}" if student else na_value,
                    str(e.course_id),
                    course.course_code if course else na_value,
                    course.course_name if course else na_value,
                    format_date_value(e.enrolled_at, lang),
                ]
            )
        filename = f"enrollments_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        audit.log_from_request(
            request=request,
            action=AuditAction.BULK_EXPORT,
            resource=AuditResource.ENROLLMENT,
            details={"count": len(enrollments), "format": "pdf", "filename": filename},
            success=True,
        )
        return _pdf_table_response(t("sheet_enrollments", lang), headers, rows, filename)
    except Exception as exc:
        logger.error("Export enrollments pdf failed: %s", exc, exc_info=True)
        audit.log_from_request(
            request=request,
            action=AuditAction.BULK_EXPORT,
            resource=AuditResource.ENROLLMENT,
            details={"error": str(exc), "format": "pdf"},
            success=False,
        )
        raise http_error(500, ErrorCode.EXPORT_FAILED, "Export failed", request, context={"error": str(exc)})


@router.get("/grades/excel")
@require_permission("exports:generate")
async def export_all_grades_excel(request: Request, db: Session = Depends(get_db)):
    """Export all grades to Excel"""
    audit = AuditLogger(db)
    try:
        Grade, Student, Course = import_names("models", "Grade", "Student", "Course")
        lang = get_lang(request)
        na_value = not_available(lang)

        grades = db.query(Grade).filter(Grade.deleted_at.is_(None)).all()

        wb = openpyxl.Workbook()
        ws: Any = wb.active
        ws.title = t("sheet_all_grades", lang)
        headers = get_header_row("all_grades", lang)
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        for row, g in enumerate(grades, 2):
            student = db.query(Student).filter(Student.id == g.student_id, Student.deleted_at.is_(None)).first()
            course = db.query(Course).filter(Course.id == g.course_id, Course.deleted_at.is_(None)).first()
            pct = (g.grade / g.max_grade) * 100 if g.max_grade else 0

            ws.cell(row=row, column=1, value=g.id)
            ws.cell(row=row, column=2, value=g.student_id)
            ws.cell(row=row, column=3, value=f"{student.first_name} {student.last_name}" if student else na_value)
            ws.cell(row=row, column=4, value=g.course_id)
            ws.cell(row=row, column=5, value=course.course_name if course else na_value)
            ws.cell(row=row, column=6, value=translate_assignment_name(g.assignment_name, lang))
            ws.cell(row=row, column=7, value=translate_grade_category(g.category or na_value, lang))
            ws.cell(row=row, column=8, value=g.grade)
            ws.cell(row=row, column=9, value=g.max_grade)
            ws.cell(row=row, column=10, value=f"{pct:.2f}%")
            ws.cell(row=row, column=11, value=g.weight)
            ws.cell(
                row=row,
                column=12,
                value=format_date_value(g.date_submitted, lang) if g.date_submitted else na_value,
            )

        for col in range(1, 13):
            ws.column_dimensions[get_column_letter(col)].width = 15

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        filename = f"all_grades_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        # Log successful export
        audit.log_from_request(
            request=request,
            action=AuditAction.BULK_EXPORT,
            resource=AuditResource.GRADE,
            details={"count": len(grades), "format": "excel", "filename": filename},
            success=True,
        )
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )
    except Exception as exc:
        logger.error("Export all grades excel failed: %s", exc, exc_info=True)
        # Log failed export
        audit.log_from_request(
            request=request,
            action=AuditAction.BULK_EXPORT,
            resource=AuditResource.GRADE,
            details={"error": str(exc), "format": "excel"},
            success=False,
        )
        raise http_error(
            500,
            ErrorCode.EXPORT_FAILED,
            "Export failed",
            request,
            context={"error": str(exc)},
        )


@router.get("/grades/csv")
@require_permission("exports:generate")
async def export_all_grades_csv(request: Request, db: Session = Depends(get_db)):
    audit = AuditLogger(db)
    try:
        Grade, Student, Course = import_names("models", "Grade", "Student", "Course")
        lang = get_lang(request)
        na_value = not_available(lang)
        grades = db.query(Grade).filter(Grade.deleted_at.is_(None)).all()
        headers = get_header_row("all_grades", lang)
        rows = []
        for g in grades:
            student = db.query(Student).filter(Student.id == g.student_id, Student.deleted_at.is_(None)).first()
            course = db.query(Course).filter(Course.id == g.course_id, Course.deleted_at.is_(None)).first()
            pct = (g.grade / g.max_grade) * 100 if g.max_grade else 0
            rows.append(
                [
                    g.id,
                    g.student_id,
                    f"{student.first_name} {student.last_name}" if student else na_value,
                    g.course_id,
                    course.course_name if course else na_value,
                    translate_assignment_name(g.assignment_name, lang),
                    translate_grade_category(g.category or na_value, lang),
                    g.grade,
                    g.max_grade,
                    f"{pct:.2f}%",
                    g.weight,
                    format_date_value(g.date_submitted, lang) if g.date_submitted else na_value,
                ]
            )
        filename = f"all_grades_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        audit.log_from_request(
            request=request,
            action=AuditAction.BULK_EXPORT,
            resource=AuditResource.GRADE,
            details={"count": len(grades), "format": "csv", "filename": filename},
            success=True,
        )
        return _csv_response(headers, rows, filename)
    except Exception as exc:
        logger.error("Export all grades csv failed: %s", exc, exc_info=True)
        audit.log_from_request(
            request=request,
            action=AuditAction.BULK_EXPORT,
            resource=AuditResource.GRADE,
            details={"error": str(exc), "format": "csv"},
            success=False,
        )
        raise http_error(500, ErrorCode.EXPORT_FAILED, "Export failed", request, context={"error": str(exc)})


@router.get("/grades/pdf")
@require_permission("exports:generate")
async def export_all_grades_pdf(request: Request, db: Session = Depends(get_db)):
    audit = AuditLogger(db)
    try:
        Grade, Student, Course = import_names("models", "Grade", "Student", "Course")
        lang = get_lang(request)
        na_value = not_available(lang)
        grades = db.query(Grade).filter(Grade.deleted_at.is_(None)).all()
        headers = get_header_row("all_grades", lang)
        rows = []
        for g in grades:
            student = db.query(Student).filter(Student.id == g.student_id, Student.deleted_at.is_(None)).first()
            course = db.query(Course).filter(Course.id == g.course_id, Course.deleted_at.is_(None)).first()
            pct = (g.grade / g.max_grade) * 100 if g.max_grade else 0
            rows.append(
                [
                    str(g.id),
                    str(g.student_id),
                    f"{student.first_name} {student.last_name}" if student else na_value,
                    str(g.course_id),
                    course.course_name if course else na_value,
                    translate_assignment_name(g.assignment_name, lang),
                    translate_grade_category(g.category or na_value, lang),
                    str(g.grade),
                    str(g.max_grade),
                    f"{pct:.2f}%",
                    str(g.weight),
                    format_date_value(g.date_submitted, lang) if g.date_submitted else na_value,
                ]
            )
        filename = f"all_grades_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        audit.log_from_request(
            request=request,
            action=AuditAction.BULK_EXPORT,
            resource=AuditResource.GRADE,
            details={"count": len(grades), "format": "pdf", "filename": filename},
            success=True,
        )
        return _pdf_table_response(t("sheet_all_grades", lang), headers, rows, filename)
    except Exception as exc:
        logger.error("Export all grades pdf failed: %s", exc, exc_info=True)
        audit.log_from_request(
            request=request,
            action=AuditAction.BULK_EXPORT,
            resource=AuditResource.GRADE,
            details={"error": str(exc), "format": "pdf"},
            success=False,
        )
        raise http_error(500, ErrorCode.EXPORT_FAILED, "Export failed", request, context={"error": str(exc)})


@router.get("/performance/excel")
@require_permission("exports:generate")
async def export_daily_performance_excel(request: Request, db: Session = Depends(get_db)):
    """Export all daily performance records to Excel"""
    audit = AuditLogger(db)
    try:
        DailyPerformance, Student, Course = import_names("models", "DailyPerformance", "Student", "Course")
        lang = get_lang(request)
        na_value = not_available(lang)

        performances = db.query(DailyPerformance).filter(DailyPerformance.deleted_at.is_(None)).all()

        wb = openpyxl.Workbook()
        ws: Any = wb.active
        ws.title = t("sheet_daily_performance", lang)
        headers = get_header_row("daily_performance", lang)
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        for row, p in enumerate(performances, 2):
            student = db.query(Student).filter(Student.id == p.student_id, Student.deleted_at.is_(None)).first()
            course = db.query(Course).filter(Course.id == p.course_id, Course.deleted_at.is_(None)).first()

            ws.cell(row=row, column=1, value=p.id)
            ws.cell(row=row, column=2, value=p.student_id)
            ws.cell(row=row, column=3, value=f"{student.first_name} {student.last_name}" if student else na_value)
            ws.cell(row=row, column=4, value=p.course_id)
            ws.cell(row=row, column=5, value=course.course_name if course else na_value)
            ws.cell(row=row, column=6, value=format_date_value(p.date, lang))
            ws.cell(row=row, column=7, value=translate_grade_category(p.category or na_value, lang))
            ws.cell(row=row, column=8, value=p.score)
            ws.cell(row=row, column=9, value=p.max_score)
            ws.cell(row=row, column=10, value=f"{p.percentage:.2f}%")
            ws.cell(row=row, column=11, value=p.notes or "")

        for col in range(1, 12):
            ws.column_dimensions[get_column_letter(col)].width = 15

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        filename = f"daily_performance_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        # Log successful export
        audit.log_from_request(
            request=request,
            action=AuditAction.BULK_EXPORT,
            resource=AuditResource.PERFORMANCE,
            details={"count": len(performances), "format": "excel", "filename": filename},
            success=True,
        )
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )
    except Exception as exc:
        logger.error("Export daily performance excel failed: %s", exc, exc_info=True)
        # Log failed export
        audit.log_from_request(
            request=request,
            action=AuditAction.BULK_EXPORT,
            resource=AuditResource.PERFORMANCE,
            details={"error": str(exc), "format": "excel"},
            success=False,
        )
        raise http_error(
            500,
            ErrorCode.EXPORT_FAILED,
            "Export failed",
            request,
            context={"error": str(exc)},
        )


@router.get("/performance/csv")
@require_permission("exports:generate")
async def export_daily_performance_csv(request: Request, db: Session = Depends(get_db)):
    audit = AuditLogger(db)
    try:
        DailyPerformance, Student, Course = import_names("models", "DailyPerformance", "Student", "Course")
        lang = get_lang(request)
        na_value = not_available(lang)
        performances = db.query(DailyPerformance).filter(DailyPerformance.deleted_at.is_(None)).all()
        headers = get_header_row("daily_performance", lang)
        rows = []
        for p in performances:
            student = db.query(Student).filter(Student.id == p.student_id, Student.deleted_at.is_(None)).first()
            course = db.query(Course).filter(Course.id == p.course_id, Course.deleted_at.is_(None)).first()
            rows.append(
                [
                    p.id,
                    p.student_id,
                    f"{student.first_name} {student.last_name}" if student else na_value,
                    p.course_id,
                    course.course_name if course else na_value,
                    format_date_value(p.date, lang),
                    translate_grade_category(p.category or na_value, lang),
                    p.score,
                    p.max_score,
                    f"{p.percentage:.2f}%",
                    p.notes or "",
                ]
            )
        filename = f"daily_performance_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        audit.log_from_request(
            request=request,
            action=AuditAction.BULK_EXPORT,
            resource=AuditResource.PERFORMANCE,
            details={"count": len(performances), "format": "csv", "filename": filename},
            success=True,
        )
        return _csv_response(headers, rows, filename)
    except Exception as exc:
        logger.error("Export daily performance csv failed: %s", exc, exc_info=True)
        audit.log_from_request(
            request=request,
            action=AuditAction.BULK_EXPORT,
            resource=AuditResource.PERFORMANCE,
            details={"error": str(exc), "format": "csv"},
            success=False,
        )
        raise http_error(500, ErrorCode.EXPORT_FAILED, "Export failed", request, context={"error": str(exc)})


@router.get("/performance/pdf")
@require_permission("exports:generate")
async def export_daily_performance_pdf(request: Request, db: Session = Depends(get_db)):
    audit = AuditLogger(db)
    try:
        DailyPerformance, Student, Course = import_names("models", "DailyPerformance", "Student", "Course")
        lang = get_lang(request)
        na_value = not_available(lang)
        performances = db.query(DailyPerformance).filter(DailyPerformance.deleted_at.is_(None)).all()
        headers = get_header_row("daily_performance", lang)
        rows = []
        for p in performances:
            student = db.query(Student).filter(Student.id == p.student_id, Student.deleted_at.is_(None)).first()
            course = db.query(Course).filter(Course.id == p.course_id, Course.deleted_at.is_(None)).first()
            rows.append(
                [
                    str(p.id),
                    str(p.student_id),
                    f"{student.first_name} {student.last_name}" if student else na_value,
                    str(p.course_id),
                    course.course_name if course else na_value,
                    format_date_value(p.date, lang),
                    translate_grade_category(p.category or na_value, lang),
                    str(p.score),
                    str(p.max_score),
                    f"{p.percentage:.2f}%",
                    p.notes or "",
                ]
            )
        filename = f"daily_performance_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        audit.log_from_request(
            request=request,
            action=AuditAction.BULK_EXPORT,
            resource=AuditResource.PERFORMANCE,
            details={"count": len(performances), "format": "pdf", "filename": filename},
            success=True,
        )
        return _pdf_table_response(t("sheet_daily_performance", lang), headers, rows, filename)
    except Exception as exc:
        logger.error("Export daily performance pdf failed: %s", exc, exc_info=True)
        audit.log_from_request(
            request=request,
            action=AuditAction.BULK_EXPORT,
            resource=AuditResource.PERFORMANCE,
            details={"error": str(exc), "format": "pdf"},
            success=False,
        )
        raise http_error(500, ErrorCode.EXPORT_FAILED, "Export failed", request, context={"error": str(exc)})


@router.get("/highlights/excel")
@require_permission("exports:generate")
async def export_highlights_excel(request: Request, db: Session = Depends(get_db)):
    """Export all student highlights to Excel"""
    audit = AuditLogger(db)
    try:
        Highlight, Student = import_names("models", "Highlight", "Student")
        lang = get_lang(request)
        na_value = not_available(lang)

        highlights = db.query(Highlight).filter(Highlight.deleted_at.is_(None)).all()

        wb = openpyxl.Workbook()
        ws: Any = wb.active
        ws.title = t("sheet_highlights", lang)
        headers = get_header_row("highlights", lang)
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        for row, h in enumerate(highlights, 2):
            student = db.query(Student).filter(Student.id == h.student_id, Student.deleted_at.is_(None)).first()

            ws.cell(row=row, column=1, value=h.id)
            ws.cell(row=row, column=2, value=h.student_id)
            ws.cell(row=row, column=3, value=f"{student.first_name} {student.last_name}" if student else na_value)
            ws.cell(row=row, column=4, value=h.semester)
            ws.cell(row=row, column=5, value=translate_grade_category(h.category or na_value, lang))
            ws.cell(row=row, column=6, value=h.rating or na_value)
            ws.cell(row=row, column=7, value=h.highlight_text)
            ws.cell(row=row, column=8, value=format_date_value(h.date_created, lang))
            ws.cell(row=row, column=9, value=yes_no(bool(h.is_positive), lang))

        for col in range(1, 10):
            ws.column_dimensions[get_column_letter(col)].width = 18

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        filename = f"highlights_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        # Log successful export
        audit.log_from_request(
            request=request,
            action=AuditAction.BULK_EXPORT,
            resource=AuditResource.HIGHLIGHT,
            details={"count": len(highlights), "format": "excel", "filename": filename},
            success=True,
        )
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )
    except Exception as exc:
        logger.error("Export highlights excel failed: %s", exc, exc_info=True)
        # Log failed export
        audit.log_from_request(
            request=request,
            action=AuditAction.BULK_EXPORT,
            resource=AuditResource.HIGHLIGHT,
            details={"error": str(exc), "format": "excel"},
            success=False,
        )
        raise http_error(
            500,
            ErrorCode.EXPORT_FAILED,
            "Export failed",
            request,
            context={"error": str(exc)},
        )


@router.get("/highlights/csv")
@require_permission("exports:generate")
async def export_highlights_csv(request: Request, db: Session = Depends(get_db)):
    audit = AuditLogger(db)
    try:
        Highlight, Student = import_names("models", "Highlight", "Student")
        lang = get_lang(request)
        na_value = not_available(lang)
        highlights = db.query(Highlight).filter(Highlight.deleted_at.is_(None)).all()
        headers = get_header_row("highlights", lang)
        rows = []
        for h in highlights:
            student = db.query(Student).filter(Student.id == h.student_id, Student.deleted_at.is_(None)).first()
            rows.append(
                [
                    h.id,
                    h.student_id,
                    f"{student.first_name} {student.last_name}" if student else na_value,
                    h.semester,
                    translate_grade_category(h.category or na_value, lang),
                    h.rating or na_value,
                    h.highlight_text,
                    format_date_value(h.date_created, lang),
                    yes_no(bool(h.is_positive), lang),
                ]
            )
        filename = f"highlights_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        audit.log_from_request(
            request=request,
            action=AuditAction.BULK_EXPORT,
            resource=AuditResource.HIGHLIGHT,
            details={"count": len(highlights), "format": "csv", "filename": filename},
            success=True,
        )
        return _csv_response(headers, rows, filename)
    except Exception as exc:
        logger.error("Export highlights csv failed: %s", exc, exc_info=True)
        audit.log_from_request(
            request=request,
            action=AuditAction.BULK_EXPORT,
            resource=AuditResource.HIGHLIGHT,
            details={"error": str(exc), "format": "csv"},
            success=False,
        )
        raise http_error(500, ErrorCode.EXPORT_FAILED, "Export failed", request, context={"error": str(exc)})


@router.get("/highlights/pdf")
@require_permission("exports:generate")
async def export_highlights_pdf(request: Request, db: Session = Depends(get_db)):
    audit = AuditLogger(db)
    try:
        Highlight, Student = import_names("models", "Highlight", "Student")
        lang = get_lang(request)
        na_value = not_available(lang)
        highlights = db.query(Highlight).filter(Highlight.deleted_at.is_(None)).all()
        headers = get_header_row("highlights", lang)
        rows = []
        for h in highlights:
            student = db.query(Student).filter(Student.id == h.student_id, Student.deleted_at.is_(None)).first()
            rows.append(
                [
                    str(h.id),
                    str(h.student_id),
                    f"{student.first_name} {student.last_name}" if student else na_value,
                    h.semester,
                    translate_grade_category(h.category or na_value, lang),
                    h.rating or na_value,
                    h.highlight_text,
                    format_date_value(h.date_created, lang),
                    yes_no(bool(h.is_positive), lang),
                ]
            )
        filename = f"highlights_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        audit.log_from_request(
            request=request,
            action=AuditAction.BULK_EXPORT,
            resource=AuditResource.HIGHLIGHT,
            details={"count": len(highlights), "format": "pdf", "filename": filename},
            success=True,
        )
        return _pdf_table_response(t("sheet_highlights", lang), headers, rows, filename)
    except Exception as exc:
        logger.error("Export highlights pdf failed: %s", exc, exc_info=True)
        audit.log_from_request(
            request=request,
            action=AuditAction.BULK_EXPORT,
            resource=AuditResource.HIGHLIGHT,
            details={"error": str(exc), "format": "pdf"},
            success=False,
        )
        raise http_error(500, ErrorCode.EXPORT_FAILED, "Export failed", request, context={"error": str(exc)})


@router.get("/student-report/pdf/{student_id}")
@require_permission("exports:generate")
async def export_student_report_pdf(student_id: int, request: Request, db: Session = Depends(get_db)):
    """Generate comprehensive student report PDF with grades, attendance, and analytics"""
    try:
        Student, Grade, Attendance, Course, DailyPerformance = import_names(
            "models", "Student", "Grade", "Attendance", "Course", "DailyPerformance"
        )
        lang = get_lang(request)
        # Register DejaVu Sans font for Unicode/Greek support
        font_path = os.path.join(os.path.dirname(__file__), "../fonts/DejaVuSans.ttf")
        pdfmetrics.registerFont(TTFont("DejaVuSans", font_path))
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter, topMargin=0.5 * inch, bottomMargin=0.5 * inch)
        elements: list[Flowable] = []
        styles = getSampleStyleSheet()
        # Title
        title_style = ParagraphStyle(
            "CustomTitle",
            parent=styles["Heading1"],
            fontName="DejaVuSans",
            fontSize=20,
            textColor=colors.HexColor("#4F46E5"),
            spaceAfter=10,
            alignment=1,
        )
        student = db.query(Student).filter(Student.id == student_id, Student.deleted_at.is_(None)).first()
        if not student:
            raise http_error(
                404,
                ErrorCode.STUDENT_NOT_FOUND,
                "Student not found",
                request,
                context={"student_id": student_id},
            )
        elements.append(Paragraph(t("report_title", lang), title_style))
        # Student Info
        info_style = ParagraphStyle(
            "InfoStyle",
            parent=styles["Normal"],
            fontName="DejaVuSans",
            fontSize=12,
            spaceAfter=20,
            alignment=1,
        )
        elements.append(
            Paragraph(
                f"<b>{student.first_name} {student.last_name}</b> ({student.student_id})<br/>{student.email}",
                info_style,
            )
        )
        elements.append(Spacer(1, 0.3 * inch))
        # Attendance Summary
        subtitle_style = ParagraphStyle(
            "SubtitleStyle",
            parent=styles["Heading2"],
            fontName="DejaVuSans",
            fontSize=14,
            textColor=colors.HexColor("#4F46E5"),
            spaceAfter=10,
        )
        elements.append(Paragraph(t("attendance_summary", lang), subtitle_style))
        attendance_records = (
            db.query(Attendance).filter(Attendance.student_id == student_id, Attendance.deleted_at.is_(None)).all()
        )
        total_att = len(attendance_records)
        present = len([a for a in attendance_records if a.status == "Present"])
        absent = len([a for a in attendance_records if a.status == "Absent"])
        late = len([a for a in attendance_records if a.status == "Late"])
        excused = len([a for a in attendance_records if a.status == "Excused"])
        att_rate = (present / total_att * 100) if total_att > 0 else 0
        att_headers = [
            t("total_classes", lang),
            t("present", lang),
            t("absent", lang),
            t("late", lang),
            t("excused", lang),
            t("attendance_rate", lang),
        ]
        att_rows = [[str(total_att), str(present), str(absent), str(late), str(excused), f"{att_rate:.1f}%"]]
        att_table = _build_wrapped_table(att_headers, att_rows, doc.width, align="CENTER")
        elements.append(att_table)
        elements.append(Spacer(1, 0.3 * inch))
        # Grades by Course
        elements.append(Paragraph(t("grades_by_course", lang), subtitle_style))
        grades = db.query(Grade).filter(Grade.student_id == student_id, Grade.deleted_at.is_(None)).all()
        course_grades = {}
        for g in grades:
            if g.course_id not in course_grades:
                course_grades[g.course_id] = []
            course_grades[g.course_id].append(g)
        for course_id, course_grade_list in course_grades.items():
            course = db.query(Course).filter(Course.id == course_id, Course.deleted_at.is_(None)).first()
            if not course:
                continue
            elements.append(
                Paragraph(
                    f"<b>{course.course_code} - {course.course_name}</b>",
                    ParagraphStyle("CourseTitle", parent=styles["Normal"], fontName="DejaVuSans"),
                )
            )
            elements.append(Spacer(1, 0.1 * inch))
            grade_headers = [
                t("assignment", lang),
                t("category", lang),
                t("score", lang),
                t("max_score", lang),
                t("percentage", lang),
                t("letter", lang),
            ]
            grade_rows: list[list[Any]] = []
            total_pct = 0
            for g in course_grade_list:
                pct = (g.grade / g.max_grade * 100) if g.max_grade else 0
                total_pct += pct
                grade_rows.append(
                    [
                        translate_assignment_name(g.assignment_name, lang),
                        translate_grade_category(g.category or not_available(lang), lang),
                        str(g.grade),
                        str(g.max_grade),
                        f"{pct:.1f}%",
                        _letter_grade(pct),
                    ]
                )
            avg_pct = total_pct / len(course_grade_list) if course_grade_list else 0
            grade_rows.append(["", "", "", t("average", lang), f"{avg_pct:.1f}%", _letter_grade(avg_pct)])
            grade_table = _build_wrapped_table(
                grade_headers,
                grade_rows,
                doc.width,
                align="CENTER",
                extra_styles=[("BACKGROUND", (0, -1), (-1, -1), colors.lightgrey)],
            )
            elements.append(grade_table)
            elements.append(Spacer(1, 0.2 * inch))
        # Daily Performance Summary
        daily_perf = (
            db.query(DailyPerformance)
            .filter(
                DailyPerformance.student_id == student_id,
                DailyPerformance.deleted_at.is_(None),
            )
            .all()
        )
        if daily_perf:
            elements.append(Paragraph(t("daily_performance_summary", lang), subtitle_style))
            avg_perf = sum(dp.percentage for dp in daily_perf) / len(daily_perf)
            perf_text = (
                f"{t('total_daily_entries', lang)}: {len(daily_perf)}<br/>{t('avg_performance', lang)}: {avg_perf:.1f}%"
            )
            elements.append(
                Paragraph(perf_text, ParagraphStyle("PerfText", parent=styles["Normal"], fontName="DejaVuSans"))
            )
            elements.append(Spacer(1, 0.2 * inch))
        # Footer
        footer_style = ParagraphStyle(
            "FooterStyle",
            parent=styles["Normal"],
            fontName="DejaVuSans",
            fontSize=8,
            textColor=colors.grey,
            alignment=1,
        )
        elements.append(Spacer(1, 0.3 * inch))
        elements.append(
            Paragraph(
                f"{t('generated_on', lang)} {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}<br/>{t('system', lang)}",
                footer_style,
            )
        )
        doc.build(elements)
        buffer.seek(0)
        filename = f"student_report_{student.student_id}_{datetime.now().strftime('%Y%m%d')}.pdf"
        return StreamingResponse(
            buffer,
            media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )
    except HTTPException:
        raise
    except Exception as exc:
        logger.error("Export student report pdf failed: %s", exc, exc_info=True)
        raise http_error(
            500,
            ErrorCode.EXPORT_FAILED,
            "Export failed",
            request,
            context={"error": str(exc)},
        )


@router.get("/courses/pdf")
@require_permission("exports:generate")
async def export_courses_pdf(request: Request, db: Session = Depends(get_db)):
    """Export all courses to PDF"""
    try:
        (Course,) = import_names("models", "Course")
        lang = get_lang(request)
        courses = db.query(Course).filter(Course.deleted_at.is_(None)).all()
        filename = f"courses_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        headers = [
            t("label_course_code", lang),
            t("label_course_name", lang),
            t("label_semester", lang),
            t("label_credits", lang),
            t("label_hours", lang),
            t("label_periods_per_week", lang),
        ]
        rows = [
            [
                c.course_code,
                c.course_name,
                c.semester or not_available(lang),
                str(c.credits or 0),
                str(c.hours_per_week or 0),
                str(c.periods_per_week or 0),
            ]
            for c in courses
        ]
        return _pdf_table_response(t("title_course_catalog", lang), headers, rows, filename)
    except Exception as exc:
        logger.error("Export courses pdf failed: %s", exc, exc_info=True)
        raise http_error(
            500,
            ErrorCode.EXPORT_FAILED,
            "Export failed",
            request,
            context={"error": str(exc)},
        )


@router.get("/analytics/course/{course_id}/pdf")
@require_permission("exports:generate")
async def export_course_analytics_pdf(course_id: int, request: Request, db: Session = Depends(get_db)):
    """Export course analytics report to PDF"""
    try:
        Course, Grade, CourseEnrollment = import_names("models", "Course", "Grade", "CourseEnrollment")
        lang = get_lang(request)
        font_path = os.path.join(os.path.dirname(__file__), "../fonts/DejaVuSans.ttf")
        pdfmetrics.registerFont(TTFont("DejaVuSans", font_path))

        course = db.query(Course).filter(Course.id == course_id, Course.deleted_at.is_(None)).first()
        if not course:
            raise http_error(
                404,
                ErrorCode.COURSE_NOT_FOUND,
                "Course not found",
                request,
                context={"course_id": course_id},
            )

        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter, topMargin=0.5 * inch)
        elements: list[Flowable] = []
        styles = getSampleStyleSheet()

        # Title
        title_style = ParagraphStyle(
            "CustomTitle",
            parent=styles["Heading1"],
            fontName="DejaVuSans",
            fontSize=20,
            textColor=colors.HexColor("#4F46E5"),
            spaceAfter=10,
            alignment=1,
        )
        elements.append(Paragraph(t("title_course_analytics", lang), title_style))

        # Course Info
        info_style = ParagraphStyle(
            "InfoStyle",
            parent=styles["Normal"],
            fontName="DejaVuSans",
            fontSize=12,
            spaceAfter=20,
            alignment=1,
        )
        elements.append(
            Paragraph(
                f"<b>{course.course_code} - {course.course_name}</b><br/>{t('label_semester', lang)}: {course.semester or not_available(lang)}<br/>{t('label_credits', lang)}: {course.credits or 0}",
                info_style,
            )
        )
        elements.append(Spacer(1, 0.3 * inch))

        # Get enrollments
        enrollments = (
            db.query(CourseEnrollment)
            .filter(
                CourseEnrollment.course_id == course_id,
                CourseEnrollment.deleted_at.is_(None),
            )
            .all()
        )

        # Get all grades for this course
        grades = db.query(Grade).filter(Grade.course_id == course_id, Grade.deleted_at.is_(None)).all()

        # Calculate statistics
        student_ids = set([e.student_id for e in enrollments])
        total_students = len(student_ids)
        total_assignments = len(grades)

        percentages: list[float] = []
        if grades:
            percentages = [(g.grade / g.max_grade * 100) if g.max_grade else 0 for g in grades]
            avg_grade = sum(percentages) / len(percentages)
            highest = max(percentages)
            lowest = min(percentages)
        else:
            avg_grade = highest = lowest = 0

        # Statistics Table
        subtitle_style = ParagraphStyle(
            "SubtitleStyle",
            parent=styles["Heading2"],
            fontName="DejaVuSans",
            fontSize=14,
            textColor=colors.HexColor("#4F46E5"),
            spaceAfter=10,
        )
        elements.append(Paragraph(t("title_course_statistics", lang), subtitle_style))

        stats_headers = [t("label_metric", lang), t("label_value", lang)]
        stats_rows = [
            [t("label_total_students_enrolled", lang), str(total_students)],
            [t("label_total_assignments", lang), str(total_assignments)],
            [t("label_average_grade", lang), f"{avg_grade:.1f}%"],
            [t("label_highest_grade", lang), f"{highest:.1f}%"],
            [t("label_lowest_grade", lang), f"{lowest:.1f}%"],
        ]
        stats_table = _build_wrapped_table(stats_headers, stats_rows, doc.width, align="LEFT")
        elements.append(stats_table)
        elements.append(Spacer(1, 0.3 * inch))

        # Grade Distribution
        if grades:
            elements.append(Paragraph(t("title_grade_distribution", lang), subtitle_style))

            # Count letter grades
            a_count = sum(1 for p in percentages if p >= 90)
            b_count = sum(1 for p in percentages if 80 <= p < 90)
            c_count = sum(1 for p in percentages if 70 <= p < 80)
            d_count = sum(1 for p in percentages if 60 <= p < 70)
            f_count = sum(1 for p in percentages if p < 60)

            dist_headers = [t("label_letter_grade", lang), t("label_count", lang), t("label_percentage", lang)]
            dist_rows = [
                [t("distribution_a", lang), str(a_count), f"{a_count / len(grades) * 100:.1f}%"],
                [t("distribution_b", lang), str(b_count), f"{b_count / len(grades) * 100:.1f}%"],
                [t("distribution_c", lang), str(c_count), f"{c_count / len(grades) * 100:.1f}%"],
                [t("distribution_d", lang), str(d_count), f"{d_count / len(grades) * 100:.1f}%"],
                [t("distribution_f", lang), str(f_count), f"{f_count / len(grades) * 100:.1f}%"],
            ]
            dist_table = _build_wrapped_table(dist_headers, dist_rows, doc.width, align="CENTER")
            elements.append(dist_table)

        # Footer
        footer_style = ParagraphStyle(
            "FooterStyle",
            parent=styles["Normal"],
            fontName="DejaVuSans",
            fontSize=8,
            textColor=colors.grey,
            alignment=1,
        )
        elements.append(Spacer(1, 0.3 * inch))
        elements.append(
            Paragraph(
                f"{t('generated_on', lang)} {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}<br/>{t('system', lang)}",
                footer_style,
            )
        )

        doc.build(elements)
        buffer.seek(0)
        filename = f"course_analytics_{course.course_code}_{datetime.now().strftime('%Y%m%d')}.pdf"
        return StreamingResponse(
            buffer,
            media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )
    except HTTPException:
        raise
    except Exception as exc:
        logger.error("Export course analytics pdf failed: %s", exc, exc_info=True)
        raise http_error(
            500,
            ErrorCode.EXPORT_FAILED,
            "Export failed",
            request,
            context={"error": str(exc)},
        )
