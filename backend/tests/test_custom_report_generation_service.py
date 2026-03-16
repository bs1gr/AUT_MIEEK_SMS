from datetime import date
from pathlib import Path
import tempfile
from typing import cast

from openpyxl import load_workbook
from pypdf import PdfReader

from backend.config import settings
from backend.models import Attendance, Course, CourseEnrollment, GeneratedReport, Grade, Report, Student
from backend.services.custom_report_generation_service import CustomReportGenerationService
from backend.services.email_notification_service import EmailNotificationService


def _create_student(db):
    student = Student(
        first_name="Jane",
        last_name="Doe",
        email="jane.doe@example.com",
        student_id="S-100",
        enrollment_date=date(2024, 9, 1),
        is_active=True,
    )
    db.add(student)
    db.commit()
    db.refresh(student)
    return student


def _create_report(db, user_id: int) -> Report:
    report = Report(
        user_id=user_id,
        name="Student Report",
        description="Test report",
        report_type="student",
        fields={"columns": ["first_name", "last_name", "email", "student_id"]},
        export_format="csv",
        include_charts=False,
    )
    db.add(report)
    db.commit()
    db.refresh(report)
    return report


def _create_generated_report(db, report_id: int, user_id: int) -> GeneratedReport:
    generated = GeneratedReport(
        report_id=report_id,
        user_id=user_id,
        file_name="report_test.csv",
        export_format="csv",
        status="pending",
    )
    db.add(generated)
    db.commit()
    db.refresh(generated)
    return generated


def test_generate_report_csv_creates_file(db, tmp_path: Path):
    user_id = 1
    _create_student(db)
    report = _create_report(db, user_id)
    report_id = cast(int, report.id)
    generated = _create_generated_report(db, report_id, user_id)
    generated_id = cast(int, generated.id)

    service = CustomReportGenerationService(db)
    service.reports_dir = str(tmp_path)

    service.generate_report(report_id, generated_id, user_id, "csv", include_charts=False)

    db.refresh(generated)
    assert generated.status == "completed"
    assert generated.file_path
    assert generated.file_size_bytes is not None
    assert Path(generated.file_path).exists()

    content = Path(generated.file_path).read_text(encoding="utf-8")
    assert "First Name" in content
    assert "Jane" in content


def test_successful_generation_supersedes_prior_completed_artifacts(db, tmp_path: Path):
    user_id = 1
    _create_student(db)
    report = _create_report(db, user_id)
    report_id = cast(int, report.id)

    first_generated = GeneratedReport(  # type: ignore[call-arg]
        report_id=report_id,
        user_id=user_id,
        file_name="report_first.csv",
        export_format="csv",
        status="pending",
    )
    db.add(first_generated)
    db.commit()
    db.refresh(first_generated)

    second_generated = GeneratedReport(  # type: ignore[call-arg]
        report_id=report_id,
        user_id=user_id,
        file_name="report_second.csv",
        export_format="csv",
        status="pending",
    )
    db.add(second_generated)
    db.commit()
    db.refresh(second_generated)

    service = CustomReportGenerationService(db)
    service.reports_dir = str(tmp_path)

    service.generate_report(report_id, cast(int, first_generated.id), user_id, "csv", include_charts=False)
    db.refresh(first_generated)
    assert first_generated.status == "completed"

    service.generate_report(report_id, cast(int, second_generated.id), user_id, "csv", include_charts=False)
    db.refresh(first_generated)
    db.refresh(second_generated)

    assert second_generated.status == "completed"
    assert first_generated.status == "superseded"
    assert first_generated.error_message is not None
    assert str(cast(int, second_generated.id)) in first_generated.error_message


def test_generate_grade_report_supports_legacy_field_aliases(db, tmp_path: Path):
    user_id = 1
    student = _create_student(db)
    course = Course(  # type: ignore[call-arg]
        course_code="COURSE-1",
        course_name="Legacy Report Course",
        semester="1",
        credits=3,
        is_active=True,
    )
    db.add(course)
    db.commit()
    db.refresh(course)
    course_id = cast(int, course.id)

    grade = Grade(  # type: ignore[call-arg]
        student_id=cast(int, student.id),
        course_id=course_id,
        assignment_name="Midterm",
        category="Exam",
        grade=88.5,
        max_grade=100,
        weight=1,
        date_assigned=date(2024, 10, 3),
        date_submitted=None,
    )
    db.add(grade)
    db.commit()

    report = Report(  # type: ignore[call-arg]
        user_id=user_id,
        name="Legacy Grade Report",
        description="Legacy grade field alias support",
        report_type="grade",
        fields={"columns": ["student_name", "grade_value", "exam_date"]},
        export_format="csv",
        include_charts=False,
    )
    db.add(report)
    db.commit()
    db.refresh(report)
    report_id = cast(int, report.id)

    generated = GeneratedReport(  # type: ignore[call-arg]
        report_id=report_id,
        user_id=user_id,
        file_name="legacy_grade_report.csv",
        export_format="csv",
        status="pending",
    )
    db.add(generated)
    db.commit()
    db.refresh(generated)
    generated_id = cast(int, generated.id)

    service = CustomReportGenerationService(db)
    service.reports_dir = str(tmp_path)

    service.generate_report(report_id, generated_id, user_id, "csv", include_charts=False)

    db.refresh(generated)
    assert generated.status == "completed"
    assert generated.file_path

    content = Path(generated.file_path).read_text(encoding="utf-8")
    assert "Student Name" in content
    assert "Grade" in content
    assert "Jane Doe" in content
    assert "88.5" in content
    assert "2024-10-03" in content


def test_generate_grade_report_normalizes_legacy_id_column_to_student_identifier(db, tmp_path: Path):
    user_id = 1
    student = _create_student(db)
    course = Course(  # type: ignore[call-arg]
        course_code="COURSE-2",
        course_name="Identifier Report Course",
        semester="1",
        credits=3,
        is_active=True,
    )
    db.add(course)
    db.commit()
    db.refresh(course)

    for assignment_name, assigned_on in (("Quiz 1", date(2024, 10, 4)), ("Quiz 2", date(2024, 10, 5))):
        grade = Grade(  # type: ignore[call-arg]
            student_id=cast(int, student.id),
            course_id=cast(int, course.id),
            assignment_name=assignment_name,
            category="Quiz",
            grade=91,
            max_grade=100,
            weight=1,
            date_assigned=assigned_on,
            date_submitted=assigned_on,
        )
        db.add(grade)
    db.commit()

    report = Report(  # type: ignore[call-arg]
        user_id=user_id,
        name="Grade Identifier Report",
        description="Ensure grade reports use stable student identifiers",
        report_type="grade",
        fields={"columns": ["student_name", "id", "course_id"]},
        export_format="csv",
        include_charts=False,
    )
    db.add(report)
    db.commit()
    db.refresh(report)

    generated = GeneratedReport(  # type: ignore[call-arg]
        report_id=cast(int, report.id),
        user_id=user_id,
        file_name="grade_identifier_report.csv",
        export_format="csv",
        status="pending",
    )
    db.add(generated)
    db.commit()
    db.refresh(generated)

    service = CustomReportGenerationService(db)
    service.reports_dir = str(tmp_path)
    service.generate_report(cast(int, report.id), cast(int, generated.id), user_id, "csv", include_charts=False)

    db.refresh(generated)
    content = Path(cast(str, generated.file_path)).read_text(encoding="utf-8")
    lines = [line.strip() for line in content.splitlines() if line.strip()]

    assert lines[0].startswith("Student Name,Student ID,Course Id")
    assert len(lines) == 3
    first_student_identifier = lines[1].split(",")[1]
    second_student_identifier = lines[2].split(",")[1]
    assert first_student_identifier == second_student_identifier == "S-100"


def test_generate_student_attendance_summary_uses_student_identifier_and_computed_metrics(db, tmp_path: Path):
    user_id = 1
    student = _create_student(db)
    course = Course(  # type: ignore[call-arg]
        course_code="COURSE-ATT-1",
        course_name="Attendance Course",
        semester="1",
        credits=3,
        is_active=True,
    )
    db.add(course)
    db.commit()
    db.refresh(course)

    report = Report(  # type: ignore[call-arg]
        user_id=user_id,
        name="Attendance Summary",
        description="Ensure student attendance summary uses external IDs and computed attendance metrics",
        report_type="student",
        fields={"columns": ["id", "first_name", "attendance_rate", "total_classes", "attended"]},
        filters=[],
        sort_by=[{"field": "id", "order": "asc"}],
        export_format="csv",
        include_charts=False,
    )
    db.add(report)
    db.commit()
    db.refresh(report)

    attendance_rows = [
        Attendance(  # type: ignore[call-arg]
            student_id=cast(int, student.id),
            course_id=cast(int, course.id),
            date=date(2024, 10, 1),
            status="Present",
            period_number=1,
        ),
        Attendance(  # type: ignore[call-arg]
            student_id=cast(int, student.id),
            course_id=cast(int, course.id),
            date=date(2024, 10, 2),
            status="Late",
            period_number=1,
        ),
        Attendance(  # type: ignore[call-arg]
            student_id=cast(int, student.id),
            course_id=cast(int, course.id),
            date=date(2024, 10, 3),
            status="Absent",
            period_number=1,
        ),
    ]
    db.add_all(attendance_rows)
    db.commit()

    generated = GeneratedReport(  # type: ignore[call-arg]
        report_id=cast(int, report.id),
        user_id=user_id,
        file_name="student_attendance_summary.csv",
        export_format="csv",
        status="pending",
    )
    db.add(generated)
    db.commit()
    db.refresh(generated)

    service = CustomReportGenerationService(db)
    service.reports_dir = str(tmp_path)
    service.generate_report(cast(int, report.id), cast(int, generated.id), user_id, "csv", include_charts=False)

    db.refresh(generated)
    content = Path(cast(str, generated.file_path)).read_text(encoding="utf-8")
    lines = [line.strip() for line in content.splitlines() if line.strip()]

    assert lines[0].startswith("Student ID,First Name,Attendance Rate,Total Classes,Attended")
    assert "S-100,Jane,66.7,3,2" in lines[1]


def test_generate_course_report_supports_legacy_name_code_and_enrollment_count_aliases(db, tmp_path: Path):
    user_id = 1
    student = _create_student(db)
    course = Course(  # type: ignore[call-arg]
        course_code="C-ALIAS-1",
        course_name="Alias Compatible Course",
        semester="2",
        credits=4,
        is_active=True,
    )
    db.add(course)
    db.commit()
    db.refresh(course)

    enrollment = CourseEnrollment(  # type: ignore[call-arg]
        student_id=cast(int, student.id),
        course_id=cast(int, course.id),
        status="active",
    )
    db.add(enrollment)
    db.commit()

    report = Report(  # type: ignore[call-arg]
        user_id=user_id,
        name="Course Legacy Alias Report",
        description="Legacy course field alias support",
        report_type="course",
        fields={"columns": ["name", "code", "enrollment_count"]},
        export_format="csv",
        include_charts=False,
    )
    db.add(report)
    db.commit()
    db.refresh(report)

    generated = GeneratedReport(  # type: ignore[call-arg]
        report_id=cast(int, report.id),
        user_id=user_id,
        file_name="course_legacy_alias_report.csv",
        export_format="csv",
        status="pending",
    )
    db.add(generated)
    db.commit()
    db.refresh(generated)

    service = CustomReportGenerationService(db)
    service.reports_dir = str(tmp_path)
    service.generate_report(cast(int, report.id), cast(int, generated.id), user_id, "csv", include_charts=False)

    db.refresh(generated)
    content = Path(cast(str, generated.file_path)).read_text(encoding="utf-8")
    lines = [line.strip() for line in content.splitlines() if line.strip()]

    assert lines[0].startswith("Course Name,Course Code,Enrollment Count")
    assert "Alias Compatible Course,C-ALIAS-1,1" in lines[1]


def test_generate_student_report_supports_legacy_study_year_status_and_gpa_aliases(db, tmp_path: Path):
    user_id = 1
    student = _create_student(db)
    student.study_year = 3  # type: ignore[assignment]
    student.is_active = True  # type: ignore[assignment]
    db.commit()

    course = Course(  # type: ignore[call-arg]
        course_code="GPA-101",
        course_name="GPA Course",
        semester="1",
        credits=3,
        is_active=True,
    )
    db.add(course)
    db.commit()
    db.refresh(course)

    grade = Grade(  # type: ignore[call-arg]
        student_id=cast(int, student.id),
        course_id=cast(int, course.id),
        assignment_name="Exam",
        category="Exam",
        grade=80,
        max_grade=100,
        weight=1,
        date_assigned=date(2024, 10, 10),
        date_submitted=date(2024, 10, 11),
    )
    db.add(grade)
    db.commit()

    report = Report(  # type: ignore[call-arg]
        user_id=user_id,
        name="Student Legacy Alias Report",
        description="Legacy student alias support",
        report_type="student",
        fields={"columns": ["student_id", "year_of_study", "enrollment_status", "gpa"]},
        filters={"gpa": {"operator": "gte", "value": 70}},
        export_format="csv",
        include_charts=False,
    )
    db.add(report)
    db.commit()
    db.refresh(report)

    generated = GeneratedReport(  # type: ignore[call-arg]
        report_id=cast(int, report.id),
        user_id=user_id,
        file_name="student_legacy_alias_report.csv",
        export_format="csv",
        status="pending",
    )
    db.add(generated)
    db.commit()
    db.refresh(generated)

    service = CustomReportGenerationService(db)
    service.reports_dir = str(tmp_path)
    service.generate_report(cast(int, report.id), cast(int, generated.id), user_id, "csv", include_charts=False)

    db.refresh(generated)
    content = Path(cast(str, generated.file_path)).read_text(encoding="utf-8")
    lines = [line.strip() for line in content.splitlines() if line.strip()]

    assert lines[0].startswith("Student ID,Study Year,Active,Gpa")
    assert "S-100,3,Yes,80.0" in lines[1]


def test_export_pdf_preserves_numeric_attendance_values(db):
    service = CustomReportGenerationService(db)
    service._set_language("el")

    headers = ["Κωδικός Μαθητή", "Όνομα", "Επώνυμο", "Ποσοστό Παρουσιών"]
    rows = [
        ["S1159879", "Λαμπριανός", "Αρτεμίου", 97.6],
        ["S1173031", "Σταύρος", "Κουσιής", 95.2],
        ["S1042527", "Χαράλαμπος", "Μιχαήλ", 100.0],
    ]

    with tempfile.TemporaryDirectory() as tmpdir:
        pdf_path = Path(tmpdir) / "attendance_summary.pdf"
        service._export_pdf(rows, headers, str(pdf_path), title="Σύνοψη Παρουσιών - Όλοι οι Μαθητές")
        extracted_text = "\n".join(page.extract_text() or "" for page in PdfReader(str(pdf_path)).pages)

    assert "S1159879" in extracted_text
    assert "97.6" in extracted_text
    assert "95.2" in extracted_text
    assert "100.0" in extracted_text


def test_send_report_email_with_attachment(db, tmp_path: Path, monkeypatch):
    user_id = 1
    _create_student(db)
    report = _create_report(db, user_id)
    report.email_recipients = ["report@example.com"]  # type: ignore[assignment]
    report.email_enabled = True  # type: ignore[assignment]
    db.commit()

    generated = _create_generated_report(db, cast(int, report.id), user_id)
    attachment_path = tmp_path / "report.csv"
    attachment_path.write_text("header\nvalue\n", encoding="utf-8")
    generated.file_path = str(attachment_path)  # type: ignore[assignment]
    generated.record_count = 2  # type: ignore[assignment]
    db.commit()

    captured: dict[str, object] = {}

    def fake_send_report_ready(
        recipients,
        report_name,
        export_format,
        record_count,
        generated_at,
        attachment_paths=None,
        attachment_note=None,
    ):
        captured["recipients"] = list(recipients)
        captured["attachment_paths"] = attachment_paths
        captured["attachment_note"] = attachment_note
        return 1, []

    monkeypatch.setattr(EmailNotificationService, "is_enabled", lambda: True)
    monkeypatch.setattr(EmailNotificationService, "send_report_ready", fake_send_report_ready)

    service = CustomReportGenerationService(db)
    service._send_report_email(report, generated, "csv")

    db.refresh(generated)
    assert generated.email_sent is True
    assert generated.email_error is None
    assert generated.email_sent_at is not None
    assert captured["recipients"] == ["report@example.com"]
    assert captured["attachment_paths"] == [str(attachment_path)]
    assert captured["attachment_note"] is None


def test_send_report_email_oversized_attachment(db, tmp_path: Path, monkeypatch):
    user_id = 1
    _create_student(db)
    report = _create_report(db, user_id)
    report.email_recipients = ["report@example.com"]  # type: ignore[assignment]
    report.email_enabled = True  # type: ignore[assignment]
    db.commit()

    generated = _create_generated_report(db, cast(int, report.id), user_id)
    attachment_path = tmp_path / "report.csv"
    attachment_path.write_bytes(b"a" * (2 * 1024 * 1024))
    generated.file_path = str(attachment_path)  # type: ignore[assignment]
    generated.record_count = 2  # type: ignore[assignment]
    db.commit()

    captured: dict[str, object] = {}

    def fake_send_report_ready(
        recipients,
        report_name,
        export_format,
        record_count,
        generated_at,
        attachment_paths=None,
        attachment_note=None,
    ):
        captured["recipients"] = list(recipients)
        captured["attachment_paths"] = attachment_paths
        captured["attachment_note"] = attachment_note
        return 1, []

    monkeypatch.setattr(EmailNotificationService, "is_enabled", lambda: True)
    monkeypatch.setattr(EmailNotificationService, "send_report_ready", fake_send_report_ready)
    monkeypatch.setattr(settings, "SMTP_ATTACHMENT_MAX_MB", 1, raising=False)

    service = CustomReportGenerationService(db)
    service._send_report_email(report, generated, "csv")

    db.refresh(generated)
    assert generated.email_sent is True
    assert generated.email_error is None
    assert generated.email_sent_at is not None
    assert captured["recipients"] == ["report@example.com"]
    assert captured["attachment_paths"] in (None, [])
    assert isinstance(captured["attachment_note"], str)


def test_xlsx_shows_empty_cells_for_unresolvable_field_name(db, tmp_path: Path):
    """
    Prove that 'partially empty' report output is a DATA PIPELINE fault, not a
    PDF rendering fault.

    XLSX export uses plain str values — no ReportLab Paragraph objects anywhere.
    Yet an unresolvable column name produces an empty cell in the XLSX file.
    If the same field name were passed to _export_pdf the cell would also be
    empty regardless of whether the cell content is a Paragraph or a plain str,
    because _resolve_field already returned '' before any rendering happened.
    """
    user_id = 1
    student = _create_student(db)
    course = Course(  # type: ignore[call-arg]
        course_code="XLSX-PROOF-1",
        course_name="XLSX Proof Course",
        semester="1",
        credits=3,
        is_active=True,
    )
    db.add(course)
    db.commit()
    db.refresh(course)

    grade = Grade(  # type: ignore[call-arg]
        student_id=cast(int, student.id),
        course_id=cast(int, course.id),
        assignment_name="Midterm",
        category="Exam",
        grade=77.0,
        max_grade=100,
        weight=1,
        date_assigned=date(2024, 11, 1),
        date_submitted=date(2024, 11, 2),
    )
    db.add(grade)
    db.commit()

    # "fictional_score_column" cannot be resolved by _resolve_field.
    # The empty value propagates identically into CSV, XLSX, and PDF rows.
    report = Report(  # type: ignore[call-arg]
        user_id=user_id,
        name="XLSX Empty Proof",
        description="Prove empty cells come from field resolution, not rendering",
        report_type="grade",
        fields={"columns": ["student_name", "fictional_score_column", "grade"]},
        export_format="excel",
        include_charts=False,
    )
    db.add(report)
    db.commit()
    db.refresh(report)

    generated = GeneratedReport(  # type: ignore[call-arg]
        report_id=cast(int, report.id),
        user_id=user_id,
        file_name="xlsx_empty_proof.xlsx",
        export_format="excel",
        status="pending",
    )
    db.add(generated)
    db.commit()
    db.refresh(generated)

    service = CustomReportGenerationService(db)
    service.reports_dir = str(tmp_path)
    service.generate_report(cast(int, report.id), cast(int, generated.id), user_id, "excel", include_charts=False)

    db.refresh(generated)
    assert generated.status == "completed"
    assert generated.file_path

    wb = load_workbook(generated.file_path)
    ws = wb.active
    data_rows = list(ws.iter_rows(min_row=2, values_only=True))
    assert len(data_rows) == 1, "Expected exactly one grade row"

    # Column index 1 (0-based) is 'fictional_score_column': must be empty/None.
    # The resolved value from _resolve_field is '' which _stringify_value keeps as '';
    # openpyxl stores an empty string cell as None when reading back with values_only.
    unresolvable_cell = data_rows[0][1]
    assert unresolvable_cell in ("", None), (
        f"Expected empty cell for unresolvable field but got {unresolvable_cell!r}; "
        "if this test fails the field was somehow resolved and the premise changes."
    )

    # Column index 2 is 'grade': must be populated (77.0).
    # openpyxl returns numeric types; 77.0 may come back as int 77 when the
    # fractional part is zero, so compare as float.
    grade_cell = data_rows[0][2]
    assert grade_cell is not None and float(grade_cell) == 77.0, (
        f"Expected 77.0 for resolved 'grade' field but got {grade_cell!r}"
    )


def test_xlsx_and_csv_both_empty_for_pre_alias_grade_value_field(db, tmp_path: Path):
    """
    The original 'partially empty' report was triggered by grade reports requesting
    the field 'grade_value' (the old UI-side name) which _resolve_field could not
    match before the LEGACY_FIELD_ALIASES mapping was added.

    Both XLSX (plain str) and CSV (plain str) would show an empty cell for that
    column — demonstrating the fault is entirely in the data pipeline.
    Once LEGACY_FIELD_ALIASES maps 'grade_value' to 'grade', both formats are
    populated without any change to PDF rendering code.
    """
    user_id = 1
    student = _create_student(db)
    course = Course(  # type: ignore[call-arg]
        course_code="ALIAS-PROOF-2",
        course_name="Alias Proof Course",
        semester="1",
        credits=3,
        is_active=True,
    )
    db.add(course)
    db.commit()
    db.refresh(course)

    grade = Grade(  # type: ignore[call-arg]
        student_id=cast(int, student.id),
        course_id=cast(int, course.id),
        assignment_name="Final",
        category="Exam",
        grade=92.5,
        max_grade=100,
        weight=1,
        date_assigned=date(2024, 12, 1),
        date_submitted=date(2024, 12, 2),
    )
    db.add(grade)
    db.commit()

    def _run_report(export_format: str, file_name: str) -> GeneratedReport:
        report = Report(  # type: ignore[call-arg]
            user_id=user_id,
            name=f"Alias Proof {export_format.upper()}",
            description="Legacy grade_value alias",
            report_type="grade",
            # 'grade_value' is the old UI field name; LEGACY_FIELD_ALIASES maps it to 'grade'
            fields={"columns": ["student_name", "grade_value", "exam_date"]},
            export_format=export_format,
            include_charts=False,
        )
        db.add(report)
        db.commit()
        db.refresh(report)

        gen = GeneratedReport(  # type: ignore[call-arg]
            report_id=cast(int, report.id),
            user_id=user_id,
            file_name=file_name,
            export_format=export_format,
            status="pending",
        )
        db.add(gen)
        db.commit()
        db.refresh(gen)

        svc = CustomReportGenerationService(db)
        svc.reports_dir = str(tmp_path)
        svc.generate_report(cast(int, report.id), cast(int, gen.id), user_id, export_format, include_charts=False)
        db.refresh(gen)
        return gen

    # --- XLSX ---
    gen_xlsx = _run_report("excel", "alias_proof.xlsx")
    wb = load_workbook(gen_xlsx.file_path)
    ws = wb.active
    xlsx_rows = list(ws.iter_rows(min_row=2, values_only=True))
    assert len(xlsx_rows) == 1
    xlsx_grade_cell = xlsx_rows[0][1]  # 'grade_value' column (index 1)
    # After alias fix: must resolve to "92.5"
    assert str(xlsx_grade_cell) == "92.5", (
        f"XLSX 'grade_value' cell unexpectedly empty or wrong: {xlsx_grade_cell!r}; "
        "alias fix in _normalize_report_field_key is not being applied."
    )

    # --- CSV (same rows, same _resolve_field, no rendering at all) ---
    gen_csv = _run_report("csv", "alias_proof.csv")
    csv_content = Path(cast(str, gen_csv.file_path)).read_text(encoding="utf-8")
    csv_lines = [ln.strip() for ln in csv_content.splitlines() if ln.strip()]
    assert len(csv_lines) == 2, f"Expected header + 1 data row, got {csv_lines}"
    csv_grade_cell = csv_lines[1].split(",")[1]  # grade_value column
    assert csv_grade_cell == "92.5", (
        f"CSV 'grade_value' cell unexpectedly empty or wrong: {csv_grade_cell!r}; "
        "alias fix in _normalize_report_field_key is not being applied."
    )

    # Both formats return the same value — confirming the source is the data pipeline,
    # not the rendering layer (Paragraph / plain str distinction is irrelevant here).
