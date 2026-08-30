import csv
import zipfile
from datetime import date
from io import BytesIO, StringIO

from openpyxl import load_workbook


def _create_student(client, code: str, idx: int):
    response = client.post(
        "/api/v1/students/",
        json={
            "student_id": code,
            "email": f"{code.lower()}@test.com",
            "first_name": f"Alpha{idx}",
            "last_name": "Tester",
        },
    )
    assert response.status_code == 201
    return response.json()


def _create_course(client, code: str):
    response = client.post(
        "/api/v1/courses/",
        json={
            "course_code": code,
            "course_name": "Operations Analytics",
            "semester": "Fall",
            "credits": 3,
        },
    )
    assert response.status_code == 201
    return response.json()


def _find_row_by_first_cell(ws, value):
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row:
            continue
        if row[0] == value:
            return row
    raise AssertionError(f"Value {value} not found in sheet {ws.title}")


def test_attendance_analytics_export_contains_period_breakdown(client):
    student_one = _create_student(client, "AN001", 1)
    student_two = _create_student(client, "AN002", 2)
    course = _create_course(client, "OPS101")
    today = date.today().isoformat()

    payloads = [
        {"student_id": student_one["id"], "status": "Present", "period_number": 1},
        {"student_id": student_one["id"], "status": "Absent", "period_number": 2},
        {"student_id": student_two["id"], "status": "Late", "period_number": 1},
    ]

    for payload in payloads:
        response = client.post(
            "/api/v1/attendance/",
            json={
                "student_id": payload["student_id"],
                "course_id": course["id"],
                "date": today,
                "status": payload["status"],
                "period_number": payload["period_number"],
            },
        )
        assert response.status_code == 201

    export_response = client.get("/api/v1/export/attendance/analytics/excel")
    assert export_response.status_code == 200

    workbook = load_workbook(BytesIO(export_response.content))
    assert "Overview" in workbook.sheetnames
    assert "Course Summary" in workbook.sheetnames
    assert "Period Summary" in workbook.sheetnames
    assert "Student Summary" in workbook.sheetnames

    course_summary = workbook["Course Summary"]
    course_row = _find_row_by_first_cell(course_summary, course["course_code"])
    # Row schema: code, name, total, present, absent, late, excused, rate
    assert course_row[2] == 3  # total records for course
    assert course_row[3] == 1  # present
    assert course_row[4] == 1  # absent
    assert course_row[5] == 1  # late

    period_summary = workbook["Period Summary"]
    period_one = _find_row_by_first_cell(period_summary, 1)
    period_two = _find_row_by_first_cell(period_summary, 2)
    # Row schema: period, total, present, absent, late, excused, rate
    assert period_one[1] == 2  # two records in period 1
    assert period_one[2] == 1  # present count in period 1
    assert period_one[4] == 1  # late count in period 1
    assert period_two[1] == 1
    assert period_two[3] == 1  # absent count in period 2

    student_summary = workbook["Student Summary"]
    student_one_row = _find_row_by_first_cell(student_summary, student_one["student_id"])
    student_two_row = _find_row_by_first_cell(student_summary, student_two["student_id"])
    # Row schema: student_id, name, total, present, absent, late, excused, dominant
    assert student_one_row[2] == 2
    assert student_one_row[3] == 1
    assert student_one_row[4] == 1
    assert student_one_row[7] == "Present"
    assert student_two_row[2] == 1
    assert student_two_row[5] == 1


def test_export_all_zip_enriches_rows_from_preloaded_lookups(client):
    """Regression test for the N+1 fix in export_all_zip: enrollments.csv,
    all_grades.csv, and daily_performance.csv must still be enriched with the
    correct student name and course code/name, now sourced from an in-memory
    id->row lookup built once instead of a query per row.
    """
    student = _create_student(client, "ZIP001", 1)
    course = _create_course(client, "ZIP101")

    enroll_resp = client.post(
        f"/api/v1/enrollments/course/{course['id']}", json={"student_ids": [student["id"]]}
    )
    assert enroll_resp.status_code == 200

    grade_resp = client.post(
        "/api/v1/grades/",
        json={
            "student_id": student["id"],
            "course_id": course["id"],
            "assignment_name": "Zip Assignment",
            "category": "Homework",
            "grade": 88.0,
            "max_grade": 100.0,
            "weight": 1.0,
            "date_assigned": date.today().isoformat(),
            "date_submitted": date.today().isoformat(),
        },
    )
    assert grade_resp.status_code == 201

    perf_resp = client.post(
        "/api/v1/daily-performance/",
        json={
            "student_id": student["id"],
            "course_id": course["id"],
            "date": date.today().isoformat(),
            "category": "Participation",
            "score": 9.0,
            "max_score": 10.0,
            "notes": "Zip test",
        },
    )
    assert perf_resp.status_code == 200

    resp = client.get("/api/v1/export/all/zip")
    assert resp.status_code == 200, resp.text

    student_name = f"{student['first_name']} {student['last_name']}"
    with zipfile.ZipFile(BytesIO(resp.content)) as zf:
        enrollments_csv = zf.read("enrollments.csv").decode("utf-8-sig")
        enrollment_row = next(r for r in csv.reader(StringIO(enrollments_csv)) if r and r[1] == str(student["id"]))
        assert student_name in enrollment_row
        assert course["course_code"] in enrollment_row

        grades_csv = zf.read("all_grades.csv").decode("utf-8-sig")
        grade_row = next(r for r in csv.reader(StringIO(grades_csv)) if r and r[1] == str(student["id"]))
        assert student_name in grade_row
        assert course["course_name"] in grade_row

        performance_csv = zf.read("daily_performance.csv").decode("utf-8-sig")
        performance_row = next(
            r for r in csv.reader(StringIO(performance_csv)) if r and r[1] == str(student["id"])
        )
        assert student_name in performance_row
        assert course["course_name"] in performance_row
