"""
Custom report generation service for Phase 6 Reporting Enhancements.

Handles report data retrieval and export to PDF/Excel/CSV.
Uses a background-task-friendly entrypoint that opens its own DB session.
"""

from __future__ import annotations

import csv
import logging
import os
import re
import time
from datetime import date, datetime, timezone
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session, joinedload
from xml.sax.saxutils import escape

try:
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_CENTER
    from reportlab.lib.pagesizes import landscape, letter
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import inch
    from reportlab.platypus import Flowable, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False
    Flowable = Any  # type: ignore

from backend.config import settings
from backend.models import Attendance, Course, DailyPerformance, GeneratedReport, Grade, Report, Student, User
from backend.services.email_notification_service import EmailNotificationService
from backend.services.report_exporters import register_report_fonts

logger = logging.getLogger(__name__)


class CustomReportGenerationService:
    """Generate report files for custom report definitions."""

    STUDENT_IDENTIFIER_REPORT_TYPES = {"student", "grade", "attendance", "daily_performance"}
    RELATIONAL_STUDENT_REPORT_TYPES = {"grade", "attendance", "daily_performance"}
    COMPUTED_STUDENT_FIELDS = {
        "attendance_rate",
        "total_classes",
        "attended",
        "gpa",
        "passed_courses",
        "failed_courses",
    }

    LEGACY_FIELD_ALIASES = {
        "grade_value": "grade",
        "exam_date": "date_submitted",
        "attendance_percentage": "attendance_rate",
        "attendance_percent": "attendance_rate",
        "year_of_study": "study_year",
        "enrollment_status": "is_active",
    }

    REPORT_TYPE_FIELD_ALIASES: Dict[str, Dict[str, str]] = {
        "course": {
            "name": "course_name",
            "code": "course_code",
        },
    }

    FIELD_LABELS = {
        "first_name": {"en": "First Name", "el": "Όνομα"},
        "last_name": {"en": "Last Name", "el": "Επώνυμο"},
        "student_name": {"en": "Student Name", "el": "Όνομα Μαθητή"},
        "student_id": {"en": "Student ID", "el": "Αριθμός Μητρώου"},
        "email": {"en": "Email", "el": "Email"},
        "is_active": {"en": "Active", "el": "Ενεργός"},
        "course_code": {"en": "Course Code", "el": "Κωδικός Μαθήματος"},
        "course_name": {"en": "Course Name", "el": "Όνομα Μαθήματος"},
        "course_title": {"en": "Course Title", "el": "Τίτλος Μαθήματος"},
        "semester": {"en": "Semester", "el": "Εξάμηνο"},
        "credits": {"en": "Credits", "el": "Μονάδες"},
        "periods_per_week": {"en": "Periods per Week", "el": "Περίοδοι ανά Εβδομάδα"},
        "assignment_name": {"en": "Assignment Name", "el": "Όνομα Εργασίας"},
        "category": {"en": "Category", "el": "Κατηγορία"},
        "grade": {"en": "Grade", "el": "Βαθμός"},
        "max_grade": {"en": "Max Grade", "el": "Μέγιστος Βαθμός"},
        "weight": {"en": "Weight", "el": "Βαρύτητα"},
        "date_assigned": {"en": "Date Assigned", "el": "Ημερομηνία Ανάθεσης"},
        "date_submitted": {"en": "Date Submitted", "el": "Ημερομηνία Υποβολής"},
        "grade_value": {"en": "Grade", "el": "Βαθμός"},
        "exam_date": {"en": "Exam Date", "el": "Ημερομηνία Εξέτασης"},
        "date": {"en": "Date", "el": "Ημερομηνία"},
        "status": {"en": "Status", "el": "Κατάσταση"},
        "attendance_rate": {"en": "Attendance Rate", "el": "Ποσοστό Παρουσιών"},
        "total_classes": {"en": "Total Classes", "el": "Σύνολο Μαθημάτων"},
        "attended": {"en": "Attended", "el": "Παρουσίες"},
        "score": {"en": "Score", "el": "Βαθμολογία"},
        "max_score": {"en": "Max Score", "el": "Μέγιστη Βαθμολογία"},
        "percentage": {"en": "Percentage", "el": "Ποσοστό"},
        "study_year": {"en": "Study Year", "el": "Έτος Σπουδών"},
        "academic_year": {"en": "Academic Year", "el": "Ακαδημαϊκό Έτος"},
        "class_division": {"en": "Class Division", "el": "Τμήμα"},
        "gpa": {"en": "GPA", "el": "Μ.Ο. Βαθμολογίας"},
        "passed_courses": {"en": "Active Participation", "el": "Ενεργή Συμμετοχή"},
        "failed_courses": {"en": "Failed Courses", "el": "Αποτυχημένα Μαθήματα"},
    }

    # Translation dictionary for data values (categories, statuses, etc.)
    VALUE_TRANSLATIONS = {
        # Assignment categories (case-insensitive matching)
        "exam": {"en": "Exam", "el": "Εξέταση"},
        "midterm": {"en": "Midterm", "el": "Ενδιάμεση"},
        "midterm exam": {"en": "Midterm Exam", "el": "Ενδιάμεση Εξέταση"},
        "final": {"en": "Final", "el": "Τελική"},
        "final exam": {"en": "Final Exam", "el": "Τελική Εξέταση"},
        "quiz": {"en": "Quiz", "el": "Κουίζ"},
        "test": {"en": "Test", "el": "Τεστ"},
        "assignment": {"en": "Assignment", "el": "Εργασία"},
        "homework": {"en": "Homework", "el": "Κατ' οίκον"},
        "homework/assignments": {"en": "Homework/Assignments", "el": "Εργασίες/Ασκήσεις"},
        "project": {"en": "Project", "el": "Έργο"},
        "lab": {"en": "Lab", "el": "Εργαστήριο"},
        "laboratory": {"en": "Laboratory", "el": "Εργαστήριο"},
        "participation": {"en": "Participation", "el": "Συμμετοχή"},
        "class participation": {"en": "Class Participation", "el": "Συμμετοχή στο Μάθημα"},
        "continuous assessment": {"en": "Continuous Assessment", "el": "Συνεχής Αξιολόγηση"},
        "presentation": {"en": "Presentation", "el": "Παρουσίαση"},
        "essay": {"en": "Essay", "el": "Δοκίμιο"},
        "report": {"en": "Report", "el": "Αναφορά"},
        "paper": {"en": "Paper", "el": "Εργασία"},
        "practical": {"en": "Practical", "el": "Πρακτική"},
        "oral": {"en": "Oral", "el": "Προφορική"},
        "written": {"en": "Written", "el": "Γραπτή"},
        # Attendance statuses
        "present": {"en": "Present", "el": "Παρών"},
        "absent": {"en": "Absent", "el": "Απών"},
        "late": {"en": "Late", "el": "Καθυστέρηση"},
        "excused": {"en": "Excused", "el": "Δικαιολογημένος"},
        "unexcused": {"en": "Unexcused", "el": "Αδικαιολόγητος"},
        "tardy": {"en": "Tardy", "el": "Καθυστέρηση"},
        # Common assignment name patterns
        "sample exam": {"en": "Sample Exam", "el": "Δοκιμαστική Εξέταση"},
        "practice exam": {"en": "Practice Exam", "el": "Πρακτική Εξέταση"},
        "mock exam": {"en": "Mock Exam", "el": "Δοκιμαστική Εξέταση"},
        "review": {"en": "Review", "el": "Επανάληψη"},
        "exercise": {"en": "Exercise", "el": "Άσκηση"},
        "worksheet": {"en": "Worksheet", "el": "Φύλλο Εργασίας"},
        # Semester/academic terms
        "fall": {"en": "Fall", "el": "Χειμερινό"},
        "spring": {"en": "Spring", "el": "Εαρινό"},
        "summer": {"en": "Summer", "el": "Θερινό"},
        "winter": {"en": "Winter", "el": "Χειμερινό"},
        "semester 1": {"en": "Semester 1", "el": "Εξάμηνο 1"},
        "semester 2": {"en": "Semester 2", "el": "Εξάμηνο 2"},
        # Academic status
        "active": {"en": "Active", "el": "Ενεργός"},
        "inactive": {"en": "Inactive", "el": "Ανενεργός"},
        "enrolled": {"en": "Enrolled", "el": "Εγγεγραμμένος"},
        "graduated": {"en": "Graduated", "el": "Απόφοιτος"},
        "withdrawn": {"en": "Withdrawn", "el": "Αποχώρησε"},
        "suspended": {"en": "Suspended", "el": "Αναστολή"},
        # Common grade/score terms
        "pass": {"en": "Pass", "el": "Επιτυχία"},
        "fail": {"en": "Fail", "el": "Αποτυχία"},
        "incomplete": {"en": "Incomplete", "el": "Ατελές"},
        "pending": {"en": "Pending", "el": "Εκκρεμεί"},
        "submitted": {"en": "Submitted", "el": "Υποβλήθηκε"},
        "graded": {"en": "Graded", "el": "Βαθμολογήθηκε"},
        "missing": {"en": "Missing", "el": "Λείπει"},
        # Course types
        "lecture": {"en": "Lecture", "el": "Διάλεξη"},
        "tutorial": {"en": "Tutorial", "el": "Φροντιστήριο"},
        "seminar": {"en": "Seminar", "el": "Σεμινάριο"},
        "workshop": {"en": "Workshop", "el": "Εργαστήριο"},
    }

    def __init__(self, db: Session):
        self.db = db
        self.reports_dir = os.path.join(os.path.dirname(__file__), "..", "reports")
        os.makedirs(self.reports_dir, exist_ok=True)
        self._allow_sensitive_fields = False
        self._language = "en"

    @staticmethod
    def _normalize_language(value: Optional[str]) -> str:
        if not value:
            return "en"
        normalized = value.strip().lower()
        if normalized.startswith("el"):
            return "el"
        return "en"

    def _set_language(self, value: Optional[str]) -> None:
        self._language = self._normalize_language(value)

    def _translate_value(self, value: Any, field_name: str = "") -> Any:
        """
        Translate data values like category names, assignment names, status values.
        Word-by-word translation with support for multi-word phrases.
        """
        if not isinstance(value, str) or not value:
            return value

        # Only translate specific fields
        translatable_fields = {"category", "status", "assignment_name"}
        if field_name.lower() not in translatable_fields:
            return value

        # Keep English if language is English
        if self._language == "en":
            return value

        value_lower = value.strip().lower()

        # Check for exact full match first
        if value_lower in self.VALUE_TRANSLATIONS:
            translation = self.VALUE_TRANSLATIONS[value_lower].get(self._language)
            if translation:
                return translation

        # For assignment_name and category, do word-by-word translation
        # Sort translation keys by length (descending) to match multi-word phrases first
        sorted_keys = sorted(self.VALUE_TRANSLATIONS.keys(), key=len, reverse=True)

        result = value
        result_lower = value_lower

        # Track what's been replaced to avoid double-translation
        already_replaced_positions: set[tuple[int, str]] = set()

        for key in sorted_keys:
            if key in result_lower:
                # Find all occurrences of this key
                start = 0
                while True:
                    pos = result_lower.find(key, start)
                    if pos == -1:
                        break

                    # Check if this position overlaps with already replaced text
                    if any(pos <= p < pos + len(key) or p <= pos < p + len(k) for p, k in already_replaced_positions):
                        start = pos + 1
                        continue

                    # Check word boundaries (don't replace partial words)
                    before_ok = pos == 0 or not result_lower[pos - 1].isalnum()
                    after_ok = pos + len(key) >= len(result_lower) or not result_lower[pos + len(key)].isalnum()

                    if before_ok and after_ok:
                        # Get the translation
                        translated = self.VALUE_TRANSLATIONS[key].get(self._language, key)

                        # Preserve capitalization style from original
                        original_fragment = result[pos : pos + len(key)]
                        if original_fragment.isupper():
                            # All caps → keep translation as-is
                            pass
                        elif original_fragment[0].isupper():
                            # Title case → capitalize first letter
                            translated = (
                                translated[0].upper() + translated[1:] if len(translated) > 1 else translated.upper()
                            )

                        # Replace in both result and result_lower
                        result = result[:pos] + translated + result[pos + len(key) :]
                        result_lower = result_lower[:pos] + translated.lower() + result_lower[pos + len(key) :]

                        # Mark this position as replaced
                        already_replaced_positions.add((pos, key))

                        # Move start forward
                        start = pos + len(translated)
                    else:
                        start = pos + 1

        return result

    @staticmethod
    def _normalize_field_key(field: str) -> str:
        raw = (field or "").strip()
        if not raw:
            return ""

        def _normalize_segment(segment: str) -> str:
            normalized = segment.strip().replace("-", "_").replace(" ", "_")
            normalized = re.sub(r"(.)([A-Z][a-z]+)", r"\1_\2", normalized)
            normalized = re.sub(r"([a-z0-9])([A-Z])", r"\1_\2", normalized)
            normalized = re.sub(r"_+", "_", normalized)
            return normalized.lower()

        if "." in raw:
            return ".".join(_normalize_segment(part) for part in raw.split("."))
        return _normalize_segment(raw)

    @classmethod
    def _canonical_field_key(cls, field: str) -> str:
        normalized = cls._normalize_field_key(field)
        return cls.LEGACY_FIELD_ALIASES.get(normalized, normalized)

    @classmethod
    def _normalize_report_field_key(cls, report_type: str, field: str) -> str:
        canonical = cls._canonical_field_key(field)
        normalized_report_type = (report_type or "").strip().lower()
        report_aliases = cls.REPORT_TYPE_FIELD_ALIASES.get(normalized_report_type, {})
        canonical = report_aliases.get(canonical, canonical)
        if canonical == "id" and normalized_report_type in cls.STUDENT_IDENTIFIER_REPORT_TYPES:
            return "student_id"
        return canonical

    @classmethod
    def _is_sensitive_field(cls, field: str) -> bool:
        sensitive = {
            "health_issue",
            "note",
            "mobile_phone",
            "phone",
            "email",
        }
        normalized = cls._normalize_field_key(field)
        if normalized in sensitive:
            return True
        if "." in normalized:
            suffix = normalized.split(".")[-1]
            return suffix in sensitive
        return False

    def _set_sensitive_field_access(self, user_id: int) -> None:
        user = self.db.query(User).filter(User.id == user_id).first()
        self._allow_sensitive_fields = bool(user and getattr(user, "role", None) == "admin")

    @staticmethod
    def run_generation_task(
        report_id: int,
        generated_report_id: int,
        user_id: int,
        export_format: str,
        include_charts: bool,
        email_recipients: Optional[List[str]] = None,
        email_enabled: Optional[bool] = None,
        language: Optional[str] = None,
    ) -> None:
        """Background task entrypoint with isolated DB session."""
        from backend.db import SessionLocal

        db = SessionLocal()
        try:
            service = CustomReportGenerationService(db)
            service.generate_report(
                report_id,
                generated_report_id,
                user_id,
                export_format,
                include_charts,
                email_recipients=email_recipients,
                email_enabled=email_enabled,
                language=language,
            )
        finally:
            db.close()

    def generate_report(
        self,
        report_id: int,
        generated_report_id: int,
        user_id: int,
        export_format: str,
        include_charts: bool,
        email_recipients: Optional[List[str]] = None,
        email_enabled: Optional[bool] = None,
        language: Optional[str] = None,
    ) -> None:
        """Generate a report file and update the GeneratedReport record."""
        start_time = time.perf_counter()
        generated = self._get_generated_report(report_id, generated_report_id, user_id)
        if not generated:
            logger.error("Generated report record not found: %s", generated_report_id)
            return

        report = self._get_report(report_id, user_id)
        if not report:
            self._mark_failed(generated, "Report not found or access denied")
            return

        self._set_sensitive_field_access(user_id)
        self._set_language(language)

        self._update_status(generated, "generating")

        try:
            rows, headers = self._build_report_rows(report)

            # Extract group_by from report fields
            group_by_field = None
            group_by_col_index = None
            raw_report_fields = report.fields
            report_fields: Dict[str, Any] = raw_report_fields if isinstance(raw_report_fields, dict) else {}
            if report_fields:
                group_by_field = report_fields.get("group_by")
            if group_by_field and isinstance(group_by_field, str):
                # Find the column index matching the group_by field
                report_type_str = str(report.report_type or "").lower()
                normalized_group_by = self._normalize_report_field_key(report_type_str, group_by_field)
                columns = self._normalize_columns(report_type_str, report.fields)
                for idx, (key, _label) in enumerate(columns):
                    if key == normalized_group_by:
                        group_by_col_index = idx
                        break

            file_path = self._export_report(
                rows,
                headers,
                export_format,
                str(generated.file_name),
                title=str(report.name) if report.name else None,
                group_by_col_index=group_by_col_index,
            )  # type: ignore[arg-type]
            duration = time.perf_counter() - start_time

            generated.file_path = file_path  # type: ignore[assignment]
            generated.file_size_bytes = os.path.getsize(file_path) if file_path else None  # type: ignore[assignment]
            generated.record_count = len(rows)  # type: ignore[assignment]
            generated.generation_duration_seconds = round(float(duration), 3)  # type: ignore[arg-type,assignment]
            generated.status = "completed"  # type: ignore[assignment]
            generated.error_message = None  # type: ignore[assignment]

            report.last_run_at = datetime.now(timezone.utc)  # type: ignore[assignment]
            self._supersede_prior_generated_reports(report_id, user_id, generated_report_id)
            self.db.commit()

            self._send_report_email(
                report=report,
                generated=generated,
                export_format=export_format,
                email_recipients=email_recipients,
                email_enabled=email_enabled,
            )
        except Exception as exc:
            logger.error("Report generation failed: %s", exc, exc_info=True)
            self._mark_failed(generated, str(exc))

    def _send_report_email(
        self,
        report: Report,
        generated: GeneratedReport,
        export_format: str,
        email_recipients: Optional[List[str]] = None,
        email_enabled: Optional[bool] = None,
    ) -> None:
        recipients = email_recipients if email_recipients is not None else report.email_recipients
        if isinstance(recipients, str):
            recipients = [recipients]
        if not recipients:
            return

        send_enabled = email_enabled if email_enabled is not None else bool(report.email_enabled)
        if not send_enabled:
            return

        if not EmailNotificationService.is_enabled():
            logger.warning("Email delivery skipped: SMTP not configured")
            return

        file_path = str(generated.file_path or "")
        attachment_paths: list[str] = []
        attachment_note: Optional[str] = None
        max_mb = int(getattr(settings, "SMTP_ATTACHMENT_MAX_MB", 10) or 10)
        max_bytes = max_mb * 1024 * 1024

        if file_path and os.path.exists(file_path):
            try:
                file_size = os.path.getsize(file_path)
                if file_size <= max_bytes:
                    attachment_paths = [file_path]
                else:
                    attachment_note = (
                        f"Attachment not included because the file exceeds {max_mb} MB. "
                        "You can download it from the reports page."
                    )
            except Exception as exc:
                logger.warning("Failed to prepare attachment: %s", exc)
        else:
            attachment_note = "Attachment not found on disk. You can download it from the reports page."

        generated_at = generated.generated_at.isoformat() if generated.generated_at else ""
        success_count, failed = EmailNotificationService.send_report_ready(
            recipients=recipients,
            report_name=str(report.name),
            export_format=export_format,
            record_count=int(generated.record_count) if generated.record_count is not None else None,
            generated_at=generated_at,
            attachment_paths=attachment_paths or None,
            attachment_note=attachment_note,
        )

        if failed:
            generated.email_sent = False  # type: ignore[assignment]
            generated.email_error = f"Failed to send to: {', '.join(failed)}"  # type: ignore[assignment]
        else:
            generated.email_sent = True  # type: ignore[assignment]
            generated.email_error = None  # type: ignore[assignment]
        if success_count > 0:
            generated.email_sent_at = datetime.now(timezone.utc)  # type: ignore[assignment]
        self.db.commit()

    def _get_report(self, report_id: int, user_id: int) -> Optional[Report]:
        return (
            self.db.query(Report)
            .filter(Report.id == report_id, Report.user_id == user_id, Report.deleted_at.is_(None))
            .first()
        )

    def _get_generated_report(self, report_id: int, generated_id: int, user_id: int) -> Optional[GeneratedReport]:
        return (
            self.db.query(GeneratedReport)
            .filter(
                GeneratedReport.id == generated_id,
                GeneratedReport.report_id == report_id,
                GeneratedReport.user_id == user_id,
            )
            .first()
        )

    def _update_status(self, generated: GeneratedReport, status: str) -> None:
        generated.status = status  # type: ignore[assignment]
        self.db.commit()

    def _mark_failed(self, generated: GeneratedReport, message: str) -> None:
        generated.status = "failed"  # type: ignore[assignment]
        generated.error_message = message  # type: ignore[assignment]
        self.db.commit()

    def _supersede_prior_generated_reports(self, report_id: int, user_id: int, generated_report_id: int) -> None:
        superseded_at = datetime.now(timezone.utc).isoformat()
        prior_completed_reports = (
            self.db.query(GeneratedReport)
            .filter(
                GeneratedReport.report_id == report_id,
                GeneratedReport.user_id == user_id,
                GeneratedReport.id != generated_report_id,
                GeneratedReport.status == "completed",
            )
            .all()
        )

        for previous_report in prior_completed_reports:
            previous_report.status = "superseded"  # type: ignore[assignment]
            previous_report.error_message = f"Superseded by generated report {generated_report_id} at {superseded_at}"  # type: ignore[assignment]

    def _build_report_rows(self, report: Report) -> Tuple[List[List[Any]], List[str]]:
        report_type = str(report.report_type or "").lower()
        model, query = self._build_query(report)
        # Handle filters as a list of filter objects or dict
        filters_list: list = report.filters or []  # type: ignore[assignment]
        sort_list: list = report.sort_by or []  # type: ignore[assignment]

        query = self._apply_filters(query, model, filters_list, report_type)
        query = self._apply_sort(query, model, sort_list, report_type)

        columns = self._normalize_columns(report_type, report.fields)
        if not self._allow_sensitive_fields:
            columns = [(key, label) for key, label in columns if not self._is_sensitive_field(key)]
        headers = [label for _, label in columns]

        records = query.all()
        records = self._apply_post_filters(records, filters_list, report_type)
        records = self._apply_post_sort(records, sort_list, report_type)

        rows: List[List[Any]] = []
        for record in records:
            row = [self._resolve_field(record, key) for key, _ in columns]
            rows.append(row)

        return rows, headers

    def _build_query(self, report: Report):
        report_type = str(report.report_type or "").lower()

        if report_type == "student":
            query = (
                self.db.query(Student)
                .options(
                    joinedload(Student.attendances),
                    joinedload(Student.grades),
                    joinedload(Student.enrollments),
                )
                .filter(Student.deleted_at.is_(None))
            )
            return Student, query

        if report_type == "course":
            query = (
                self.db.query(Course)
                .options(
                    joinedload(Course.enrollments),
                )
                .filter(Course.deleted_at.is_(None))
            )
            return Course, query

        if report_type == "grade":
            query = (
                self.db.query(Grade)
                .options(joinedload(Grade.student), joinedload(Grade.course))
                .filter(Grade.deleted_at.is_(None))
            )
            return Grade, query

        if report_type == "attendance":
            query = (
                self.db.query(Attendance)
                .options(joinedload(Attendance.student), joinedload(Attendance.course))
                .filter(Attendance.deleted_at.is_(None))
            )
            return Attendance, query

        if report_type == "daily_performance":
            query = (
                self.db.query(DailyPerformance)
                .options(joinedload(DailyPerformance.student), joinedload(DailyPerformance.course))
                .filter(DailyPerformance.deleted_at.is_(None))
            )
            return DailyPerformance, query

        raise ValueError(f"Unsupported report type: {report.report_type}")

    def _can_apply_query_field(self, model: Any, report_type: str, field: str) -> bool:
        normalized_report_type = (report_type or "").strip().lower()
        if normalized_report_type in self.RELATIONAL_STUDENT_REPORT_TYPES and field == "student_id":
            return False
        if normalized_report_type == "student" and field in self.COMPUTED_STUDENT_FIELDS:
            return False
        return hasattr(model, field)

    def _apply_filters(self, query, model, filters: Any, report_type: str) -> Any:  # type: ignore[no-untyped-def]
        """Apply filters to the query. Handles multiple filter formats."""
        if not filters:
            return query

        # Convert list of filter objects to proper queries
        if isinstance(filters, list):
            for filter_obj in filters:
                if not isinstance(filter_obj, dict):
                    continue

                field = filter_obj.get("field")
                operator = filter_obj.get("operator", "equals")
                value = filter_obj.get("value")
                normalized_field = self._normalize_report_field_key(report_type, str(field or ""))

                if not field or not self._can_apply_query_field(model, report_type, normalized_field):
                    continue

                column = getattr(model, normalized_field)

                # Apply operator-specific logic
                if operator == "equals":
                    query = query.filter(column == value)  # type: ignore[operator]
                elif operator == "not_equals":
                    query = query.filter(column != value)  # type: ignore[operator]
                elif operator == "contains":
                    query = query.filter(column.ilike(f"%{value}%"))  # type: ignore[operator]
                elif operator == "not_contains":
                    query = query.filter(~column.ilike(f"%{value}%"))  # type: ignore[operator]
                elif operator == "starts_with":
                    query = query.filter(column.ilike(f"{value}%"))  # type: ignore[operator]
                elif operator == "ends_with":
                    query = query.filter(column.ilike(f"%{value}"))  # type: ignore[operator]
                elif operator == "greater_than" or operator == "gt":
                    query = query.filter(column > value)  # type: ignore[operator]
                elif operator == "less_than" or operator == "lt":
                    query = query.filter(column < value)  # type: ignore[operator]
                elif operator == "greater_than_or_equal":
                    query = query.filter(column >= value)  # type: ignore[operator]
                elif operator == "less_than_or_equal":
                    query = query.filter(column <= value)  # type: ignore[operator]
                elif operator == "in":
                    if isinstance(value, (list, tuple)):
                        query = query.filter(column.in_(value))  # type: ignore[operator]
                elif operator == "between":
                    if isinstance(value, dict) and "from" in value and "to" in value:
                        query = query.filter(column.between(value["from"], value["to"]))  # type: ignore[operator]

        # Nested dict format: {"grade": {"operator": "less_than", "value": 60}}
        elif isinstance(filters, dict):
            for field_name, filter_spec in filters.items():
                normalized_field = self._normalize_report_field_key(report_type, str(field_name))
                # If value is a dict with operator/value, it's the new format
                if isinstance(filter_spec, dict) and "operator" in filter_spec and "value" in filter_spec:
                    operator = filter_spec.get("operator", "equals")
                    value = filter_spec.get("value")

                    if not self._can_apply_query_field(model, report_type, normalized_field):
                        continue

                    column = getattr(model, normalized_field)

                    # Apply operator-specific logic
                    if operator == "equals":
                        query = query.filter(column == value)  # type: ignore[operator]
                    elif operator == "not_equals":
                        query = query.filter(column != value)  # type: ignore[operator]
                    elif operator == "contains":
                        query = query.filter(column.ilike(f"%{value}%"))  # type: ignore[operator]
                    elif operator == "not_contains":
                        query = query.filter(~column.ilike(f"%{value}%"))  # type: ignore[operator]
                    elif operator == "starts_with":
                        query = query.filter(column.ilike(f"{value}%"))  # type: ignore[operator]
                    elif operator == "ends_with":
                        query = query.filter(column.ilike(f"%{value}"))  # type: ignore[operator]
                    elif operator == "greater_than" or operator == "gt":
                        query = query.filter(column > value)  # type: ignore[operator]
                    elif operator == "less_than" or operator == "lt":
                        query = query.filter(column < value)  # type: ignore[operator]
                    elif operator == "greater_than_or_equal":
                        query = query.filter(column >= value)  # type: ignore[operator]
                    elif operator == "less_than_or_equal":
                        query = query.filter(column <= value)  # type: ignore[operator]
                    elif operator == "in":
                        if isinstance(value, (list, tuple)):
                            query = query.filter(column.in_(value))  # type: ignore[operator]
                    elif operator == "between":
                        if isinstance(value, dict) and "from" in value and "to" in value:
                            query = query.filter(column.between(value["from"], value["to"]))  # type: ignore[operator]
                # Legacy simple format: {"field_name": value}
                else:
                    if self._can_apply_query_field(model, report_type, normalized_field):
                        column = getattr(model, normalized_field)
                        query = query.filter(column == filter_spec)  # type: ignore[operator]

        return query

    def _apply_sort(self, query, model, sort_by: Any, report_type: str) -> Any:  # type: ignore[no-untyped-def]
        """Apply sorting to the query. Handles both list and dict formats."""
        # Handle list of sort objects
        if isinstance(sort_by, list):
            for sort_obj in sort_by:
                if not isinstance(sort_obj, dict):
                    continue

                field = sort_obj.get("field")
                order = str(sort_obj.get("order", "asc")).lower()
                normalized_field = self._normalize_report_field_key(report_type, str(field or ""))

                if field and self._can_apply_query_field(model, report_type, normalized_field):
                    column = getattr(model, normalized_field)
                    query = query.order_by(column.desc() if order == "desc" else column.asc())  # type: ignore[operator]

        # Legacy dict format support
        elif isinstance(sort_by, dict):
            field = sort_by.get("field") if isinstance(sort_by, dict) else None
            direction = str(sort_by.get("direction", "asc")).lower() if isinstance(sort_by, dict) else "asc"
            normalized_field = self._normalize_report_field_key(report_type, str(field or ""))
            if field and self._can_apply_query_field(model, report_type, normalized_field):
                column = getattr(model, normalized_field)
                query = query.order_by(column.desc() if direction == "desc" else column.asc())  # type: ignore[operator]

        return query

    def _normalize_columns(self, report_type: str, fields: Any) -> List[Tuple[str, str]]:  # type: ignore[no-untyped-def]
        columns = self._extract_columns(fields)
        if columns:
            normalized_columns: List[Tuple[str, str]] = []
            for key, label in columns:
                normalized_key = self._normalize_report_field_key(report_type, key)
                original_default_label = self._label_from_key(key, self._language)
                normalized_label = label
                if not label or label == original_default_label:
                    normalized_label = self._label_from_key(normalized_key, self._language)
                normalized_columns.append((normalized_key, normalized_label))
            return normalized_columns

        defaults: Dict[str, List[str]] = {
            "student": ["first_name", "last_name", "email", "student_id", "is_active"],
            "course": ["course_code", "course_name", "semester", "credits", "periods_per_week"],
            "grade": ["student_name", "course_code", "grade", "max_grade", "date_assigned"],
            "attendance": ["student_name", "course_code", "date", "status"],
            "daily_performance": [
                "student_name",
                "course_code",
                "date",
                "category",
                "score",
                "max_score",
                "percentage",
            ],
        }
        fallback = defaults.get(report_type.lower(), [])
        return [(key, self._label_from_key(key, self._language)) for key in fallback]

    def _get_student_attendance_metrics(self, student: Student) -> Dict[str, float | int]:
        attendance_records = [
            attendance
            for attendance in getattr(student, "attendances", []) or []
            if getattr(attendance, "deleted_at", None) is None
        ]

        total_classes = len(attendance_records)
        attended = sum(
            1
            for attendance in attendance_records
            if str(getattr(attendance, "status", "")).lower() in {"present", "late"}
        )
        attendance_rate = round((attended / total_classes) * 100, 1) if total_classes > 0 else 0.0

        return {
            "attendance_rate": attendance_rate,
            "total_classes": total_classes,
            "attended": attended,
        }

    def _get_student_grade_metrics(self, student: Student) -> Dict[str, float | int]:
        grade_records = [
            grade for grade in getattr(student, "grades", []) or [] if getattr(grade, "deleted_at", None) is None
        ]

        percentages: List[float] = []
        passed_courses = 0
        failed_courses = 0

        for grade in grade_records:
            grade_value = getattr(grade, "grade", None)
            max_grade = getattr(grade, "max_grade", None)
            if grade_value is None or not max_grade:
                continue
            try:
                percentage = (float(grade_value) / float(max_grade)) * 100
            except Exception:
                continue

            percentages.append(percentage)
            if percentage >= 50:
                passed_courses += 1
            else:
                failed_courses += 1

        gpa = round(sum(percentages) / len(percentages), 1) if percentages else 0.0
        return {
            "gpa": gpa,
            "passed_courses": passed_courses,
            "failed_courses": failed_courses,
        }

    def _get_course_metrics(self, course: Course) -> Dict[str, int]:
        enrollments = [
            enrollment
            for enrollment in getattr(course, "enrollments", []) or []
            if getattr(enrollment, "deleted_at", None) is None
        ]
        return {
            "enrollment_count": len(enrollments),
        }

    def _get_resolved_filter_value(self, record: Any, report_type: str, field: str) -> Any:
        normalized_field = self._normalize_report_field_key(report_type, field)
        return self._resolve_field(record, normalized_field)

    @staticmethod
    def _coerce_comparable_value(value: Any) -> Any:
        if value is None or value == "":
            return ""
        if isinstance(value, str):
            stripped = value.strip()
            try:
                return float(stripped)
            except ValueError:
                return stripped.lower()
        return value

    def _matches_filter(self, record: Any, report_type: str, field: str, operator: str, expected: Any) -> bool:
        actual = self._get_resolved_filter_value(record, report_type, field)
        normalized_operator = str(operator or "equals").lower()

        actual_cmp = self._coerce_comparable_value(actual)
        expected_cmp = self._coerce_comparable_value(expected)

        if normalized_operator in {"equals", "eq"}:
            return actual_cmp == expected_cmp
        if normalized_operator in {"not_equals", "ne"}:
            return actual_cmp != expected_cmp
        if normalized_operator == "contains":
            return str(expected_cmp) in str(actual_cmp)
        if normalized_operator == "not_contains":
            return str(expected_cmp) not in str(actual_cmp)
        if normalized_operator == "starts_with":
            return str(actual_cmp).startswith(str(expected_cmp))
        if normalized_operator == "ends_with":
            return str(actual_cmp).endswith(str(expected_cmp))
        if normalized_operator in {"greater_than", "gt"}:
            return actual_cmp > expected_cmp
        if normalized_operator in {"greater_than_or_equal", "gte"}:
            return actual_cmp >= expected_cmp
        if normalized_operator in {"less_than", "lt"}:
            return actual_cmp < expected_cmp
        if normalized_operator in {"less_than_or_equal", "lte"}:
            return actual_cmp <= expected_cmp
        if normalized_operator == "in":
            return actual_cmp in (expected_cmp if isinstance(expected_cmp, (list, tuple, set)) else [expected_cmp])
        return True

    def _apply_post_filters(self, records: List[Any], filters: Any, report_type: str) -> List[Any]:
        if not filters:
            return records

        normalized_filters: List[Tuple[str, str, Any]] = []
        if isinstance(filters, list):
            for filter_obj in filters:
                if isinstance(filter_obj, dict) and filter_obj.get("field"):
                    normalized_filters.append(
                        (
                            str(filter_obj.get("field")),
                            str(filter_obj.get("operator", "equals")),
                            filter_obj.get("value"),
                        )
                    )
        elif isinstance(filters, dict):
            for field_name, filter_spec in filters.items():
                if isinstance(filter_spec, dict) and "value" in filter_spec:
                    normalized_filters.append(
                        (str(field_name), str(filter_spec.get("operator", "equals")), filter_spec.get("value"))
                    )
                else:
                    normalized_filters.append((str(field_name), "equals", filter_spec))

        filtered_records = records
        for field, operator, value in normalized_filters:
            filtered_records = [
                record
                for record in filtered_records
                if self._matches_filter(record, report_type, field, operator, value)
            ]
        return filtered_records

    def _apply_post_sort(self, records: List[Any], sort_by: Any, report_type: str) -> List[Any]:
        if not sort_by:
            return records

        sort_rules: List[Tuple[str, str]] = []
        if isinstance(sort_by, list):
            for sort_obj in sort_by:
                if isinstance(sort_obj, dict) and sort_obj.get("field"):
                    sort_rules.append((str(sort_obj.get("field")), str(sort_obj.get("order", "asc")).lower()))
        elif isinstance(sort_by, dict) and sort_by.get("field"):
            sort_rules.append((str(sort_by.get("field")), str(sort_by.get("direction", "asc")).lower()))

        sorted_records = list(records)
        for field, direction in reversed(sort_rules):
            sorted_records.sort(
                key=lambda record: self._coerce_comparable_value(
                    self._get_resolved_filter_value(record, report_type, field)
                ),
                reverse=direction == "desc",
            )
        return sorted_records

    def _extract_columns(self, fields: Any) -> List[Tuple[str, str]]:
        columns: Sequence[Any] = []
        if isinstance(fields, dict):
            columns = fields.get("columns") or fields.get("fields") or []
        elif isinstance(fields, list):
            columns = fields

        result: List[Tuple[str, str]] = []
        for col in columns:
            if isinstance(col, dict):
                key = col.get("key") or col.get("field") or col.get("name")
                if not key:
                    continue
                label = col.get("label") or ""
                resolved_label = self._localize_label(str(key), str(label))
                result.append((str(key), resolved_label))
            else:
                key = str(col)
                result.append((key, self._label_from_key(key, self._language)))
        return result

    def _label_from_key(self, key: str, language: Optional[str] = None) -> str:
        normalized = self._normalize_field_key(key)
        canonical_key = self._canonical_field_key(key)
        labels = self.FIELD_LABELS.get(normalized)
        if not labels and canonical_key != normalized:
            labels = self.FIELD_LABELS.get(canonical_key)
        default_label = canonical_key.replace("_", " ").replace(".", " ").title()
        if not labels:
            return default_label

        target_language = self._normalize_language(language)
        if target_language == "el":
            return labels.get("el") or labels.get("en") or default_label
        return labels.get("en") or default_label

    def _localize_label(self, key: str, label: str) -> str:
        if not label:
            return self._label_from_key(key, self._language)

        if self._normalize_language(self._language) != "el":
            return label

        english_default = self._label_from_key(key, "en")
        normalized_label = label.strip().lower()
        normalized_key = self._normalize_field_key(key)
        canonical_key = self._canonical_field_key(key)
        normalized_candidates = {
            normalized_key,
            normalized_key.replace("_", " ").replace(".", " "),
            canonical_key,
            canonical_key.replace("_", " ").replace(".", " "),
            english_default.strip().lower(),
        }
        if normalized_label in normalized_candidates:
            return self._label_from_key(key, self._language)
        if normalized_label == english_default.strip().lower():
            return self._label_from_key(key, "el")
        return label

    @staticmethod
    def _format_temporal_value(value: Any) -> Any:
        if isinstance(value, (datetime, date)):
            return value.isoformat()
        return value

    def _resolve_field(self, record: Any, field: str) -> Any:
        """Resolve a field value from a record, handling relationships and nested fields."""

        canonical_field = self._canonical_field_key(field)

        if not self._allow_sensitive_fields and self._is_sensitive_field(field):
            return ""

        if canonical_field == "percentage":
            value = None
            if hasattr(record, "score") and hasattr(record, "max_score"):
                score = getattr(record, "score", None)
                max_score = getattr(record, "max_score", None)
                if score is not None and max_score:
                    try:
                        value = (float(score) / float(max_score)) * 100
                    except Exception:
                        value = None
            elif hasattr(record, "grade") and hasattr(record, "max_grade"):
                grade = getattr(record, "grade", None)
                max_grade = getattr(record, "max_grade", None)
                if grade is not None and max_grade:
                    try:
                        value = (float(grade) / float(max_grade)) * 100
                    except Exception:
                        value = None
            elif hasattr(record, "percentage"):
                try:
                    value = float(getattr(record, "percentage"))
                except Exception:
                    value = None

            if value is None:
                return ""
            return f"{value:.2f}%"

        # Special handling for computed fields
        if canonical_field == "student_name":
            student = getattr(record, "student", None)
            if student:
                return f"{student.first_name} {student.last_name}".strip()
            if hasattr(record, "first_name") and hasattr(record, "last_name"):
                return f"{record.first_name} {record.last_name}".strip()
            return ""

        if canonical_field == "student_id":
            student = getattr(record, "student", None)
            if student and hasattr(student, "student_id"):
                external_student_id = getattr(student, "student_id", None)
                if external_student_id is not None:
                    return external_student_id
            if hasattr(record, "student_id"):
                return getattr(record, "student_id", "") or ""
            return ""

        if isinstance(record, Student) and canonical_field in self.COMPUTED_STUDENT_FIELDS:
            attendance_metrics = self._get_student_attendance_metrics(record)
            if canonical_field in attendance_metrics:
                return attendance_metrics.get(canonical_field, "")

            grade_metrics = self._get_student_grade_metrics(record)
            return grade_metrics.get(canonical_field, "")

        if isinstance(record, Course) and canonical_field == "enrollment_count":
            course_metrics = self._get_course_metrics(record)
            return course_metrics.get(canonical_field, "")

        if canonical_field == "course_code":
            course = getattr(record, "course", None)
            if course and hasattr(course, "course_code"):
                return course.course_code
            if hasattr(record, "course_code"):
                return record.course_code
            return ""

        if canonical_field == "course_name":
            course = getattr(record, "course", None)
            if course and hasattr(course, "course_name"):
                return course.course_name
            if hasattr(record, "course_name"):
                return record.course_name
            return ""

        if canonical_field == "date_submitted":
            submitted_value = getattr(record, "date_submitted", None)
            if submitted_value is None:
                assigned_value = getattr(record, "date_assigned", None)
                if isinstance(assigned_value, (datetime, date)):
                    return self._format_temporal_value(assigned_value)
                if assigned_value is not None:
                    return assigned_value
                return ""
            return self._format_temporal_value(submitted_value)

        # Handle nested dot notation
        if "." in canonical_field:
            value = record
            for part in canonical_field.split("."):
                value = getattr(value, part, None)
                if value is None:
                    return ""
            return value

        # Try to get directly from record
        if hasattr(record, canonical_field):
            value = getattr(record, canonical_field, None)
            if value is None:
                return ""
            value = self._format_temporal_value(value)
            # Translate value if applicable (category, status, assignment_name)
            return self._translate_value(value, canonical_field)

        # For Grade records, try to resolve from related models
        # e.g., first_name -> student.first_name, last_name -> student.last_name
        if hasattr(record, "student"):
            student = getattr(record, "student", None)
            if student and hasattr(student, canonical_field):
                value = getattr(student, canonical_field, None)
                if value is None:
                    return ""
                value = self._format_temporal_value(value)
                return self._translate_value(value, canonical_field)

        # Try Course relationship
        if hasattr(record, "course"):
            course = getattr(record, "course", None)
            if course and hasattr(course, canonical_field):
                value = getattr(course, canonical_field, None)
                if value is None:
                    return ""
                value = self._format_temporal_value(value)
                return self._translate_value(value, canonical_field)

        return ""

    def _export_report(
        self,
        rows: List[List[Any]],
        headers: List[str],
        export_format: str,
        file_name: str,
        title: Optional[str] = None,
        group_by_col_index: Optional[int] = None,
    ) -> str:  # type: ignore[override]
        format_lower = (export_format or "pdf").lower()
        file_path = os.path.join(self.reports_dir, file_name)

        # Sort rows by group column so groups are contiguous
        if group_by_col_index is not None and rows:
            rows = sorted(rows, key=lambda r: str(r[group_by_col_index]) if group_by_col_index < len(r) else "")

        if format_lower == "csv":
            self._export_csv(rows, headers, file_path, group_by_col_index=group_by_col_index)
            return file_path

        if format_lower == "excel":
            self._export_excel(rows, headers, file_path, group_by_col_index=group_by_col_index)
            return file_path

        if format_lower == "pdf":
            self._export_pdf(rows, headers, file_path, title=title, group_by_col_index=group_by_col_index)
            return file_path

        raise ValueError(f"Unsupported export format: {export_format}")

    def _export_csv(
        self,
        rows: Iterable[Sequence[Any]],
        headers: Sequence[str],
        file_path: str,
        group_by_col_index: Optional[int] = None,
    ) -> None:
        with open(file_path, "w", newline="", encoding="utf-8") as handle:
            writer = csv.writer(handle)
            writer.writerow(headers)
            current_group = None
            for row in rows:
                if group_by_col_index is not None and group_by_col_index < len(row):
                    group_val = self._stringify_value(row[group_by_col_index])
                    if group_val != current_group:
                        current_group = group_val
                        group_label = f"--- {headers[group_by_col_index]}: {group_val} ---"
                        writer.writerow([group_label] + [""] * (len(headers) - 1))
                writer.writerow([self._stringify_value(value) for value in row])

    def _export_excel(
        self,
        rows: Iterable[Sequence[Any]],
        headers: Sequence[str],
        file_path: str,
        group_by_col_index: Optional[int] = None,
    ) -> None:
        wb = Workbook()
        ws = wb.active
        ws.title = "Report"

        header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        group_fill = PatternFill(start_color="E0E7FF", end_color="E0E7FF", fill_type="solid")
        group_font = Font(bold=True, color="1E40AF")

        for col_index, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_index, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        current_row = 2
        current_group = None
        for row in rows:
            if group_by_col_index is not None and group_by_col_index < len(row):
                group_val = self._stringify_value(row[group_by_col_index])
                if group_val != current_group:
                    current_group = group_val
                    group_label = f"{headers[group_by_col_index]}: {group_val}"
                    cell = ws.cell(row=current_row, column=1, value=group_label)
                    cell.font = group_font
                    cell.fill = group_fill
                    ws.merge_cells(
                        start_row=current_row,
                        start_column=1,
                        end_row=current_row,
                        end_column=len(headers),
                    )
                    current_row += 1

            for col_index, value in enumerate(row, 1):
                ws.cell(row=current_row, column=col_index, value=self._stringify_value(value))
            current_row += 1

        for col_index, header in enumerate(headers, 1):
            width = max(len(str(header)), 10)
            ws.column_dimensions[get_column_letter(col_index)].width = max(width, 12)

        wb.save(file_path)

    def _export_pdf(
        self,
        rows: List[List[Any]],
        headers: List[str],
        file_path: str,
        title: Optional[str] = None,
        group_by_col_index: Optional[int] = None,
    ) -> None:
        if not REPORTLAB_AVAILABLE:
            raise ImportError("ReportLab is required for PDF generation. Install with: pip install reportlab")

        page_size = letter
        if len(headers) > 8:
            page_size = landscape(letter)

        doc = SimpleDocTemplate(
            file_path,
            pagesize=page_size,
            leftMargin=0.5 * inch,
            rightMargin=0.5 * inch,
            topMargin=0.5 * inch,
            bottomMargin=0.5 * inch,
        )
        styles = getSampleStyleSheet()
        base_font, base_font_bold = register_report_fonts()
        body_font_size = 10
        if len(headers) > 6:
            body_font_size = 9
        if len(headers) > 8:
            body_font_size = 8
        if len(headers) > 10:
            body_font_size = 7

        title_style = ParagraphStyle(
            "ReportTitle",
            parent=styles["Heading1"],
            fontSize=16,
            textColor=colors.HexColor("#1f77b4"),
            spaceAfter=12,
            fontName=base_font_bold,
            alignment=TA_CENTER,
        )
        header_style = ParagraphStyle(
            "ReportHeader",
            parent=styles["BodyText"],
            fontName=base_font_bold,
            fontSize=body_font_size,
            leading=body_font_size + 2,
            textColor=colors.whitesmoke,
            alignment=TA_CENTER,
        )
        cell_style = ParagraphStyle(
            "ReportCell",
            parent=styles["BodyText"],
            fontName=base_font,
            fontSize=body_font_size,
            leading=body_font_size + 2,
        )
        group_style = ParagraphStyle(
            "GroupHeader",
            parent=styles["BodyText"],
            fontName=base_font_bold,
            fontSize=body_font_size,
            leading=body_font_size + 2,
            textColor=colors.HexColor("#1E40AF"),
        )

        report_title = title
        if not report_title:
            report_title = "Προσαρμοσμένη Αναφορά" if self._language == "el" else "Custom Report"
        elements: List[Flowable] = [Paragraph(report_title, title_style), Spacer(1, 0.2 * inch)]

        safe_headers: List[Flowable] = [Paragraph(escape(str(header)), header_style) for header in headers]
        col_count = max(len(headers), 1)

        # Build table_data with optional group header rows
        table_data: List[List[Flowable]] = [safe_headers]
        group_row_indices: List[int] = []  # 0-based indices in table_data
        current_group = None

        for row in rows:
            if group_by_col_index is not None and group_by_col_index < len(row):
                group_val = str(self._stringify_value(row[group_by_col_index]))
                if group_val != current_group:
                    current_group = group_val
                    group_label = f"{headers[group_by_col_index]}: {group_val}"
                    group_row: List[Flowable] = [Paragraph(escape(group_label), group_style)]
                    group_row += [Paragraph("", cell_style)] * (col_count - 1)
                    group_row_indices.append(len(table_data))
                    table_data.append(group_row)

            safe_row: List[Flowable] = [
                Paragraph(escape(str(self._stringify_value(value))), cell_style) for value in row
            ]
            table_data.append(safe_row)

        def _estimate_column_width(index: int) -> float:
            header_length = len(str(headers[index]))
            row_lengths = [len(str(self._stringify_value(row[index]))) for row in rows if index < len(row)]
            max_length = max([header_length, *row_lengths, 6])
            font_factor = body_font_size * 0.62
            return max(72.0, max_length * font_factor + 12)

        estimated_widths = [_estimate_column_width(index) for index in range(col_count)]
        total_estimated_width = sum(estimated_widths) or doc.width
        scale = min(1.0, doc.width / total_estimated_width)

        min_col_width = 20.0
        col_widths = [max(min_col_width, width * scale) for width in estimated_widths]

        total_width = sum(col_widths)
        if total_width > doc.width and total_width > 0:
            rescale = doc.width / total_width
            col_widths = [max(min_col_width, width * rescale) for width in col_widths]
            total_width = sum(col_widths)

        if total_width < doc.width:
            extra = (doc.width - total_width) / col_count
            col_widths = [width + extra for width in col_widths]

        table = Table(table_data, repeatRows=1, colWidths=col_widths)

        style_commands: list = [
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1f77b4")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
            ("ALIGN", (0, 0), (-1, 0), "CENTER"),
            ("ALIGN", (0, 1), (-1, -1), "LEFT"),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("FONTNAME", (0, 0), (-1, 0), base_font_bold),
            ("FONTNAME", (0, 1), (-1, -1), base_font),
            ("FONTSIZE", (0, 0), (-1, 0), body_font_size),
            ("FONTSIZE", (0, 1), (-1, -1), body_font_size),
            ("LEFTPADDING", (0, 0), (-1, -1), 6),
            ("RIGHTPADDING", (0, 0), (-1, -1), 6),
            ("BOTTOMPADDING", (0, 0), (-1, 0), 8),
            ("BOTTOMPADDING", (0, 1), (-1, -1), 6),
            ("TOPPADDING", (0, 1), (-1, -1), 6),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ]

        if group_row_indices:
            # Per-row alternating backgrounds, skipping group headers
            data_row_counter = 0
            for i in range(1, len(table_data)):
                if i in group_row_indices:
                    continue
                bg = colors.white if data_row_counter % 2 == 0 else colors.HexColor("#F3F4F6")
                style_commands.append(("BACKGROUND", (0, i), (-1, i), bg))
                data_row_counter += 1
            # Style group header rows
            for gi in group_row_indices:
                style_commands.append(("BACKGROUND", (0, gi), (-1, gi), colors.HexColor("#DBEAFE")))
                style_commands.append(("SPAN", (0, gi), (-1, gi)))
                style_commands.append(("FONTNAME", (0, gi), (-1, gi), base_font_bold))
        else:
            style_commands.append(("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F3F4F6")]))

        table.setStyle(TableStyle(style_commands))
        elements.append(table)
        doc.build(elements)

    def _stringify_value(self, value: Any) -> Any:
        if isinstance(value, (datetime, date)):
            return value.isoformat()
        if value is None:
            return ""
        if isinstance(value, bool):
            if self._language == "el":
                return "Ναι" if value else "Όχι"
            return "Yes" if value else "No"
        return value
